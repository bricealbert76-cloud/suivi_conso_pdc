import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import csv
import re
import glob
import threading
import traceback
import time
import unicodedata
import calendar
from collections import OrderedDict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

from .constants import *
from .theme import *
from .utils import *
from .window_crapull import CrapullWindow
from .window_diana import DianaWindow
from .window_pdc import PdcMajWindow

class SuiviConsoApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"Suivi Consommation DSI — v{VERSION}")
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)
        self.minsize(920, 700)

        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.file3_path = tk.StringVar()
        self.file4_path = tk.StringVar()
        self.status_var = tk.StringVar(value="Prêt.")

        self._build_ui()
        self._center_window(980, 780)
        # Intervenants périmètre (hors Surveillances des comm.)
        self._allowed_norms      = set()
        self._allowed_list       = []
        self._nucleus_exclusions = set()  # intervenants uniquement Surv.Comm. ou Analytic

        # FIX 2 — Verrou pour éviter les appels concurrents à _preprocess_sources
        self._preprocess_lock = threading.Lock()

        # ── Point d'entrée unique au démarrage : fichier 4 → intervenants → présélections → prétraitement
        self.after(100, self._preselectfile4)

        # Structures pré-chargées pour la fenêtre MAJ PDC
        # (plus chargées au démarrage — uniquement au clic sur le bouton MAJ PDC)
        self.pdc_usernames    = []
        self.pdc_jh_data      = {}
        self.pdc_eur_data     = {}
        self.pdc_projets_rgu  = set()
        self.pdc_eur_detail   = {}
        self.diana_jh_data    = {}
        self.diana_eur_data   = {}
        self.histo_tjm_data   = {}
        self._pdc_cache_ready = False

    def _center_window(self, w, h):
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    # ── Interface ────────────────────────────────────────────────────────
    def _build_ui(self):
        hdr = tk.Frame(self, bg=CA_GREEN, pady=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⬡  SUIVI CONSOMMATION DSI",
                 bg=CA_GREEN, fg="#ffffff", font=FONT_TITLE,
                 padx=24).pack(side="left")
        tk.Label(hdr, text=f"v{VERSION}", bg=CA_GREEN, fg="#ffffff",
                 font=FONT_SMALL).pack(side="left", padx=8, pady=4, anchor="s")
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        body = tk.Frame(self, bg=BG_MAIN)
        body.pack(fill="both", expand=True, padx=24, pady=16)

        left = tk.Frame(body, bg=BG_PANEL, bd=0,
                        highlightbackground=BORDER, highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0, 12), ipadx=14, ipady=14)
        self._build_file_panel(left)

        right = tk.Frame(body, bg=BG_PANEL, bd=0,
                         highlightbackground=BORDER, highlightthickness=1)
        right.pack(side="left", fill="both", expand=True, ipadx=14, ipady=14)
        self._build_result_panel(right)

        bar = tk.Frame(self, bg=BG_CARD, height=30)
        bar.pack(fill="x", side="bottom")
        tk.Label(bar, textvariable=self.status_var, bg=BG_CARD,
                 fg=TEXT_SEC, font=FONT_SMALL, anchor="w", padx=12).pack(fill="x")

    def _build_file_panel(self, parent):
        # ── Panneau défilable : tout le contenu dans un Canvas + Scrollbar ──
        canvas = tk.Canvas(parent, bg=BG_PANEL, bd=0, highlightthickness=0)
        vsb    = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner  = tk.Frame(canvas, bg=BG_PANEL)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(win_id, width=e.width))

        # Molette de souris active seulement quand le curseur survole le panneau
        def _mw(e): canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _mw))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Le contenu ci-dessous est bâti sur `inner` (alias `parent` local)
        parent = inner

        tk.Label(parent, text="SOURCES DE DONNÉES", bg=BG_PANEL,
                 fg=LBL_HEAD, font=FONT_HEAD).pack(anchor="w", padx=10, pady=(10, 14))

        specs = [
            ("1", "Saisies collaborateurs (.txt)",
             "Fichier texte séparé par '~'",
             self.file1_path, self._browse_file1),
            ("2", "Plan de charges — jours/hommes (.xlsx/.csv)",
             "Consommés réels & prévisionnels JH",
             self.file2_path, self._browse_file2),
            ("3", "Plan de charges — euros (.xlsx/.csv)",
             "Consommés réels & prévisionnels €",
             self.file3_path, self._browse_file3),
            ("4", "Intervenants périmètre (.csv)",
             "Référentiel intervenants — filtre hors Surveillances des comm.",
             self.file4_path, self._browse_file4),
        ]
        for num, label, hint, var, cmd in specs:
            self._build_file_row(parent, num, label, hint, var, cmd)

        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=10, pady=14)

        bf = tk.Frame(parent, bg=BG_PANEL)
        bf.pack(padx=10, pady=4, fill="x")

        self.btn_suivi = make_button(bf, "▶  Suivi Conso",
                                     self._run_suivi_conso,
                                     style="ca", width=22)
        self.btn_suivi.pack(fill="x")

        self.btn_tjm = make_button(bf, "📊  Générer Histo TJM",
                                   self._run_histo_tjm,
                                   style="ca", width=22)
        self.btn_tjm.pack(fill="x", pady=(8, 0))

        self.btn_crapull = make_button(bf, "🔍  Vérif. données Crapull",
                                       self._open_crapull_window,
                                       width=22)
        self.btn_crapull.pack(fill="x", pady=(8, 0))
        self.btn_crapull.config(state="disabled")

        self.btn_diana = make_button(bf, "📋  Vérif. données Diana",
                                     self._open_diana_window,
                                     width=22)
        self.btn_diana.pack(fill="x", pady=(8, 0))
        self.btn_diana.config(state="disabled")

        self.btn_pdc_maj = make_button(bf, "📝  MAJ Plan de charges",
                                       self._open_pdc_maj_window,
                                       width=22)
        self.btn_pdc_maj.pack(fill="x", pady=(8, 0))
        self.btn_pdc_maj.config(state="disabled")

        make_button(bf, "✕  Effacer journal",
                    self._clear_log, width=22).pack(fill="x", pady=(8, 0))

    def _build_file_row(self, parent, num, label, hint, var, cmd):
        card = tk.Frame(parent, bg=BG_CARD,
                        highlightbackground=BORDER, highlightthickness=1)
        card.pack(fill="x", padx=10, pady=4, ipady=6, ipadx=8)

        top = tk.Frame(card, bg=BG_CARD)
        top.pack(fill="x")
        tk.Label(top, text=f" {num} ", bg=ACCENT2, fg="#ffffff",
                 font=FONT_SMALL_BOLD).pack(side="left", padx=(0, 6))
        tk.Label(top, text=label, bg=BG_CARD, fg=TEXT_PRI,
                 font=FONT_BODY).pack(side="left")
        tk.Label(card, text=hint, bg=BG_CARD, fg=TEXT_SEC,
                 font=FONT_SMALL).pack(anchor="w", padx=2)

        row = tk.Frame(card, bg=BG_CARD)
        row.pack(fill="x", pady=(4, 0))
        tk.Entry(row, textvariable=var, bg=BG_ENTRY, fg=TEXT_SEC,
                 font=FONT_SMALL, insertbackground=TEXT_PRI, relief="flat", bd=0,
                 highlightbackground=BORDER, highlightthickness=1,
                 width=28).pack(side="left", fill="x", expand=True, ipady=4)
        make_button(row, "…", cmd, width=3).pack(side="left", padx=(6, 0))

    def _build_result_panel(self, parent):
        tk.Label(parent, text="RÉSULTATS & JOURNAL",
                 bg=BG_PANEL, fg=LBL_HEAD, font=FONT_HEAD).pack(
            anchor="w", padx=10, pady=(10, 8))

        sty = ttk.Style()
        sty.theme_use("default")
        sty.configure("DSI.TNotebook", background=BG_PANEL, borderwidth=0)
        sty.configure("DSI.TNotebook.Tab", background=BG_ENTRY,
                      foreground=TEXT_SEC, font=FONT_TAB_UNSEL,
                      padding=[10, 4], borderwidth=0)
        sty.map("DSI.TNotebook.Tab",
                background=[("selected", BG_CARD)],
                foreground=[("selected", "#000000")],
                font=[("selected", FONT_TAB_UNSEL)])

        nb = ttk.Notebook(parent, style="DSI.TNotebook")
        nb.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Journal
        t1 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t1, text=" Journal ")
        self.log_box = scrolledtext.ScrolledText(
            t1, bg=BG_CARD, fg=TEXT_PRI, font=FONT_MONO,
            bd=0, relief="flat", insertbackground=TEXT_PRI,
            state="disabled", wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=4, pady=4)
        self._configure_log_tags()

        # Intervenants périmètre
        t2 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t2, text=" Intervenants périmètre ")
        tk.Label(t2, text="Intervenants retenus (hors Surveillances des comm.)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))
        self.rgu_box = scrolledtext.ScrolledText(
            t2, bg=BG_CARD, fg=ACCENT, font=FONT_MONO,
            bd=0, relief="flat", state="disabled", wrap="word")
        self.rgu_box.pack(fill="both", expand=True, padx=4, pady=4)

        # Codes Projet
        t3 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t3, text=" Codes Projet ciblés ")
        tk.Label(t3, text="Codes projets filtrés (liste fixe)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))
        pb = scrolledtext.ScrolledText(t3, bg=BG_CARD, fg=WARN,
                                        font=FONT_MONO, bd=0, relief="flat")
        pb.pack(fill="both", expand=True, padx=4, pady=4)
        pb.insert("1.0", "\n".join(sorted(TARGET_PROJECTS)))
        pb.config(state="disabled")

        # Anomalies TJM
        t4 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t4, text=" Anomalies TJM ")
        tk.Label(t4,
                 text="Conflits TJM (plusieurs valeurs) + TJM manquants (conso_diana_ttnr non calculable)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))
        self.anom_box = scrolledtext.ScrolledText(
            t4, bg=BG_CARD, fg=WARN, font=FONT_MONO,
            bd=0, relief="flat", state="disabled", wrap="word")
        self.anom_box.pack(fill="both", expand=True, padx=4, pady=4)

        # Incohérence des noms
        t5 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t5, text=" Incohérence des noms ")
        tk.Label(t5,
                 text="Noms différents entre Diana et Histo_TJM (avant normalisation)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))

        inco_frame = tk.Frame(t5, bg=BG_CARD)
        inco_frame.pack(fill="both", expand=True, padx=4, pady=4)
        vsb_i = tk.Scrollbar(inco_frame, orient="vertical")
        hsb_i = tk.Scrollbar(inco_frame, orient="horizontal")
        vsb_i.pack(side="right", fill="y")
        hsb_i.pack(side="bottom", fill="x")
        inco_cols = ("nom_diana", "nom_histo_tjm", "norm_diana", "norm_histo_tjm")
        self.inco_tree = ttk.Treeview(
            inco_frame, columns=inco_cols, show="headings",
            yscrollcommand=vsb_i.set, xscrollcommand=hsb_i.set)
        vsb_i.config(command=self.inco_tree.yview)
        hsb_i.config(command=self.inco_tree.xview)
        headers_inco = {
            "nom_diana":      ("Nom Diana (original)",    200),
            "nom_histo_tjm":  ("Nom Histo_TJM (original)", 200),
            "norm_diana":     ("Nom Diana normalisé",      200),
            "norm_histo_tjm": ("Nom Histo_TJM normalisé",  200),
        }
        style_inco = ttk.Style()
        style_inco.configure("Inco.Treeview",
                              background=BG_CARD, foreground=TEXT_PRI,
                              fieldbackground=BG_CARD, rowheight=22,
                              font=FONT_SMALL)
        style_inco.configure("Inco.Treeview.Heading",
                              background=BG_ENTRY, foreground=ACCENT2,
                              font=FONT_SMALL_BOLD)
        self.inco_tree.config(style="Inco.Treeview")
        for col, (head, width) in headers_inco.items():
            self.inco_tree.heading(col, text=head)
            self.inco_tree.column(col, width=width, anchor="w", stretch=True)
        self.inco_tree.tag_configure("odd",  background=BG_CARD)
        self.inco_tree.tag_configure("even", background=BG_ENTRY)
        self.inco_tree.pack(fill="both", expand=True)

    def _configure_log_tags(self):
        self.log_box.tag_config("info",    foreground=TEXT_PRI)
        self.log_box.tag_config("ok",      foreground=ACCENT)
        self.log_box.tag_config("warn",    foreground=WARN)
        self.log_box.tag_config("error",   foreground="#f85149")
        self.log_box.tag_config("section", foreground=ACCENT2,
                                font=FONT_BODY_BOLD)

    # ── Navigateurs fichiers ──────────────────────────────────────────────
    def _preselectfile1(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        diana_dir  = os.path.join(script_dir, "diana_files")
        pattern    = os.path.join(diana_dir, "YTDTimesheet_*.txt")
        date_re    = re.compile(r"YTDTimesheet_(\d{2}-\d{2}-\d{4})\.txt$", re.IGNORECASE)

        self._log(f"  Recherche fichiers Diana : YTDTimesheet_DD-MM-YYYY.txt"
                  f" dans {diana_dir}", "info")

        if not os.path.isdir(diana_dir):
            msg = f"Répertoire ./diana_files absent ({diana_dir}) — sélection manuelle."
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)
            return

        candidates = glob.glob(pattern)
        self._log(f"  {len(candidates)} fichier(s) YTDTimesheet_*.txt trouvé(s) :", "info")
        for p in sorted(candidates):
            self._log(f"      • {os.path.basename(p)}", "info")

        best_path = None
        best_sort = ""
        for p in candidates:
            m = date_re.search(os.path.basename(p))
            if m:
                dd, mm, yyyy = m.group(1).split("-")
                sort_key = f"{yyyy}-{mm}-{dd}"
                if sort_key > best_sort:
                    best_sort = sort_key
                    best_path = p

        if best_path:
            self.file1_path.set(best_path)
            self._log(f"  ✔  Fichier 1 présélectionné : {os.path.basename(best_path)}", "ok")
            self._set_status(f"Fichier Diana présélectionné : {os.path.basename(best_path)}")
            self.after(0, self._update_diana_btn)
        else:
            msg = ("Aucun fichier YTDTimesheet_DD-MM-YYYY.txt trouvé dans ./diana_files "
                   "— vérifiez le nom des fichiers.")
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)

    def _browse_file1(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier saisies collaborateurs",
            filetypes=[("Fichiers texte", "*.txt"), ("Tous", "*.*")])
        if not path:
            return
        self._set_status("Vérification du séparateur…")
        ok, msg = self._check_separator(path)
        if ok:
            self.file1_path.set(path)
            self._log(f"✔  Fichier 1 : {os.path.basename(path)}", "ok")
            self._log(f"   Séparateur '~' confirmé sur ≤{CHECK_LINES} lignes.", "ok")
            self._update_diana_btn()
        else:
            messagebox.showerror("Contrôle séparateur",
                                 f"Le fichier sélectionné ne semble pas utiliser\n"
                                 f"le caractère '~' comme séparateur.\n\n{msg}")
            self._log(f"✘  Fichier 1 refusé : {os.path.basename(path)}", "error")
            self._log(f"   {msg}", "warn")
        self._set_status("Prêt.")

    def _preselectfiles23(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        pdc_dir    = os.path.join(script_dir, "pdc_files")

        DATE_PAT = r"(\d{4}-\d{2}-\d{2})"

        self._log(f"  Recherche fichiers PDC dans {pdc_dir}", "info")
        self._log(f"  Patterns : SynthesePrevIntervenant_Pere_JH_YYYY-MM-DD.csv/.xlsx", "info")
        self._log(f"           & SynthesePrevIntervenant_Pere_TTNR_YYYY-MM-DD.csv/.xlsx", "info")

        if not os.path.isdir(pdc_dir):
            msg = f"Répertoire ./pdc_files absent ({pdc_dir}) — sélection manuelle."
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)
            return

        def _scan_pairs(ext):
            e = ext.lstrip(".")
            pat_jh   = re.compile(
                rf"SynthesePrevIntervenant_Pere_JH_{DATE_PAT}\.{e}$", re.IGNORECASE)
            pat_ttnr = re.compile(
                rf"SynthesePrevIntervenant_Pere_TTNR_{DATE_PAT}\.{e}$", re.IGNORECASE)
            jh_files, ttnr_files = {}, {}
            for f in glob.glob(os.path.join(pdc_dir, f"*.{e}")):
                bn = os.path.basename(f)
                m = pat_jh.match(bn)
                if m: jh_files[m.group(1)] = f
                m = pat_ttnr.match(bn)
                if m: ttnr_files[m.group(1)] = f
            self._log(f"  Fichiers JH   [{ext.upper()}] : {len(jh_files)} → "
                      f"{[os.path.basename(v) for v in jh_files.values()]}", "info")
            self._log(f"  Fichiers TTNR [{ext.upper()}] : {len(ttnr_files)} → "
                      f"{[os.path.basename(v) for v in ttnr_files.values()]}", "info")
            common = sorted(set(jh_files) & set(ttnr_files), reverse=True)
            if common:
                d = common[0]
                return jh_files[d], ttnr_files[d], d, ext.upper()
            return None, None, None, None

        path_jh, path_eur, date_sel, fmt = _scan_pairs(".csv")
        if not path_jh:
            self._log("  Aucune paire CSV — recherche en XLSX...", "info")
            path_jh, path_eur, date_sel, fmt = _scan_pairs(".xlsx")

        if path_jh and path_eur:
            self.file2_path.set(path_jh)
            self.file3_path.set(path_eur)
            self._log(f"  ✔  Fichier 2 [{fmt}] présélectionné : {os.path.basename(path_jh)}", "ok")
            self._log(f"  ✔  Fichier 3 [{fmt}] présélectionné : {os.path.basename(path_eur)}", "ok")
            self._set_status(f"PDC présélectionnés ({fmt}, date {date_sel}).")
            self.after(0, self._update_crapull_btn)
        else:
            msg = ("Aucune paire PDC (JH + TTNR, même date) trouvée dans ./pdc_files "
                   "— vérifiez les noms de fichiers.")
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)

    def _browse_file2(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier plan de charges JH",
            filetypes=[("Excel ou CSV", "*.xlsx *.csv"),
                       ("Fichiers Excel", "*.xlsx"),
                       ("Fichiers CSV",   "*.csv"),
                       ("Tous", "*.*")])
        if not path:
            return
        other = self.file3_path.get()
        if other and (os.path.splitext(path)[1].lower() !=
                      os.path.splitext(other)[1].lower()):
            messagebox.showerror(
                "Formats incompatibles",
                "Les fichiers 2 et 3 doivent être dans le même format\n"
                "(tous les deux .xlsx  OU  tous les deux .csv).")
            return
        self.file2_path.set(path)
        ext = os.path.splitext(path)[1].upper()
        self._log(f"✔  Fichier 2 [{ext}] : {os.path.basename(path)}", "ok")
        self._update_crapull_btn()

    def _browse_file3(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier plan de charges €",
            filetypes=[("Excel ou CSV", "*.xlsx *.csv"),
                       ("Fichiers Excel", "*.xlsx"),
                       ("Fichiers CSV",   "*.csv"),
                       ("Tous", "*.*")])
        if not path:
            return
        other = self.file2_path.get()
        if other and (os.path.splitext(path)[1].lower() !=
                      os.path.splitext(other)[1].lower()):
            messagebox.showerror(
                "Formats incompatibles",
                "Les fichiers 2 et 3 doivent être dans le même format\n"
                "(tous les deux .xlsx  OU  tous les deux .csv).")
            return
        self.file3_path.set(path)
        ext = os.path.splitext(path)[1].upper()
        self._log(f"✔  Fichier 3 [{ext}] : {os.path.basename(path)}", "ok")
        self._update_crapull_btn()

    def _browse_file4(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier intervenants périmètre",
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous", "*.*")])
        if not path:
            return
        self.file4_path.set(path)
        self._log(f"✔  Fichier 4 [CSV] : {os.path.basename(path)}", "ok")
        self._reload_intervenants()
        # Si les fichiers source sont déjà sélectionnés → lancer le prétraitement
        if self.file1_path.get() and self.file2_path.get() and self.file3_path.get():
            self._preprocess_sources_async()

    def _preselectfile4(self):
        """Chef d'orchestre au démarrage :
        1. Cherche intervenants.csv dans le répertoire du script
        2. Charge _allowed_norms (synchrone — rapide)
        3. Présélectionne les fichiers 1, 2, 3
        4. Lance _preprocess_sources() en arrière-plan (avec spinner)
        Si intervenants.csv est absent → log + attente sélection manuelle.
        """
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidate  = os.path.join(script_dir, "intervenants.csv")
        if not os.path.isfile(candidate):
            self._log(f"  ⚠  intervenants.csv absent du répertoire : {script_dir}", "warn")
            self._log("  ℹ  Le prétraitement sera déclenché après la sélection du fichier 4.", "info")
            return

        # 1 — Charger les intervenants (synchrone, fichier CSV léger)
        self.file4_path.set(candidate)
        self._log(f"  ✔  Fichier 4 présélectionné : intervenants.csv", "ok")
        self._reload_intervenants()   # _allowed_norms garanti chargé ici

        # 2 — Présélectionner les fichiers source 1, 2, 3 (synchrone)
        self._preselectfile1()
        self._preselectfiles23()

        # 3 — Lancer le prétraitement en arrière-plan si les fichiers sont présents
        if self.file1_path.get() and self.file2_path.get() and self.file3_path.get():
            self._preprocess_sources_async()
        else:
            self._log("  ℹ  Prétraitement différé (fichiers source incomplets).", "info")

    def _reload_intervenants(self):
        path = self.file4_path.get()
        self._allowed_norms, self._allowed_list = _load_allowed_intervenants(path)
        self._nucleus_exclusions = _load_nucleus_exclusions(path)
        self._update_rgu_box(self._allowed_list)
        self._log(f"  Intervenants périmètre chargés : {len(self._allowed_list)}", "info")

    # ── Spinner dans le journal ───────────────────────────────────────────
    _SPINNER_FRAMES = ("◐", "◓", "◑", "◒")

    def _spinner_start(self, label):
        """Insère une ligne 'label ◐' dans le journal et la fait tourner sur place.
        Retourne un token (index de ligne) pour l'arrêter via _spinner_stop(token).
        """
        ts   = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] ⏳  {label}  {self._SPINNER_FRAMES[0]}\n"

        self.log_box.config(state="normal")
        self.log_box.insert("end", line, "warn")
        self.log_box.see("end")
        # Mémoriser l'indice de début de cette ligne pour la mettre à jour sur place
        line_index = self.log_box.index("end - 1 line linestart")
        self.log_box.config(state="disabled")

        self._spinner_active = True
        self._spinner_label  = label
        self._spinner_ts     = ts
        self._spinner_frame  = 0
        self._spinner_index  = line_index
        self._spinner_tick()
        return line_index

    def _spinner_tick(self):
        if not getattr(self, "_spinner_active", False):
            return
        self._spinner_frame = (self._spinner_frame + 1) % len(self._SPINNER_FRAMES)
        frame = self._SPINNER_FRAMES[self._spinner_frame]
        new_line = (f"[{self._spinner_ts}] ⏳  {self._spinner_label}  {frame}\n")
        self.log_box.config(state="normal")
        start = self._spinner_index
        end   = f"{start} lineend + 1 char"
        self.log_box.delete(start, end)
        self.log_box.insert(start, new_line, "warn")
        self.log_box.see("end")
        self.log_box.config(state="disabled")
        self._spinner_after_id = self.after(120, self._spinner_tick)

    def _spinner_stop(self, label_done, duration_str=""):
        """Arrête le spinner et remplace la ligne par un message de fin."""
        self._spinner_active = False
        if hasattr(self, "_spinner_after_id"):
            self.after_cancel(self._spinner_after_id)
        suffix = f"  ({duration_str})" if duration_str else ""
        done_line = f"[{self._spinner_ts}] ✔  {label_done}{suffix}\n"
        self.log_box.config(state="normal")
        start = self._spinner_index
        end   = f"{start} lineend + 1 char"
        self.log_box.delete(start, end)
        self.log_box.insert(start, done_line, "ok")
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    # ── Prétraitement asynchrone ──────────────────────────────────────────
    def _disable_all_buttons(self):
        """Désactive tous les boutons d'action pendant le prétraitement."""
        for btn in (self.btn_suivi, self.btn_tjm,
                    self.btn_crapull, self.btn_diana, self.btn_pdc_maj):
            btn.config(state="disabled")

    def _restore_buttons(self):
        """Restaure l'état des boutons selon les fichiers sélectionnés."""
        self.btn_suivi.config(state="normal")
        self.btn_tjm.config(state="normal")
        self._update_diana_btn()
        self._update_crapull_btn()

    def _preprocess_sources_async(self, on_done=None):
        """Lance _preprocess_sources() dans un thread avec spinner animé.
        Désactive tous les boutons pendant l'exécution, les restaure à la fin.
        """
        self._disable_all_buttons()
        self._set_status("Prétraitement des CSV filtrés en cours…")
        self._spinner_start("Génération des CSV filtrés")

        def _run():
            t0 = time.perf_counter()
            try:
                path_diana, path_jh, path_eur = self._preprocess_sources()
                elapsed = time.perf_counter() - t0
                dur_str = (f"{elapsed:.1f}s" if elapsed < 60
                           else f"{int(elapsed)//60}m {int(elapsed)%60:02d}s")
                self.after(0, lambda: self._spinner_stop(
                    "CSV filtrés prêts", dur_str))
                self.after(0, lambda: self._set_status(
                    f"Prétraitement terminé en {dur_str} — prêt."))
                self.after(0, self._restore_buttons)
                if on_done:
                    self.after(0, on_done)
            except Exception as exc:
                self.after(0, lambda: self._spinner_stop("Erreur prétraitement"))
                self.after(0, lambda: self._log(f"  ✘  Prétraitement : {exc}", "error"))
                self.after(0, lambda: self._log(traceback.format_exc(), "error"))
                self.after(0, lambda: self._set_status(f"Erreur prétraitement : {exc}"))
                self.after(0, self._restore_buttons)

        threading.Thread(target=_run, daemon=True).start()

    # ── Contrôle séparateur ───────────────────────────────────────────────
    def _check_separator(self, path):
        try:
            found = checked = 0
            with open(path, "r", encoding="utf-8", errors="replace") as fh:
                for i, line in enumerate(fh):
                    if i >= CHECK_LINES:
                        break
                    if line.strip():
                        checked += 1
                        if SEP_INPUT in line:
                            found += 1
            if checked == 0:
                return False, "Fichier vide ou non lisible."
            if found / checked >= 0.5:
                return True, f"{found}/{checked} lignes contiennent '~'."
            return False, (f"Seulement {found}/{checked} lignes contiennent '~'. "
                           f"Séparateur incorrect ou fichier mal formaté.")
        except Exception as exc:
            return False, str(exc)

    # ══════════════════════════════════════════════════════════════════════
    #  PRÉTRAITEMENT DES FICHIERS SOURCE  (v4.63)
    # ══════════════════════════════════════════════════════════════════════
    def _preprocess_sources(self):
        """Génère les CSV filtrés pour Diana et les 2 PDC si nécessaire,
        puis retourne les chemins effectifs à utiliser pour la suite du traitement.

        Logique par fichier :
          • Si le CSV _avec_filtre existe déjà → l'utiliser directement.
          • Sinon :
              - Diana .txt   → générer CSV sans filtre + CSV _avec_filtre
              - PDC .xlsx    → générer CSV sans filtre + CSV _avec_filtre
              - PDC .csv     → générer CSV _avec_filtre uniquement
                               (le source est déjà un CSV brut)

        Retourne (path_diana_eff, path_jh_eff, path_eur_eff).

        FIX 2 — Protégé par _preprocess_lock pour éviter toute exécution
        concurrente (risque de corruption des CSV filtrés si deux traitements
        sont déclenchés simultanément depuis des threads différents).
        """
        # FIX 2 — Acquisition du verrou : un seul prétraitement à la fois
        with self._preprocess_lock:
            self._log("━" * 54, "section")
            self._log(" PRÉTRAITEMENT — génération CSV filtrés", "section")
            self._log("━" * 54, "section")

            path_diana_eff = self._preprocess_diana()
            path_jh_eff    = self._preprocess_pdc(self.file2_path.get(), label="JH")
            path_eur_eff   = self._preprocess_pdc(self.file3_path.get(), label="EUR")

            self._log("━" * 54, "section")
            self._log(" PRÉTRAITEMENT TERMINÉ", "section")
            self._log(f"  Diana : {os.path.basename(path_diana_eff)}", "ok")
            self._log(f"  JH    : {os.path.basename(path_jh_eff)}", "ok")
            self._log(f"  EUR   : {os.path.basename(path_eur_eff)}", "ok")
            self._log("━" * 54, "section")

            return path_diana_eff, path_jh_eff, path_eur_eff

    # ── Prétraitement Diana ───────────────────────────────────────────────
    def _preprocess_diana(self):
        """Génère (si besoin) les CSV Diana sans filtre et avec filtre.
        Retourne le chemin du CSV _avec_filtre.
        """
        path_src = self.file1_path.get()
        if not path_src or not os.path.isfile(path_src):
            raise FileNotFoundError(f"Fichier Diana introuvable : {path_src}")

        base_no_ext  = os.path.splitext(path_src)[0]
        path_csv_all = base_no_ext + ".csv"
        path_csv_flt = base_no_ext + "_avec_filtre.csv"

        # Si le CSV filtré existe déjà → réutilisation directe
        if os.path.isfile(path_csv_flt):
            self._log(f"  ✔  Diana CSV filtré existant réutilisé : "
                      f"{os.path.basename(path_csv_flt)}", "ok")
            return path_csv_flt

        # Lecture du fichier .txt source
        self._log(f"  ⏳  Lecture Diana source : {os.path.basename(path_src)}", "info")
        t0 = time.perf_counter()

        enc = "utf-8"
        for candidate in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
            try:
                with open(path_src, "r", encoding=candidate) as f:
                    f.read(8192)
                enc = candidate
                break
            except UnicodeDecodeError:
                continue

        with open(path_src, "r", encoding=enc, errors="replace") as fh:
            raw_lines = fh.readlines()

        if not raw_lines:
            raise ValueError("Le fichier Diana source est vide.")

        # En-tête : suppression du 1er caractère s'il n'est pas alphabétique
        header_raw = raw_lines[0].rstrip("\n\r")
        if header_raw and not header_raw[0].isalpha():
            header_raw = header_raw[1:]
        headers_src = [h.strip().strip('"') for h in header_raw.split(SEP_INPUT)]

        def _ci(name):
            """Index d'une colonne (insensible à la casse)."""
            nl = name.lower()
            for i, h in enumerate(headers_src):
                if h.lower() == nl:
                    return i
            return -1

        idx_pc   = _ci(PROJECT_CODE_FIELD)
        idx_uf   = _ci("User First Name")
        idx_ul   = _ci("User Last Name")

        # Indices des colonnes à conserver dans le CSV filtré
        # (DIANA_COLS_FILTRE, insensible à la casse)
        flt_indices = []
        flt_headers = []
        for col in DIANA_COLS_FILTRE:
            i = _ci(col)
            if i >= 0:
                flt_indices.append(i)
                flt_headers.append(headers_src[i])   # nom original du fichier source
            else:
                self._log(f"  ⚠  Diana : colonne '{col}' absente du fichier source — ignorée.", "warn")

        # ── Génération CSV sans filtre (toutes colonnes, toutes lignes) ──
        self._log(f"  ⏳  Génération CSV sans filtre : {os.path.basename(path_csv_all)}", "info")
        nb_all = 0
        with open(path_csv_all, "w", newline="", encoding="utf-8-sig") as fout:
            fout.write(SEP_OUTPUT.join(headers_src) + "\n")
            for raw in raw_lines[1:]:
                line = raw.rstrip("\n\r")
                if not line.strip():
                    continue
                fields = [f.strip().strip('"') for f in line.split(SEP_INPUT)]
                fout.write(SEP_OUTPUT.join(fields) + "\n")
                nb_all += 1
        self._log(f"  ✔  CSV sans filtre : {nb_all} lignes → {os.path.basename(path_csv_all)}", "ok")

        # ── Génération CSV avec filtre ────────────────────────────────────
        self._log(f"  ⏳  Génération CSV avec filtre : {os.path.basename(path_csv_flt)}", "info")
        nb_flt = 0
        with open(path_csv_flt, "w", newline="", encoding="utf-8-sig") as fout:
            fout.write(SEP_OUTPUT.join(flt_headers) + "\n")
            for raw in raw_lines[1:]:
                line = raw.rstrip("\n\r")
                if not line.strip():
                    continue
                fields = [f.strip().strip('"') for f in line.split(SEP_INPUT)]
                while len(fields) < len(headers_src):
                    fields.append("")

                pc_val     = fields[idx_pc]  if idx_pc >= 0  else ""
                u_orig     = (f"{fields[idx_ul]}, {fields[idx_uf]}"
                              if idx_ul >= 0 and idx_uf >= 0 else "")
                match_interv = _norm_name(u_orig) in self._allowed_norms if self._allowed_norms else False
                match_proj   = pc_val in TARGET_PROJECTS

                if not (match_interv or match_proj):
                    continue

                row_out = [fields[i] if i < len(fields) else "" for i in flt_indices]
                fout.write(SEP_OUTPUT.join(row_out) + "\n")
                nb_flt += 1

        self._log_duration("Prétraitement Diana", t0)
        self._log(f"  ✔  CSV filtré : {nb_flt} lignes → {os.path.basename(path_csv_flt)}", "ok")
        return path_csv_flt

    # ── Prétraitement PDC (JH ou EUR) ────────────────────────────────────
    def _preprocess_pdc(self, path_src, label):
        """Génère (si besoin) les CSV PDC sans filtre (si source XLSX) et avec filtre.
        Retourne le chemin du CSV _avec_filtre.
        """
        if not path_src or not os.path.isfile(path_src):
            raise FileNotFoundError(f"Fichier PDC {label} introuvable : {path_src}")

        ext_src      = os.path.splitext(path_src)[1].lower()
        base_no_ext  = os.path.splitext(path_src)[0]
        path_csv_all = base_no_ext + ".csv"          # sans filtre (XLSX seulement)
        path_csv_flt = base_no_ext + "_avec_filtre.csv"

        # Si le CSV filtré existe déjà → réutilisation directe
        if os.path.isfile(path_csv_flt):
            self._log(f"  ✔  PDC {label} CSV filtré existant réutilisé : "
                      f"{os.path.basename(path_csv_flt)}", "ok")
            return path_csv_flt

        if not HAS_PANDAS:
            raise RuntimeError("pandas est requis pour le prétraitement PDC : pip install pandas openpyxl")

        self._log(f"  ⏳  Lecture PDC {label} source : {os.path.basename(path_src)}", "info")
        t0 = time.perf_counter()

        # ── Lecture du fichier source ─────────────────────────────────────
        if ext_src == ".xlsx":
            df = pd.read_excel(path_src, dtype=str, header=3)
            df.columns = [c.strip() for c in df.columns]

            # Génération CSV sans filtre (toutes colonnes, toutes lignes)
            self._log(f"  ⏳  Génération CSV sans filtre PDC {label} : "
                      f"{os.path.basename(path_csv_all)}", "info")
            df.to_csv(path_csv_all, index=False, sep=SEP_OUTPUT, encoding="utf-8-sig")
            self._log(f"  ✔  CSV sans filtre PDC {label} : {len(df)} lignes → "
                      f"{os.path.basename(path_csv_all)}", "ok")

        else:  # .csv
            enc = "utf-8-sig"
            for candidate in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
                try:
                    with open(path_src, "r", encoding=candidate) as f:
                        sample = f.read(8192)
                    enc = candidate
                    break
                except UnicodeDecodeError:
                    continue
            else:
                with open(path_src, "r", encoding="latin-1", errors="replace") as f:
                    sample = f.read(8192)
                enc = "latin-1"

            sep_src = ";" if sample.count(";") >= sample.count(",") else ","

            # Détection position du header (ligne 1 ou 4)
            df_probe = pd.read_csv(path_src, dtype=str, sep=sep_src,
                                   encoding=enc, encoding_errors="replace",
                                   on_bad_lines="skip", nrows=5)
            df_probe.columns = [c.strip() for c in df_probe.columns]
            cols_needed = {COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS}
            skiprows = None if cols_needed.issubset(set(df_probe.columns)) else 3

            df = pd.read_csv(path_src, dtype=str, sep=sep_src,
                             encoding=enc, encoding_errors="replace",
                             skiprows=skiprows, on_bad_lines="skip")
            df.columns = [c.strip() for c in df.columns]
            # Pas de CSV "sans filtre" pour les sources déjà en CSV

        df = df.fillna("")

        # ── Normalisation des noms de colonnes pour le filtre ─────────────
        def _norm_col(s):
            return s.lower().replace(" ", "").replace("_", "")

        col_map = {_norm_col(c): c for c in df.columns}

        def _find(name):
            if name in df.columns:
                return name
            return col_map.get(_norm_col(name))

        interv_col = _find(COL_INTERVENANT)
        proj_col   = _find(COL_PROJET)

        if not interv_col:
            self._log(f"  ⚠  PDC {label} : colonne '{COL_INTERVENANT}' introuvable — "
                      f"filtre intervenant désactivé.", "warn")
        if not proj_col:
            self._log(f"  ⚠  PDC {label} : colonne '{COL_PROJET}' introuvable — "
                      f"filtre projet désactivé.", "warn")

        # ── Colonnes à conserver dans le CSV filtré ───────────────────────
        keep_cols = []
        for col in PDC_COLS_FILTRE_MANDATORY:
            real = _find(col)
            if real:
                keep_cols.append(real)
            else:
                self._log(f"  ⚠  PDC {label} : colonne obligatoire '{col}' absente.", "warn")
        for col in PDC_COLS_FILTRE_OPTIONAL:
            real = _find(col)
            if real and real not in keep_cols:
                keep_cols.append(real)

        # ── Application du filtre ─────────────────────────────────────────
        mask = pd.Series([False] * len(df), index=df.index)
        if interv_col:
            mask |= df[interv_col].apply(
                lambda v: _norm_name(str(v).strip()) in self._allowed_norms
            )
        if proj_col:
            mask |= df[proj_col].astype(str).str.strip().isin(TARGET_PROJECTS)

        df_flt = df[mask][keep_cols]

        # ── Écriture CSV filtré ───────────────────────────────────────────
        df_flt.to_csv(path_csv_flt, index=False, sep=SEP_OUTPUT, encoding="utf-8-sig")
        self._log_duration(f"Prétraitement PDC {label}", t0)
        self._log(f"  ✔  CSV filtré PDC {label} : {len(df_flt)} lignes / "
                  f"{len(df)} total → {os.path.basename(path_csv_flt)}", "ok")

        return path_csv_flt

    # ══════════════════════════════════════════════════════════════════════
    #  TRAITEMENT 1 : Suivi Conso
    # ══════════════════════════════════════════════════════════════════════
    def _run_suivi_conso(self):
        if not self.file1_path.get():
            messagebox.showwarning("Fichier manquant",
                                   "Veuillez sélectionner le fichier n°1 (saisies .txt).")
            return
        p2 = self.file2_path.get()
        p3 = self.file3_path.get()
        if p2 and p3:
            ext2 = os.path.splitext(p2)[1].lower()
            ext3 = os.path.splitext(p3)[1].lower()
            if ext2 != ext3:
                messagebox.showerror(
                    "Formats incompatibles",
                    "Les fichiers 2 et 3 doivent être dans le même format\n"
                    "(tous les deux .xlsx  OU  tous les deux .csv).\n\n"
                    f"Fichier 2 : {ext2.upper()}\n"
                    f"Fichier 3 : {ext3.upper()}")
                return
        self.btn_suivi.config(state="disabled")
        threading.Thread(target=self._process_file1, daemon=True).start()

    def _process_file1(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))

        # Fichiers de sortie
        ts_run    = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_name = f"suivi_conso_nucleus_{ts_run}.xlsx"
        path_xlsx = os.path.join(script_dir, xlsx_name)

        # Fichier de transcodification Stream
        path_transco   = os.path.join(script_dir, "transco_Stream.xlsx")
        # Fichier Histo_TJM
        path_histo_tjm = os.path.join(script_dir, "Histo_TJM.xlsx")

        self._log("━" * 54, "section")
        self._log(" TRAITEMENT SUIVI CONSO — démarrage", "section")
        self._log("━" * 54, "section")
        self._set_status("Prétraitement des fichiers source…")
        t_global = time.perf_counter()

        try:
            if not HAS_OPENPYXL:
                raise RuntimeError("openpyxl est requis. Installez-le : pip install openpyxl")

            # ══════════════════════════════════════════════════════════════
            # ÉTAPE 0 : Prétraitement — génération / réutilisation CSV filtrés
            # ══════════════════════════════════════════════════════════════
            path_in, path_pdc_jh, path_pdc_eur = self._preprocess_sources()

            # Recalcul du nom CSV de sortie basé sur le fichier source original
            path_in_orig = self.file1_path.get()
            base_name    = os.path.splitext(os.path.basename(path_in_orig))[0]
            path_csv     = os.path.join(script_dir, base_name + "_suivi_conso.csv")

            self._log(f"  Source Diana  : {os.path.basename(path_in)}", "info")
            self._log(f"  Source PDC JH : {os.path.basename(path_pdc_jh)}", "info")
            self._log(f"  Source PDC EUR: {os.path.basename(path_pdc_eur)}", "info")
            self._log(f"  Sortie XLSX   : {xlsx_name}", "info")
            self._log(f"  Sortie CSV    : {os.path.basename(path_csv)}", "info")
            self._set_status("Traitement en cours…")

            # ── Chargement de la table de transcodification Stream ──────────
            t_step = time.perf_counter()
            transco_stream = {}
            if os.path.isfile(path_transco):
                wb_tr = openpyxl.load_workbook(path_transco, read_only=True, data_only=True)
                if "Transco_Stream" in wb_tr.sheetnames:
                    ws_tr = wb_tr["Transco_Stream"]
                    for i, row in enumerate(ws_tr.iter_rows(values_only=True)):
                        if i == 0:
                            continue
                        if row[0] is not None:
                            wpn_key    = str(row[0]).strip()
                            task_key   = str(row[1]).strip() if row[1] is not None else ""
                            stream_val = str(row[2]).strip() if row[2] is not None else "NT"
                            transco_stream[(wpn_key, task_key)] = stream_val
                    wb_tr.close()
                    self._log(f"  Transco Stream chargé : {len(transco_stream)} entrée(s).", "ok")
                    self._log_duration("Chargement Transco Stream", t_step)
                else:
                    self._log("  ⚠  Onglet 'Transco_Stream' introuvable dans transco_Stream.xlsx.", "warn")
            else:
                self._log(f"  ⚠  Fichier transco_Stream.xlsx absent ({path_transco}). "
                          f"Stream = 'NT' pour toutes les lignes.", "warn")

            # ── Chargement de Histo_TJM.xlsx ────────────────────────────────
            t_step = time.perf_counter()
            histo_tjm     = {}
            histo_tjm_raw = {}
            if os.path.isfile(path_histo_tjm):
                wb_tjm = openpyxl.load_workbook(path_histo_tjm, read_only=True, data_only=True)
                ws_tjm = wb_tjm.active
                nb_tjm = 0
                for i, row in enumerate(ws_tjm.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    if row[0] is None:
                        continue
                    raw_tjm  = str(row[0]).strip()
                    username = _norm_name(raw_tjm)
                    histo_tjm_raw[username] = raw_tjm
                    for mois_num in range(1, 13):
                        val = row[mois_num] if len(row) > mois_num else None
                        if val is not None:
                            try:
                                histo_tjm[(username, mois_num)] = float(val)
                                nb_tjm += 1
                            except (ValueError, TypeError):
                                pass
                wb_tjm.close()
                self._log(f"  Histo_TJM chargé : {nb_tjm} valeur(s) TJM.", "ok")
                self._log_duration("Chargement Histo_TJM", t_step)
            else:
                self._log(f"  ⚠  Histo_TJM.xlsx absent ({path_histo_tjm}). "
                          f"conso_diana_ttnr ne pourra pas être calculé.", "warn")

            # ── Chargement des plans de charge JH et EUR ────────────────────
            # Désormais on passe les chemins CSV filtrés retournés par _preprocess_sources
            def _load_pdc(path, label):
                """Lit un fichier plan de charge CSV filtré,
                   retourne dict (intervenant, mois_num, project_code)→float.
                   Le CSV filtré a toujours le header en ligne 1 (sep=;, utf-8-sig).
                """
                result = {}
                if not path:
                    self._log(f"  ⚠  Fichier {label} non sélectionné — conso PDC {label} absente.", "warn")
                    return result
                if not HAS_PANDAS:
                    self._log(f"  ⚠  pandas requis pour lire {label}.", "warn")
                    return result
                try:
                    df = pd.read_csv(path, dtype=str, sep=SEP_OUTPUT,
                                     encoding="utf-8-sig", encoding_errors="replace",
                                     on_bad_lines="skip")
                    df.columns = [c.strip() for c in df.columns]

                    for col in [COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS]:
                        if col not in df.columns:
                            self._log(f"  ⚠  Colonne '{col}' absente dans {label}.", "warn")
                            return result
                    df[COL_INTERVENANT] = df[COL_INTERVENANT].fillna("").str.strip()
                    df[COL_MOIS]        = df[COL_MOIS].fillna("").str.strip()
                    df[COL_PROJET]      = df[COL_PROJET].fillna("").str.strip()
                    df[COL_NBJOURS]     = (df[COL_NBJOURS].astype(str)
                                           .str.replace(",", ".", regex=False).str.strip())
                    df = df[df[COL_INTERVENANT] != ""]
                    df[COL_NBJOURS] = pd.to_numeric(df[COL_NBJOURS], errors="coerce").fillna(0)

                    def _norm_mois(val):
                        v = str(val).strip().lower()
                        try:
                            return MOIS_ORDRE.get(int(float(v)))
                        except (ValueError, TypeError):
                            pass
                        return MOIS_ORDRE.get(v)

                    df["mois_num"] = df[COL_MOIS].apply(_norm_mois)
                    df = df.dropna(subset=["mois_num"])
                    df["mois_num"] = df["mois_num"].astype(int)

                    for _, row in df.iterrows():
                        key = (_norm_name(str(row[COL_INTERVENANT]).strip()),
                               int(row["mois_num"]),
                               str(row[COL_PROJET]).strip())
                        result[key] = result.get(key, 0.0) + float(row[COL_NBJOURS])

                    self._log(f"  PDC {label} chargé : {len(result)} combinaison(s) "
                              f"(intervenant, mois, projet).", "ok")
                except Exception as e:
                    self._log(f"  ⚠  Erreur lecture {label} : {e}", "warn")
                return result

            self._log(f"  ⏳  Chargement PDC JH en cours...", "info")
            self._set_status("Chargement PDC JH en cours...")
            t_step = time.perf_counter()
            pdc_jh  = _load_pdc(path_pdc_jh,  "JH")
            self._log_duration("Chargement PDC JH", t_step)
            self._log(f"  ⏳  Chargement PDC EUR en cours...", "info")
            self._set_status("Chargement PDC EUR en cours...")
            t_step = time.perf_counter()
            pdc_eur = _load_pdc(path_pdc_eur, "EUR")
            self._log_duration("Chargement PDC EUR", t_step)

            # ── Lecture du fichier Diana CSV filtré ──────────────────────────
            # Le CSV filtré a sep=; et encodage utf-8-sig — lecture simple
            t_step = time.perf_counter()
            with open(path_in, "r", encoding="utf-8-sig", errors="replace") as fh:
                raw_lines = fh.readlines()

            if not raw_lines:
                raise ValueError("Le fichier Diana CSV filtré est vide.")

            headers = [h.strip().strip('"') for h in raw_lines[0].rstrip("\n\r").split(SEP_OUTPUT)]
            self._log(f"  En-tête Diana CSV filtré : {len(headers)} colonnes détectées.", "info")

            def col_index(name):
                nl = name.lower()
                for i, h in enumerate(headers):
                    if h.lower() == nl:
                        return i
                return -1

            idx_pc = col_index(PROJECT_CODE_FIELD)
            if idx_pc < 0:
                self._log(f"  ⚠  Colonne '{PROJECT_CODE_FIELD}' introuvable — filtre codes projet désactivé.", "warn")
            if not self._allowed_norms:
                self._log("  ⚠  Fichier intervenants non chargé (fichier 4) — filtre périmètre désactivé.", "warn")

            # Index des colonnes sources
            SRC_COLS = [
                "Project Code", "Project Is Alive", "WorkPackage Name", "Task Name",
                "Organization Unit Name", "User First Name", "User Last Name",
                "User Country Name", "User Contract Type", "TimeSheet Mode Name",
                "Timesheet Entry Details Calendar Date", "Status", "Is Published",
                "Timesheet Entry Details Value", "Application Cluster Name",
            ]
            IDX_PROJECT_CODE     = col_index("Project Code")
            IDX_PROJECT_NAME_SRC = col_index("Project Name")
            IDX_WPN              = col_index("WorkPackage Name")
            IDX_TASK             = col_index("Task Name")
            IDX_UFIRST           = col_index("User First Name")
            IDX_ULAST            = col_index("User Last Name")
            IDX_DATE             = col_index("Timesheet Entry Details Calendar Date")

            src_indices = [col_index(c) for c in SRC_COLS]
            missing_src = [SRC_COLS[i] for i, idx in enumerate(src_indices) if idx < 0]
            if missing_src:
                self._log(f"  ⚠  Colonnes sources introuvables : {missing_src}", "warn")

            SRC_DIMS = [c for c in SRC_COLS
                        if c not in ("Timesheet Entry Details Calendar Date",
                                     "Timesheet Entry Details Value",
                                     "User First Name", "User Last Name")]
            OUT_HEADERS = SRC_DIMS + [
                "ProjectName", "UserName", "annee", "dianamonth", "stream",
                "conso_diana_jh", "TJM", "conso_diana_ttnr",
            ]
            NB_MEASURES = 3

            # ── Parcours des lignes Diana CSV filtré ─────────────────────────
            # Le fichier est déjà filtré (intervenant périmètre OU projet ciblé)
            # → on parcourt toutes les lignes sans re-filtrer
            kept_rows  = []
            total_data = 0
            nt_wpn     = set()

            # FIX 1 — Indices constants déplacés AVANT la boucle :
            # SRC_COLS.index() était recalculé à chaque itération (N fois inutilement)
            idx_date_src   = SRC_COLS.index("Timesheet Entry Details Calendar Date")
            idx_value_src  = SRC_COLS.index("Timesheet Entry Details Value")
            idx_ufirst_src = SRC_COLS.index("User First Name")
            idx_ulast_src  = SRC_COLS.index("User Last Name")
            skip_src = {idx_date_src, idx_value_src, idx_ufirst_src, idx_ulast_src}

            for raw in raw_lines[1:]:
                line = raw.rstrip("\n\r")
                if not line.strip():
                    continue
                total_data += 1
                fields = [f.strip().strip('"') for f in line.split(SEP_OUTPUT)]

                while len(fields) < len(headers):
                    fields.append("")

                # fv() reste dans la boucle : capture `fields` qui change à chaque ligne
                def fv(idx, _f=fields):
                    return _f[idx] if 0 <= idx < len(_f) else ""

                out_row = []
                for pos, idx in enumerate(src_indices):
                    if pos in skip_src:
                        continue
                    out_row.append(fv(idx) if idx >= 0 else "")

                proj_code_val     = fv(IDX_PROJECT_CODE)
                proj_name_src     = fv(IDX_PROJECT_NAME_SRC)
                calc_project_name = (f"{proj_code_val} - {proj_name_src}"
                                     if proj_name_src else proj_code_val)

                calc_username     = f"{fv(IDX_ULAST)}, {fv(IDX_UFIRST)}"
                calc_username_key = _norm_name(calc_username)

                date_val   = fv(IDX_DATE)
                calc_annee = ""
                calc_mois  = ""
                if date_val and len(date_val) >= 7:
                    parts = date_val.split("-")
                    if len(parts) >= 2:
                        calc_annee = parts[0]
                        calc_mois  = parts[1]

                calc_dianamonth = (f"{calc_annee} - {calc_mois}"
                                   if calc_annee and calc_mois else "")

                proj_code_for_stream = fv(IDX_PROJECT_CODE)
                if proj_code_for_stream != "P02508":
                    calc_stream = "NA"
                else:
                    wpn_val  = fv(IDX_WPN).strip()
                    task_val = fv(IDX_TASK).strip()
                    calc_stream = transco_stream.get((wpn_val, task_val), "NT")
                    if calc_stream == "NT":
                        nt_wpn.add(f"{wpn_val} | {task_val}")

                raw_value = fv(src_indices[idx_value_src] if src_indices[idx_value_src] >= 0 else -1)
                try:
                    num_value = float(raw_value.replace(",", ".")) if raw_value else 0.0
                except ValueError:
                    num_value = 0.0

                out_row += [calc_project_name, calc_username,
                            calc_annee, calc_dianamonth, calc_stream]

                key_row = out_row[:-5] + [calc_project_name, calc_username_key,
                                          calc_annee, calc_dianamonth, calc_stream]
                dim_key     = tuple(key_row)
                wpn_for_anom = fv(IDX_WPN)
                kept_rows.append((dim_key, num_value, calc_username,
                                   calc_project_name, wpn_for_anom))

            self._log(f"  Lignes lues    : {total_data}", "info")
            self._log(f"  Lignes détail retenues : {len(kept_rows)}", "ok")
            self._log_duration("Lecture & filtrage fichier source", t_step)

            if nt_wpn:
                self._log(f"  ⚠  {len(nt_wpn)} paire(s) (WPN, Task) non transcodée(s) → Stream='NT' :", "warn")
                for w in sorted(nt_wpn):
                    self._log(f"      • {w}", "warn")

            # ── Agrégation ───────────────────────────────────────────────────
            t_step = time.perf_counter()
            agg          = OrderedDict()
            agg_dispname = {}
            agg_proj_wpn = {}
            for dim_key, num_value, username_display, proj_name, wpn in kept_rows:
                if dim_key in agg:
                    agg[dim_key] += num_value
                else:
                    agg[dim_key] = num_value
                    agg_dispname[dim_key] = username_display
                if dim_key not in agg_proj_wpn:
                    agg_proj_wpn[dim_key] = set()
                agg_proj_wpn[dim_key].add((proj_name, wpn))

            IDX_USERNAME_IN_KEY  = len(SRC_DIMS) + 1
            IDX_DIANA_IN_KEY_AGG = len(SRC_DIMS) + 3

            tjm_not_found = {}
            agg_rows = []

            for dim_key, conso_jh in agg.items():
                username_norm = dim_key[IDX_USERNAME_IN_KEY]
                username_disp = agg_dispname[dim_key]
                diana_str     = dim_key[IDX_DIANA_IN_KEY_AGG]
                try:
                    mois_num = int(str(diana_str).split("-")[-1].strip())
                except (ValueError, TypeError, IndexError):
                    mois_num = None

                tjm = histo_tjm.get((username_norm, mois_num)) if mois_num else None
                if tjm is None and mois_num is not None:
                    anom_key     = (username_disp, mois_num)
                    proj_wpn_set = agg_proj_wpn.get(dim_key, set())
                    if anom_key not in tjm_not_found:
                        tjm_not_found[anom_key] = []
                    for pn, wpn in sorted(proj_wpn_set):
                        tjm_not_found[anom_key].append((pn, wpn, conso_jh))

                conso_ttnr = round(conso_jh * tjm, 2) if tjm is not None else ""

                disp_row = list(dim_key)
                disp_row[IDX_USERNAME_IN_KEY] = username_disp
                agg_rows.append(
                    disp_row + [
                        conso_jh,
                        round(tjm, 4) if tjm is not None else "",
                        conso_ttnr,
                    ]
                )

            inco_names = {}
            for dim_key in agg.keys():
                u_norm = dim_key[IDX_USERNAME_IN_KEY]
                u_disp = agg_dispname[dim_key]
                if u_norm in histo_tjm_raw:
                    raw_tjm = histo_tjm_raw[u_norm]
                    if u_disp != raw_tjm and u_norm not in inco_names:
                        inco_names[u_norm] = (u_disp, raw_tjm)
            self._update_inco_box(inco_names)

            self._log(f"  Lignes agrégées (dianamonth) : {len(agg_rows)}", "ok")
            self._log_duration("Agrégation onglet Diana_Task", t_step)
            if tjm_not_found:
                self._log(f"  ⚠  {len(tjm_not_found)} combinaison(s) (UserName, mois) "
                          f"sans TJM dans Histo_TJM → conso_diana_ttnr vide :", "warn")
                for (username, mois_num), details in sorted(tjm_not_found.items()):
                    self._log(f"      • {username} / mois {mois_num:02d}", "warn")
                    for pn, wpn, cjh in details:
                        self._log(f"          ↳ {pn} | {wpn} ({cjh:.4f} jh)", "warn")

            # ── Écriture XLSX ────────────────────────────────────────────────
            t_step = time.perf_counter()
            from openpyxl.utils import get_column_letter

            wb_out = openpyxl.Workbook()

            hdr_font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            hdr_fill_src  = PatternFill("solid", start_color="1F4E79")
            hdr_fill_calc = PatternFill("solid", start_color="375623")
            alt_fill      = PatternFill("solid", start_color="EBF3FB")
            alt_calc      = PatternFill("solid", start_color="EBF7EB")
            thin          = Side(style="thin", color="D0D0D0")
            brd           = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr           = Alignment(horizontal="center", vertical="center")
            lft           = Alignment(horizontal="left",   vertical="center")
            rgt           = Alignment(horizontal="right",  vertical="center")

            col_widths = {
                "Project Code": 14, "Project Is Alive": 16, "WorkPackage Name": 45,
                "Task Name": 35, "Organization Unit Name": 30, "User First Name": 18,
                "User Last Name": 20, "User Country Name": 16, "User Contract Type": 18,
                "TimeSheet Mode Name": 24, "Status": 12, "Is Published": 14,
                "Application Cluster Name": 28,
                "ProjectName": 40, "UserName": 30,
                "annee": 8, "mois": 6, "dianamonth": 14, "stream": 18,
                "conso_diana_jh": 22, "TJM": 14, "conso_diana_ttnr": 22,
            }

            def _write_tcd(ws, data, all_cols, row_label, measure_label):
                from openpyxl.utils import get_column_letter as gcl
                hf  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                hfb = PatternFill("solid", start_color="1F4E79")
                hfc = PatternFill("solid", start_color="375623")
                af  = PatternFill("solid", start_color="EBF3FB")
                ac  = PatternFill("solid", start_color="EBF7EB")
                thn = Side(style="thin", color="D0D0D0")
                b   = Border(left=thn, right=thn, top=thn, bottom=thn)
                ctr = Alignment(horizontal="center", vertical="center")
                lft = Alignment(horizontal="left",   vertical="center")
                rgt = Alignment(horizontal="right",  vertical="center")
                num_fmt = "#,##0.00"

                c = ws.cell(1, 1, f"{row_label} / {measure_label}")
                c.font = hf; c.fill = hfb; c.border = b; c.alignment = ctr
                ws.column_dimensions["A"].width = 42
                for ci, dm in enumerate(all_cols, start=2):
                    c = ws.cell(1, ci, dm)
                    c.font = hf; c.fill = hfc; c.border = b; c.alignment = ctr
                    ws.column_dimensions[gcl(ci)].width = 14
                tot_ci = len(all_cols) + 2
                c = ws.cell(1, tot_ci, "TOTAL")
                c.font = hf; c.fill = hfb; c.border = b; c.alignment = ctr
                ws.column_dimensions[gcl(tot_ci)].width = 14

                row_totals = {dm: 0.0 for dm in all_cols}
                for ri, (row_key, dm_dict) in enumerate(data.items(), start=2):
                    use_alt = (ri % 2 == 0)
                    c = ws.cell(ri, 1, row_key)
                    c.font = Font(name="Arial", size=10); c.border = b; c.alignment = lft
                    if use_alt: c.fill = af
                    line_tot = 0.0
                    for ci, dm in enumerate(all_cols, start=2):
                        val = dm_dict.get(dm, 0.0)
                        cell = ws.cell(ri, ci, round(val, 2) if val else "")
                        cell.font = Font(name="Arial", size=10); cell.border = b
                        cell.alignment = rgt
                        if use_alt: cell.fill = ac
                        if val:
                            cell.number_format = num_fmt
                            line_tot += val
                            row_totals[dm] += val
                    tc = ws.cell(ri, tot_ci, round(line_tot, 2) if line_tot else "")
                    tc.font = Font(name="Arial", bold=True, size=10); tc.border = b
                    tc.alignment = rgt
                    if line_tot: tc.number_format = num_fmt
                    if use_alt: tc.fill = af

                tot_ri = len(data) + 2
                c = ws.cell(tot_ri, 1, "TOTAL")
                c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill = hfb; c.border = b; c.alignment = lft
                grand_tot = 0.0
                for ci, dm in enumerate(all_cols, start=2):
                    val = row_totals[dm]
                    cell = ws.cell(tot_ri, ci, round(val, 2) if val else "")
                    cell.font = Font(name="Arial", bold=True, size=10)
                    cell.fill = hfb; cell.border = b; cell.alignment = rgt
                    if val:
                        cell.number_format = num_fmt
                        grand_tot += val
                gc = ws.cell(tot_ri, tot_ci, round(grand_tot, 2) if grand_tot else "")
                gc.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                gc.fill = hfb; gc.border = b; gc.alignment = rgt
                if grand_tot: gc.number_format = num_fmt
                ws.freeze_panes = "B2"

            def _write_sheet(ws, headers, rows, nb_src_cols, nb_measures):
                idx_first_measure = len(headers) - nb_measures + 1
                for ci, h in enumerate(headers, start=1):
                    c = ws.cell(1, ci, h)
                    c.font      = hdr_font
                    c.fill      = hdr_fill_calc if ci > nb_src_cols else hdr_fill_src
                    c.border    = brd
                    c.alignment = ctr
                    ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 16)
                for ri, row_data in enumerate(rows, start=2):
                    use_alt = (ri % 2 == 0)
                    for ci, val in enumerate(row_data, start=1):
                        is_measure = (ci >= idx_first_measure)
                        if is_measure and val != "":
                            c = ws.cell(ri, ci,
                                        round(val, 4) if isinstance(val, float) else val)
                            c.number_format = "#,##0.0000"
                            c.alignment = rgt
                        else:
                            c = ws.cell(ri, ci, val)
                            c.alignment = lft
                        c.font   = Font(name="Arial", size=10)
                        c.border = brd
                        if use_alt:
                            c.fill = alt_calc if ci > (len(headers) - nb_measures) else alt_fill
                ws.freeze_panes = "A2"

            # Onglet 1 : Diana_Task
            ws1 = wb_out.active
            ws1.title = "Diana_Task"
            _write_sheet(ws1, OUT_HEADERS, agg_rows,
                         nb_src_cols=len(SRC_DIMS), nb_measures=NB_MEASURES)
            self._log(f"  ✔  Onglet 'Diana_Task' : {len(agg_rows)} lignes", "ok")

            # Onglet 2 : Diana_Projet
            EXCL_ONG2 = {"WorkPackage Name", "Task Name", "stream", "Application Cluster Name", "Is Published"}
            keep_idx2 = [i for i, h in enumerate(OUT_HEADERS) if h not in EXCL_ONG2]

            idx_val_oh   = OUT_HEADERS.index("conso_diana_jh")
            idx_ttnr_oh  = OUT_HEADERS.index("conso_diana_ttnr")
            idx_tjm_oh   = OUT_HEADERS.index("TJM")
            IDX_PC_IN_KEY    = OUT_HEADERS.index("Project Code")
            IDX_USER_IN_KEY  = OUT_HEADERS.index("UserName")
            IDX_DIANA_IN_KEY = OUT_HEADERS.index("dianamonth")
            measure_idx_set  = {idx_val_oh, idx_ttnr_oh, idx_tjm_oh}

            agg2_diana = OrderedDict()
            for row in agg_rows:
                dim2 = tuple(row[i] for i in keep_idx2 if i not in measure_idx_set)
                val_jh   = row[idx_val_oh]  if row[idx_val_oh]  != "" else 0.0
                val_ttnr = row[idx_ttnr_oh] if row[idx_ttnr_oh] != "" else 0.0
                tjm_row  = row[idx_tjm_oh]  if row[idx_tjm_oh]  != "" else None
                if dim2 in agg2_diana:
                    agg2_diana[dim2][0] += val_jh
                    agg2_diana[dim2][1] += val_ttnr
                    if tjm_row is not None:
                        agg2_diana[dim2][2] = max(agg2_diana[dim2][2] or 0, tjm_row)
                else:
                    agg2_diana[dim2] = [val_jh, val_ttnr, tjm_row]

            keep_idx2_dims = [i for i in keep_idx2 if i not in measure_idx_set]
            pos_pc    = keep_idx2_dims.index(IDX_PC_IN_KEY)
            pos_user  = keep_idx2_dims.index(IDX_USER_IN_KEY)
            pos_diana = keep_idx2_dims.index(IDX_DIANA_IN_KEY)

            def _fmt(v):
                if v is None or v == 0.0:
                    return ""
                return round(v, 4) if isinstance(v, float) else v

            agg_rows2 = []
            for dim2, (val_jh, val_ttnr, tjm_val) in agg2_diana.items():
                username_disp2 = dim2[pos_user]
                username_key2  = _norm_name(username_disp2)
                proj_code_r    = dim2[pos_pc]
                diana_val      = dim2[pos_diana]
                try:
                    mois_num_r = int(str(diana_val).split("-")[-1].strip())
                except (ValueError, TypeError, IndexError):
                    mois_num_r = None
                pdc_key2    = (username_key2, mois_num_r, proj_code_r) if mois_num_r else None
                val_pdc_jh  = pdc_jh.get(pdc_key2,  0.0) if pdc_key2 else 0.0
                val_pdc_eur = pdc_eur.get(pdc_key2, 0.0) if pdc_key2 else 0.0

                agg_rows2.append(
                    list(dim2) + [
                        _fmt(val_jh), _fmt(val_ttnr),
                        _fmt(val_pdc_jh), _fmt(val_pdc_eur),
                    ]
                )

            NB_MEAS2   = 5
            HDR2_ALL   = [OUT_HEADERS[i] for i in keep_idx2]
            HDR2_DIMS  = [h for h in HDR2_ALL if h not in ("conso_diana_jh",
                                                             "conso_diana_ttnr", "TJM")]
            HDR2_FINAL = HDR2_DIMS + [
                "conso_diana_jh", "conso_diana_ttnr",
                "conso_pdc_jh", "conso_pdc_eur", "ecart_jh",
            ]

            agg_rows2 = [
                row + [_fmt(
                    (row[-2] if row[-2] != "" else 0.0) - (row[-4] if row[-4] != "" else 0.0)
                )]
                for row in agg_rows2
            ]

            col_widths.update({"conso_pdc_jh": 18, "conso_pdc_eur": 18, "ecart_jh": 14})

            ws2 = wb_out.create_sheet("Diana_Projet")
            _write_sheet(ws2, HDR2_FINAL, agg_rows2,
                         nb_src_cols=len([h for h in HDR2_DIMS if h in SRC_DIMS]),
                         nb_measures=NB_MEAS2)
            self._log(f"  ✔  Onglet 'Diana_Projet' : {len(agg_rows2)} lignes", "ok")

            saved_path = path_xlsx
            wb_out.save(saved_path)
            self._log(f"  ✔  {xlsx_name} enregistré.", "ok")
            self._log_duration("Écriture fichier nucleus", t_step)

            # TCD natifs via win32com
            self._log("  ⏳  Création des TCD natifs Excel en cours...", "info")
            self._set_status("Création des TCD Excel...")
            t_step = time.perf_counter()
            tcd_ok = self._create_pivot_tables(saved_path, HDR2_FINAL, OUT_HEADERS)
            if tcd_ok:
                self._log_duration("Création TCD natifs", t_step)
            else:
                self._log("  ⚠  TCD natifs non créés (win32com indisponible). "
                          "Installez pywin32 : pip install pywin32", "warn")

            # CSV brut complet (toutes lignes du fichier Diana original)
            # On relit le fichier .txt original pour ce CSV exhaustif
            t_step = time.perf_counter()
            path_in_orig = self.file1_path.get()
            with open(path_in_orig, "r", encoding="utf-8", errors="replace") as fh:
                raw_lines_orig = fh.readlines()
            header_raw_orig = raw_lines_orig[0].rstrip("\n\r")
            if header_raw_orig and not header_raw_orig[0].isalpha():
                header_raw_orig = header_raw_orig[1:]
            headers_orig = [h.strip().strip('"') for h in header_raw_orig.split(SEP_INPUT)]
            with open(path_csv, "w", newline="", encoding="utf-8-sig") as fh:
                fh.write(SEP_OUTPUT.join(headers_orig) + "\n")
                nb_csv = 0
                for raw in raw_lines_orig[1:]:
                    line = raw.rstrip("\n\r")
                    if not line.strip():
                        continue
                    fields_all = [f.strip().strip('"') for f in line.split(SEP_INPUT)]
                    fh.write(SEP_OUTPUT.join(fields_all) + "\n")
                    nb_csv += 1
            self._log(f"  ✔  {os.path.basename(path_csv)} créé ({nb_csv} lignes)", "ok")
            self._log_duration("Écriture CSV brut complet", t_step)

            self._log(f"  Intervenants périmètre chargés : {len(self._allowed_list)}", "info")

            self._log_duration("Traitement total Suivi Conso", t_global, tag="section")
            self._log("━" * 54, "section")
            self._log(" TRAITEMENT TERMINÉ AVEC SUCCÈS ✔", "ok")
            self._log("━" * 54, "section")
            self._set_status(
                f"Terminé — {len(agg_rows)} lignes (Task) | {len(agg_rows2)} lignes (Projet) | "
                f"{total_data} lignes CSV")
            self.after(100, lambda: messagebox.showinfo(
                "Traitement terminé",
                f"Fichiers générés avec succès !\n\n"
                f"XLSX : {xlsx_name}\n"
                f"  Onglet 'Diana_Task'   : {len(agg_rows)} lignes\n"
                f"  Onglet 'Diana_Projet' : {len(agg_rows2)} lignes\n"
                f"  {len(nt_wpn)} paire(s) (WPN, Task) non transcodée(s)\n\n"
                f"CSV brut : {os.path.basename(path_csv)}\n"
                f"  → {total_data} lignes\n\n"
                f"Répertoire : {script_dir}"))

        except Exception as exc:
            self._log(f"✘ ERREUR : {exc}", "error")
            self._log(traceback.format_exc(), "error")
            self._set_status(f"Erreur : {exc}")
            self.after(100, lambda: messagebox.showerror(
                "Erreur de traitement", str(exc)))
        finally:
            self.after(100, lambda: self.btn_suivi.config(state="normal"))

    # ── TRAITEMENT 2 : Histo TJM ──────────────────────────────────────────
    def _run_histo_tjm(self):
        if not self.file2_path.get() or not self.file3_path.get():
            messagebox.showwarning(
                "Fichiers manquants",
                "Veuillez sélectionner les fichiers n°2 (JH) et n°3 (€).")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("Dépendance manquante",
                                  "Installez openpyxl :  pip install openpyxl")
            return
        if not HAS_PANDAS:
            messagebox.showerror("Dépendance manquante",
                                  "Installez pandas :  pip install pandas openpyxl")
            return
        self.btn_tjm.config(state="disabled")
        threading.Thread(target=self._process_histo_tjm, daemon=True).start()

    def _process_histo_tjm(self):
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()

        script_dir = os.path.dirname(os.path.abspath(__file__))
        path_out   = os.path.join(script_dir, "Histo_TJM.xlsx")

        t_global_tjm = time.perf_counter()
        self._log("━" * 54, "section")
        self._log(" GÉNÉRATION HISTO_TJM — démarrage", "section")
        self._log("━" * 54, "section")
        self._log(f"  Fichier JH  : {os.path.basename(path_jh)}", "info")
        self._log(f"  Fichier EUR : {os.path.basename(path_eur)}", "info")
        self._log(f"  Sortie      : {path_out}", "info")
        self._set_status("Calcul des TJM en cours…")

        try:
            t_step = time.perf_counter()
            df_jh  = pd.read_excel(path_jh,  dtype=str, header=3)
            self._log_duration("Lecture fichier JH", t_step)
            t_step = time.perf_counter()
            df_eur = pd.read_excel(path_eur, dtype=str, header=3)
            self._log_duration("Lecture fichier EUR", t_step)
            df_jh.columns  = [c.strip() for c in df_jh.columns]
            df_eur.columns = [c.strip() for c in df_eur.columns]

            self._log(f"  JH  : {len(df_jh)} lignes | "
                      f"colonnes : {list(df_jh.columns)}", "info")
            self._log(f"  EUR : {len(df_eur)} lignes | "
                      f"colonnes : {list(df_eur.columns)}", "info")

            required = [COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS]
            for label, df in [("JH (fichier 2)", df_jh), ("EUR (fichier 3)", df_eur)]:
                missing = [c for c in required if c not in df.columns]
                if missing:
                    raise ValueError(
                        f"Fichier {label} — colonnes manquantes : {missing}\n"
                        f"Colonnes présentes : {list(df.columns)}")

            def clean_df(df):
                df = df[[COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS]].copy()
                df[COL_INTERVENANT] = df[COL_INTERVENANT].fillna("").str.strip()
                df[COL_MOIS]        = df[COL_MOIS].fillna("").str.strip()
                df[COL_PROJET]      = df[COL_PROJET].fillna("").str.strip()
                df[COL_NBJOURS]     = (df[COL_NBJOURS].astype(str)
                                       .str.replace(",", ".", regex=False)
                                       .str.strip())
                df = df[df[COL_INTERVENANT] != ""]
                df = df[df[COL_NBJOURS]     != ""]
                df[COL_NBJOURS] = pd.to_numeric(df[COL_NBJOURS], errors="coerce")
                return df.dropna(subset=[COL_NBJOURS])

            df_jh  = clean_df(df_jh)
            df_eur = clean_df(df_eur)

            df_jh[COL_INTERVENANT]  = df_jh[COL_INTERVENANT].apply(_norm_name)
            df_eur[COL_INTERVENANT] = df_eur[COL_INTERVENANT].apply(_norm_name)
            self._log("  Intervenants normalisés dans JH et EUR.", "info")

            def normalize_mois(val):
                v = str(val).strip().lower()
                try:
                    return MOIS_ORDRE.get(int(float(v)))
                except (ValueError, TypeError):
                    pass
                return MOIS_ORDRE.get(v)

            df_jh["mois_num"]  = df_jh[COL_MOIS].apply(normalize_mois)
            df_eur["mois_num"] = df_eur[COL_MOIS].apply(normalize_mois)

            for label, df in [("JH", df_jh), ("EUR", df_eur)]:
                n = df["mois_num"].isna().sum()
                if n:
                    self._log(f"  ⚠  {n} ligne(s) {label} avec mois "
                              f"non reconnu — ignorées.", "warn")

            df_jh  = df_jh.dropna(subset=["mois_num"])
            df_eur = df_eur.dropna(subset=["mois_num"])
            df_jh["mois_num"]  = df_jh["mois_num"].astype(int)
            df_eur["mois_num"] = df_eur["mois_num"].astype(int)

            key = [COL_INTERVENANT, COL_PROJET, "mois_num"]
            agg_jh  = (df_jh.groupby(key, as_index=False)[COL_NBJOURS]
                       .sum().rename(columns={COL_NBJOURS: "jh"}))
            agg_eur = (df_eur.groupby(key, as_index=False)[COL_NBJOURS]
                       .sum().rename(columns={COL_NBJOURS: "eur"}))

            merged = pd.merge(agg_jh, agg_eur, on=key, how="inner")
            merged = merged[merged["jh"] > 0].copy()
            merged["tjm"] = merged["eur"] / merged["jh"]

            self._log(f"  Couples (intervenant, projet, mois) appariés : "
                      f"{len(merged)}", "info")

            key2      = [COL_INTERVENANT, "mois_num"]
            anomalies = []
            tjm_final = {}

            for (interv, mois), grp in merged.groupby(key2):
                tjms    = grp["tjm"].tolist()
                projets = grp[COL_PROJET].tolist()
                tjm_max = max(tjms)
                tjm_final[(interv, mois)] = tjm_max

                if max(tjms) - min(tjms) > 0.01:
                    detail = ", ".join(
                        f"{p}={t:.2f}€" for p, t in zip(projets, tjms))
                    anomalies.append((interv, mois, tjm_max, detail))
                    self._log(
                        f"  ⚠  CONFLIT TJM | Intervenant : {interv} | "
                        f"Mois : {mois:02d} | "
                        f"TJM/projet : {detail} "
                        f"→ retenu (max) : {tjm_max:.2f} €",
                        "warn")

            self._log(f"  Anomalies TJM : {len(anomalies)}",
                      "warn" if anomalies else "ok")
            self._update_anom_box(anomalies, {})

            intervenants = sorted(merged[COL_INTERVENANT].unique())
            self._log(f"  Intervenants distincts : {len(intervenants)}", "info")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Histo_TJM"

            hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            hdr_fill   = PatternFill("solid", start_color="1F4E79")
            center_aln = Alignment(horizontal="center", vertical="center")
            left_aln   = Alignment(horizontal="left",   vertical="center")
            num_fmt    = '#,##0.00 "€"'
            thin       = Side(style="thin", color="D0D0D0")
            brd        = Border(left=thin, right=thin, top=thin, bottom=thin)
            alt_fill   = PatternFill("solid", start_color="EBF3FB")
            data_font  = Font(name="Arial", size=10)

            c = ws.cell(1, 1, "Intervenant")
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = left_aln; c.border = brd
            ws.column_dimensions["A"].width = 34

            col_letters = list("BCDEFGHIJKLM")
            for idx, (letter, label) in enumerate(
                    zip(col_letters, MOIS_LABELS), start=1):
                c = ws.cell(1, idx + 1, label)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = center_aln; c.border = brd
                ws.column_dimensions[letter].width = 14

            for row_idx, interv in enumerate(intervenants, start=2):
                use_alt  = (row_idx % 2 == 0)
                row_fill = alt_fill if use_alt else None

                c = ws.cell(row_idx, 1, interv)
                c.font = data_font; c.alignment = left_aln; c.border = brd
                if row_fill:
                    c.fill = row_fill

                for m in range(1, 13):
                    tjm_val = tjm_final.get((interv, m))
                    cell    = ws.cell(row_idx, m + 1,
                                      round(tjm_val, 2) if tjm_val is not None else "")
                    cell.font      = data_font
                    cell.alignment = center_aln
                    cell.border    = brd
                    if row_fill:
                        cell.fill = row_fill
                    if tjm_val is not None:
                        cell.number_format = num_fmt

            ws.freeze_panes = "B2"
            wb.save(path_out)

            self._log(f"  ✔  Fichier créé → {path_out}", "ok")
            self._log_duration("Génération Histo_TJM totale", t_global_tjm, tag="section")
            self._log("━" * 54, "section")
            self._log(" HISTO_TJM GÉNÉRÉ AVEC SUCCÈS ✔", "ok")
            self._log("━" * 54, "section")
            self._set_status(
                f"Terminé — {len(intervenants)} intervenants | "
                f"{len(anomalies)} anomalie(s) → Histo_TJM.xlsx")
            self.after(100, lambda: messagebox.showinfo(
                "Histo TJM généré",
                f"Fichier Histo_TJM.xlsx créé avec succès !\n\n"
                f"  • {len(intervenants)} intervenant(s)\n"
                f"  • {len(anomalies)} anomalie(s) TJM détectée(s)\n\n"
                f"Emplacement : {path_out}"))

        except Exception as exc:
            self._log(f"✘ ERREUR : {exc}", "error")
            self._log(traceback.format_exc(), "error")
            self._set_status(f"Erreur : {exc}")
            self.after(100, lambda: messagebox.showerror(
                "Erreur de traitement", str(exc)))
        finally:
            self.after(100, lambda: self.btn_tjm.config(state="normal"))

    # ── Création TCD natifs via win32com ──────────────────────────────────
    def _create_pivot_tables(self, path_xlsx, hdr_projet, hdr_task):
        try:
            import win32com.client as win32
            import pythoncom
        except ImportError:
            return False

        pythoncom.CoInitialize()
        xl = None
        wb = None
        try:
            xl = win32.DispatchEx("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.AutomationSecurity = 1
            try:
                xl.Application.FileDialog(1)
            except Exception:
                pass

            wb = xl.Workbooks.Open(
                path_xlsx,
                UpdateLinks=False,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True,
                Notify=False)

            xlDatabase    = 1
            xlDataField   = 4
            xlRowField    = 1
            xlColumnField = 2
            xlPageField   = 3
            xlSum         = -4157

            def _make_pivot(wb, src_sheet_name, src_headers,
                            dest_sheet_name,
                            row_field, col_field, data_field,
                            filter_field=None, filter_value=None):
                ws_src    = wb.Sheets(src_sheet_name)
                last_row  = ws_src.UsedRange.Rows.Count
                last_col  = ws_src.UsedRange.Columns.Count
                src_range = ws_src.Range(
                    ws_src.Cells(1, 1),
                    ws_src.Cells(last_row, last_col))

                try:
                    ws_dest = wb.Sheets(dest_sheet_name)
                    ws_dest.Cells.Clear()
                except Exception:
                    ws_dest = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    ws_dest.Name = dest_sheet_name

                pc = wb.PivotCaches().Create(
                    SourceType=xlDatabase,
                    SourceData=src_range)

                pt = pc.CreatePivotTable(
                    TableDestination=ws_dest.Cells(3, 1),
                    TableName=dest_sheet_name)

                pt.CompactLayoutRowHeader = row_field

                if filter_field:
                    pf = pt.PivotFields(filter_field)
                    pf.Orientation = xlPageField
                    pf.Position = 1
                    pf.CurrentPage = filter_value

                pf_row = pt.PivotFields(row_field)
                pf_row.Orientation = xlRowField
                pf_row.Position = 1

                pf_col = pt.PivotFields(col_field)
                pf_col.Orientation = xlColumnField
                pf_col.Position = 1

                pf_data = pt.PivotFields(data_field)
                pf_data.Orientation = xlDataField
                pf_data.Function = xlSum
                pf_data.NumberFormat = "# ##0"
                pf_data.Name = f"Somme de {data_field}"

                pt.TableStyle2 = "PivotStyleMedium2"
                ws_dest.Columns.AutoFit()
                self._log(f"  ✔  TCD '{dest_sheet_name}' créé.", "ok")

            self._log("  ℹ  Création TCD 1 (TCD_Projets)...", "info")
            _make_pivot(wb,
                        src_sheet_name  = "Diana_Projet",
                        src_headers     = hdr_projet,
                        dest_sheet_name = "TCD_Projets",
                        row_field       = "ProjectName",
                        col_field       = "dianamonth",
                        data_field      = "conso_diana_ttnr")

            self._log("  ℹ  Création TCD 2 (TCD_P02508)...", "info")
            _make_pivot(wb,
                        src_sheet_name  = "Diana_Task",
                        src_headers     = hdr_task,
                        dest_sheet_name = "TCD_P02508",
                        row_field       = "stream",
                        col_field       = "dianamonth",
                        data_field      = "conso_diana_ttnr",
                        filter_field    = "Project Code",
                        filter_value    = "P02508")

            self._log("  ℹ  Sauvegarde du fichier avec TCD...", "info")
            xl.DisplayAlerts = False
            xl.EnableEvents  = False
            for addin in xl.COMAddIns:
                try:
                    if "label" in addin.Description.lower() or \
                       "confidential" in addin.Description.lower() or \
                       "sensitivity" in addin.ProgId.lower() or \
                       "aip" in addin.ProgId.lower():
                        addin.Connect = False
                except Exception:
                    pass
            wb.Save()
            self._log("  ℹ  wb.Save() exécuté.", "info")
            xl.EnableEvents  = True
            xl.DisplayAlerts = False
            wb.Close(SaveChanges=True)
            self._log("  ℹ  wb.Close(SaveChanges=True) exécuté.", "info")
            return True

        except Exception as exc:
            self._log(f"  ✘  Erreur création TCD : {exc}", "error")
            self._log(traceback.format_exc(), "error")
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            return False
        finally:
            try:
                if xl is not None:
                    xl.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()

    # ── Bouton MAJ PDC ────────────────────────────────────────────────────
    def _open_pdc_maj_window(self):
        """Ouvre la fenêtre de mise à jour du plan de charges.
        Déclenche le chargement du cache PDC si pas encore prêt (v4.63 : au clic, plus au démarrage).
        """
        if not HAS_PANDAS:
            messagebox.showerror("Dépendance manquante",
                                  "pandas est requis : pip install pandas openpyxl")
            return
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()
        if not path_jh or not path_eur:
            messagebox.showwarning("Fichiers manquants",
                                    "Sélectionnez d'abord les fichiers PDC (2 et 3).")
            return

        if not self._pdc_cache_ready:
            # Lancer le chargement du cache en arrière-plan,
            # puis ouvrir la fenêtre une fois le cache prêt
            self._set_status("Chargement du cache PDC en cours…")
            self._preload_pdc_cache(on_done=self._open_pdc_maj_window_now)
        else:
            self._open_pdc_maj_window_now()

    def _open_pdc_maj_window_now(self):
        """Ouvre effectivement la fenêtre PdcMajWindow (cache disponible)."""
        self._log("  ▶  Ouverture fenêtre MAJ Plan de charges…", "info")
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()
        try:
            win = PdcMajWindow(self, path_jh, path_eur,
                               cache_usernames=self.pdc_usernames,
                               cache_jh=self.pdc_jh_data,
                               cache_eur=self.pdc_eur_data,
                               cache_eur_detail=self.pdc_eur_detail,
                               diana_jh=self.diana_jh_data,
                               diana_eur=self.diana_eur_data,
                               histo_tjm=self.histo_tjm_data,
                               cache_ready=self._pdc_cache_ready,
                               nucleus_exclusions=self._nucleus_exclusions)
            win.lift()
            win.focus_force()
        except Exception as e:
            self._log(f"  ✘  Erreur ouverture MAJ PDC : {e}\n{traceback.format_exc()}", "error")

    def _update_diana_btn(self):
        has_diana = bool(self.file1_path.get())
        self.btn_diana.config(state="normal" if has_diana else "disabled")

    def _open_diana_window(self):
        path = self.file1_path.get()
        if not path:
            messagebox.showwarning("Fichier manquant",
                                    "Sélectionnez d'abord le fichier Diana (fichier 1).")
            return
        flt = os.path.splitext(path)[0] + "_avec_filtre.csv"
        win = DianaWindow(self, flt if os.path.isfile(flt) else path)
        win.focus_set()

    # ── Pré-chargement cache PDC ──────────────────────────────────────────
    def _preload_pdc_cache(self, on_done=None):
        """Pré-charge les structures PDC.
        v4.63 : déclenché au clic sur MAJ PDC (plus au démarrage).
        Utilise les CSV filtrés si disponibles via _preprocess_sources().
        on_done : callback appelé dans le thread principal une fois le cache prêt.
        """
        path1 = self.file1_path.get()
        path2 = self.file2_path.get()
        path3 = self.file3_path.get()
        if not (path1 and path2 and path3):
            self._log("  ℹ  Pré-chargement PDC ignoré (fichiers incomplets).", "info")
            return

        self._log("  ⏳  Pré-chargement cache PDC (MAJ Plan de charges)…", "info")
        self._set_status("Pré-chargement cache PDC…")

        def _run():
            try:
                annee = datetime.now().year

                # ── Résolution des chemins CSV filtrés ────────────────────────
                # On réutilise _preprocess_sources pour obtenir les bons chemins
                # (génère les CSV s'ils n'existent pas, ou réutilise s'ils existent)
                try:
                    path_diana_eff, path_jh_eff, path_eur_eff = self._preprocess_sources()
                except Exception as e_pre:
                    self.after(0, lambda: self._log(
                        f"  ⚠  Prétraitement PDC cache : {e_pre} — utilisation des chemins bruts.", "warn"))
                    path_diana_eff = path1
                    path_jh_eff    = path2
                    path_eur_eff   = path3

                # ── Intervenants périmètre ─────────────────────────────────────
                rgu_users      = {_norm_name(n): n for n in self._allowed_list}
                # usernames_list complété plus bas après lecture du PDC JH (pdc_only)

                # ── Lecture PDC CSV filtré ────────────────────────────────────
                def _read_pdc_flt(path):
                    """Lit un CSV filtré (sep=;, utf-8-sig, header ligne 1)."""
                    df = pd.read_csv(path, dtype=str, sep=SEP_OUTPUT,
                                     encoding="utf-8-sig", encoding_errors="replace",
                                     on_bad_lines="skip")
                    df.columns = [c.strip() for c in df.columns]
                    return df.fillna("")

                def _norm_col(s):
                    return s.lower().replace(" ", "").replace("_", "")

                df_jh  = _read_pdc_flt(path_jh_eff)
                df_eur = _read_pdc_flt(path_eur_eff)

                def _find(df, name):
                    if name in df.columns: return name
                    k = _norm_col(name)
                    return next((c for c in df.columns if _norm_col(c) == k), None)

                ic_jh  = _find(df_jh,  COL_INTERVENANT)
                mc_jh  = _find(df_jh,  COL_MOIS)
                pc_jh  = _find(df_jh,  COL_PROJET)
                nc_jh  = _find(df_jh,  COL_NBJOURS)
                yc_jh  = next((c for c in df_jh.columns  if "year" in c.lower()), None)
                ic_eur = _find(df_eur, COL_INTERVENANT)
                mc_eur = _find(df_eur, COL_MOIS)
                pc_eur = _find(df_eur, COL_PROJET)
                nc_eur = _find(df_eur, COL_NBJOURS)
                yc_eur = next((c for c in df_eur.columns if "year" in c.lower()), None)

                def _norm_c(s): return s.lower().replace(" ","").replace("_","")
                lc_jh = next(
                    (c for c in df_jh.columns
                     if _norm_c(c) in ("prjlibelleprojet","libelleprojet","libprojet")
                     or ("libelle" in c.lower() and "projet" in c.lower())
                     or ("libelle" in c.lower() and "prj" in c.lower())), None)
                if not lc_jh:
                    lc_jh = next(
                        (c for c in df_jh.columns if "libelle" in c.lower()), None)

                def _to_mois(v):
                    v2 = str(v).strip().lower()
                    r = MOIS_ORDRE.get(v2)
                    if r: return r
                    try: return int(float(v2))
                    except: return None

                if yc_jh:
                    df_jh = df_jh[df_jh[yc_jh].astype(str).str.strip() == str(annee)]
                if yc_eur:
                    df_eur = df_eur[df_eur[yc_eur].astype(str).str.strip() == str(annee)]

                # ── Dict JH ───────────────────────────────────────────────────
                jh_data  = {}
                pdc_only = {}
                for _, row in df_jh.iterrows():
                    interv_raw  = str(row.get(ic_jh, "")).strip()
                    interv_norm = _norm_name(interv_raw)
                    prj         = str(row.get(pc_jh, "")).strip()
                    in_csv      = interv_norm in rgu_users
                    target_proj = prj in TARGET_PROJECTS
                    if not in_csv:
                        # Capturer TOUS les non-intervenants.csv pour la combobox
                        if interv_raw and interv_norm:
                            pdc_only[interv_norm] = interv_raw
                        if not target_proj:
                            continue
                    mois = _to_mois(row.get(mc_jh, ""))
                    jh   = 0.0
                    try: jh = float(str(row.get(nc_jh, "0")).replace(",", "."))
                    except: pass
                    if not prj or mois is None: continue
                    if interv_norm not in jh_data:
                        jh_data[interv_norm] = {}
                    if prj not in jh_data[interv_norm]:
                        lib = str(row.get(lc_jh, "")) if lc_jh else ""
                        jh_data[interv_norm][prj] = {"_lib": lib}
                    d = jh_data[interv_norm][prj]
                    d[int(mois)] = d.get(int(mois), 0.0) + jh

                # Compléter libellés depuis EUR
                lc_eur = next(
                    (c for c in df_eur.columns
                     if _norm_c(c) in ("prjlibelleprojet","libelleprojet","libprojet")
                     or ("libelle" in c.lower() and "projet" in c.lower())
                     or ("libelle" in c.lower() and "prj" in c.lower())), None)
                if not lc_eur:
                    lc_eur = next(
                        (c for c in df_eur.columns if "libelle" in c.lower()), None)

                if lc_eur:
                    lib_from_eur = {}
                    for _, row in df_eur.iterrows():
                        prj2 = str(row.get(pc_eur, "")).strip()
                        lib2 = str(row.get(lc_eur, "")).strip()
                        if prj2 and lib2 and prj2 not in lib_from_eur:
                            lib_from_eur[prj2] = lib2
                    for u_norm, prj_dict in jh_data.items():
                        for prj2, prj_info in prj_dict.items():
                            if not prj_info.get("_lib") and prj2 in lib_from_eur:
                                prj_info["_lib"] = lib_from_eur[prj2]

                projets_rgu = set()
                for prj_dict in jh_data.values():
                    projets_rgu.update(prj_dict.keys())

                csv_only = sorted(
                    rgu_users[n] for n in rgu_users if n not in jh_data)
                pdc_only_list = sorted(pdc_only.values())
                # ComboBox = intervenants.csv (triés) + PDC-only (triés, en fin de liste)
                usernames_list = sorted(rgu_users.values()) + pdc_only_list

                # FIX 3 — Collecte des messages de log produits dans le thread secondaire.
                # Ils seront émis dans _done() qui s'exécute dans le thread principal
                # via self.after(0, _done), garantissant la thread-safety Tkinter.
                _pending_logs = []
                if csv_only:
                    _pending_logs.append((f"  ⚠  Dans CSV mais absents du PDC ({len(csv_only)}) :", "warn"))
                    for name in csv_only:
                        _pending_logs.append((f"    • {name}", "warn"))
                else:
                    _pending_logs.append(("  ✔  Tous les intervenants CSV ont des données PDC.", "ok"))
                if pdc_only_list:
                    _pending_logs.append((f"  ⚠  Dans PDC mais absents du CSV ({len(pdc_only_list)}) — ajoutés en fin de combobox :", "warn"))
                    for name in pdc_only_list:
                        _pending_logs.append((f"    • {name}", "warn"))
                else:
                    _pending_logs.append(("  ✔  Tous les intervenants PDC sont dans le CSV.", "ok"))

                # ── Dict EUR ──────────────────────────────────────────────────
                eur_data   = {}
                eur_detail = {}
                for _, row in df_eur.iterrows():
                    interv_norm = _norm_name(str(row.get(ic_eur, "")).strip())
                    prj  = str(row.get(pc_eur, "")).strip()
                    if prj not in TARGET_PROJECTS and interv_norm not in rgu_users: continue
                    mois_e = _to_mois(row.get(mc_eur, ""))
                    eur = 0.0
                    try: eur = float(str(row.get(nc_eur, "0")).replace(",", "."))
                    except: pass
                    if not prj: continue
                    eur_data[prj] = eur_data.get(prj, 0.0) + eur
                    if mois_e:
                        key_d = (interv_norm, prj, int(mois_e))
                        eur_detail[key_d] = eur_detail.get(key_d, 0.0) + eur

                # ── Diana JH réels (mois passés) depuis CSV filtré ────────────
                annee_str  = str(annee)
                mois_cur   = _mois_effectif()
                HDR_PC     = "Project Code"
                HDR_UFIRST = "User First Name"
                HDR_ULAST  = "User Last Name"
                HDR_DATE   = "Timesheet Entry Details Calendar Date"
                HDR_VAL    = "Timesheet Entry Details Value"

                diana_jh     = {}
                norm_to_orig = {}   # u_norm → u_orig (pour les logs TJM manquants)

                with open(path_diana_eff, "r", encoding="utf-8-sig", errors="replace") as f:
                    d_lines = f.readlines()

                d_hdrs = [h.strip().strip('"') for h in d_lines[0].split(SEP_OUTPUT)]
                def _dci(n):
                    try: return d_hdrs.index(n)
                    except ValueError: return -1

                di_pc  = _dci(HDR_PC)
                di_uf  = _dci(HDR_UFIRST); di_ul = _dci(HDR_ULAST)
                di_dt  = _dci(HDR_DATE);   di_vl = _dci(HDR_VAL)
                di_org = _dci("Organization Unit Name")

                # FIX 1 — _fv défini en dehors de la boucle avec argument par défaut
                # pour éviter la redéfinition de l'objet fonction à chaque itération
                def _fv(i, _flds=None):
                    return _flds[i] if _flds and 0 <= i < len(_flds) else ""

                for line in d_lines[1:]:
                    line = line.rstrip("\n\r")
                    if not line.strip(): continue
                    flds = [f.strip().strip('"') for f in line.split(SEP_OUTPUT)]
                    while len(flds) < len(d_hdrs): flds.append("")
                    _fv = lambda i, _f=flds: _f[i] if 0 <= i < len(_f) else ""

                    pc_v = _fv(di_pc)

                    # Retenir si Organization Unit Name contient "RGU" OU projet piloté
                    if "RGU" not in _fv(di_org).upper() and pc_v not in TARGET_PROJECTS:
                        continue

                    date_v = _fv(di_dt)
                    if not date_v or len(date_v) < 7: continue
                    try:
                        y_v = int(date_v[:4]); m_v = int(date_v[5:7])
                    except ValueError: continue
                    if y_v != annee or m_v >= mois_cur: continue

                    u_orig = f"{_fv(di_ul)}, {_fv(di_uf)}"
                    u_norm = _norm_name(u_orig)
                    if not u_norm: continue

                    if pc_v not in TARGET_PROJECTS and u_norm not in rgu_users: continue
                    jh_v = 0.0
                    try: jh_v = float(_fv(di_vl).replace(",","."))
                    except: pass

                    if u_norm not in diana_jh:
                        diana_jh[u_norm]     = {}
                        norm_to_orig[u_norm] = u_orig   # 1ère occurrence suffit
                    if pc_v not in diana_jh[u_norm]:
                        diana_jh[u_norm][pc_v] = {}
                    diana_jh[u_norm][pc_v][m_v] = diana_jh[u_norm][pc_v].get(m_v, 0.0) + jh_v

                # ── Histo_TJM ─────────────────────────────────────────────────
                path_histo      = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "Histo_TJM.xlsx")
                histo_tjm       = {}
                histo_tjm_found = os.path.isfile(path_histo)
                if HAS_OPENPYXL and histo_tjm_found:
                    wb_h = openpyxl.load_workbook(path_histo, read_only=True, data_only=True)
                    ws_h = wb_h.active
                    for ri, row in enumerate(ws_h.iter_rows(values_only=True)):
                        if ri == 0 or row[0] is None: continue
                        u_norm_h = _norm_name(str(row[0]).strip())
                        for mois_i in range(1, 13):
                            val = row[mois_i] if len(row) > mois_i else None
                            if val is not None:
                                try: histo_tjm[(u_norm_h, mois_i)] = float(val)
                                except (ValueError, TypeError): pass
                    wb_h.close()

                # ── EUR Diana (JH × TJM) ──────────────────────────────────────
                diana_eur    = {}
                tjm_manquants = {}   # {u_norm: set(mois)} — TJM absent pour JH > 0

                for u_norm, prj_dict in diana_jh.items():
                    for prj, mois_dict in prj_dict.items():
                        for mois, jh_v in mois_dict.items():
                            if not jh_v:
                                continue
                            tjm = histo_tjm.get((u_norm, mois))
                            if tjm:
                                eur_v = round(jh_v * tjm, 2)
                                if u_norm not in diana_eur: diana_eur[u_norm] = {}
                                if prj not in diana_eur[u_norm]: diana_eur[u_norm][prj] = {}
                                diana_eur[u_norm][prj][mois] = eur_v
                            else:
                                # TJM manquant : consommé EUR non calculable
                                tjm_manquants.setdefault(u_norm, set()).add(mois)

                # ── Log TJM manquants ──────────────────────────────────────────
                _pending_logs.append(("━" * 54, "section"))
                if not histo_tjm_found:
                    # Fichier absent : un seul avertissement global, pas de détail par intervenant
                    _pending_logs.append((
                        f"  ⚠  Histo_TJM.xlsx introuvable — consommé EUR Diana "
                        f"non calculé pour tous les intervenants.", "warn"))
                    _pending_logs.append((
                        f"  → Chemin recherché : {path_histo}", "warn"))
                    _pending_logs.append((
                        "  → Générer d'abord le fichier via le bouton "
                        "\"Générer Histo TJM\".", "warn"))
                elif tjm_manquants:
                    # Fichier présent mais des TJM sont absents pour certains intervenants/mois
                    _pending_logs.append((f"  Histo_TJM utilisé : {path_histo}", "info"))
                    _pending_logs.append((
                        f"  ⚠  TJM manquant(s) dans Histo_TJM.xlsx "
                        f"— consommé EUR Diana non calculé "
                        f"({len(tjm_manquants)} intervenant(s)) :", "warn"))
                    for u_norm in sorted(tjm_manquants):
                        u_disp    = norm_to_orig.get(u_norm, u_norm)
                        mois_list = sorted(tjm_manquants[u_norm])
                        mois_str  = ", ".join(MOIS_ABBREV[m - 1] for m in mois_list)
                        _pending_logs.append((
                            f"    • {u_disp} : {mois_str}", "warn"))
                    _pending_logs.append((
                        "  → Ajouter ces intervenants / mois dans Histo_TJM.xlsx "
                        "pour obtenir le consommé EUR.", "warn"))
                else:
                    # Tout est OK
                    _pending_logs.append((f"  Histo_TJM utilisé : {path_histo}", "info"))
                    _pending_logs.append((
                        "  ✔  TJM disponible pour tous les intervenants Diana "
                        "(consommé EUR calculé).", "ok"))

                # ── Mise à jour thread principal ──────────────────────────────
                # FIX 3 — _done() s'exécute dans le thread principal (via after),
                # garantissant que tous les appels Tkinter sont thread-safe.
                def _done():
                    self.pdc_usernames   = usernames_list
                    self.pdc_jh_data     = jh_data
                    self.pdc_eur_data    = eur_data
                    self.pdc_eur_detail  = eur_detail
                    self.pdc_projets_rgu = projets_rgu
                    self.diana_jh_data   = diana_jh
                    self.diana_eur_data  = diana_eur
                    self.histo_tjm_data  = histo_tjm
                    self._pdc_cache_ready = True
                    # FIX 3 — Émission des logs collectés dans le thread secondaire
                    for msg, tag in _pending_logs:
                        self._log(msg, tag)
                    self._log(
                        f"  ✔  Cache PDC prêt : {len(usernames_list)} intervenant(s) périmètre, "
                        f"{len(projets_rgu)} projet(s), "
                        f"{len(diana_jh)} username(s) Diana chargés, "
                        f"Histo_TJM : {len(histo_tjm)} entrée(s).", "ok")
                    self._set_status("Cache PDC + Diana prêt.")
                    if on_done:
                        self.after(0, on_done)

                self.after(0, _done)

            except Exception as exc:
                self.after(0, lambda: self._log(
                    f"  ✘  Erreur pré-chargement PDC : {exc}\n{traceback.format_exc()}", "error"))

        threading.Thread(target=_run, daemon=True).start()

    # ── Bouton Crapull ────────────────────────────────────────────────────
    def _update_crapull_btn(self):
        has_pdc = bool(self.file2_path.get() and self.file3_path.get())
        self.btn_crapull.config(state="normal" if has_pdc else "disabled")
        self.btn_pdc_maj.config(state="normal" if has_pdc else "disabled")

    def _open_crapull_window(self):
        if not HAS_PANDAS:
            messagebox.showerror("Dépendance manquante",
                                  "pandas est requis : pip install pandas openpyxl")
            return
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()
        if not path_jh or not path_eur:
            messagebox.showwarning("Fichiers manquants",
                                    "Sélectionnez d'abord les fichiers PDC (2 et 3).")
            return

        def _filtered(path):
            """Retourne le chemin du CSV filtré s'il existe, sinon le fichier source."""
            flt = os.path.splitext(path)[0] + "_avec_filtre.csv"
            return flt if os.path.isfile(flt) else path

        win = CrapullWindow(self, _filtered(path_jh), _filtered(path_eur))
        win.focus_set()

    # ── Helpers ───────────────────────────────────────────────────────────
    def _log(self, text, tag="info"):
        ts = datetime.now().strftime("%H:%M:%S")
        prefixed = text if text.strip().startswith("━") else f"[{ts}] {text}"
        def _do():
            self.log_box.config(state="normal")
            self.log_box.insert("end", prefixed + "\n", tag)
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _do)

    def _log_duration(self, label, t0, tag="ok"):
        elapsed = time.perf_counter() - t0
        dur_str = (f"{elapsed:.1f}s" if elapsed < 60
                   else f"{int(elapsed)//60}m {int(elapsed)%60:02d}s")
        self._log(f"  ⏱  {label} terminé en {dur_str}", tag)

    def _clear_log(self):
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")
        self._set_status("Journal effacé.")

    def _set_status(self, msg):
        self.after(0, lambda: self.status_var.set(msg))

    def _update_rgu_box(self, names):
        def _do():
            self.rgu_box.config(state="normal")
            self.rgu_box.delete("1.0", "end")
            self.rgu_box.insert(
                "1.0",
                "\n".join(names) if names
                else "(aucun intervenant périmètre chargé)")
            self.rgu_box.config(state="disabled")
        self.after(0, _do)

    def _update_inco_box(self, inco_names):
        def _do():
            self.inco_tree.delete(*self.inco_tree.get_children())
            rows = sorted(inco_names.items(), key=lambda x: x[1][0])
            for i, (norm, (diana_orig, tjm_orig)) in enumerate(rows):
                tag = "even" if i % 2 == 0 else "odd"
                self.inco_tree.insert("", "end",
                    values=(diana_orig, tjm_orig, norm, norm),
                    tags=(tag,))
        self.after(0, _do)

    def _update_anom_box(self, anomalies, tjm_not_found=None):
        def _do():
            self.anom_box.config(state="normal")
            self.anom_box.delete("1.0", "end")
            content_lines = []

            if anomalies:
                content_lines.append("═" * 60)
                content_lines.append(f"  CONFLITS TJM ({len(anomalies)} cas)")
                content_lines.append("  Plusieurs TJM différents pour un même intervenant/mois")
                content_lines.append("═" * 60)
                for interv, mois, tjm_max, detail in anomalies:
                    content_lines.append(
                        f"\n[Mois {mois:02d}] {interv}\n"
                        f"  TJM par projet   : {detail}\n"
                        f"  TJM retenu (max) : {tjm_max:.2f} €")
            else:
                content_lines.append("✔  Aucun conflit TJM détecté.")

            content_lines.append("")
            if tjm_not_found:
                content_lines.append("═" * 60)
                content_lines.append(f"  TJM MANQUANTS ({len(tjm_not_found)} combinaison(s))")
                content_lines.append("  conso_diana_ttnr non calculable (TJM absent dans Histo_TJM)")
                content_lines.append("═" * 60)
                for (username, mois_num), details in sorted(tjm_not_found.items()):
                    content_lines.append(
                        f"\n[Mois {mois_num:02d}] {username}")
                    for proj_name, wpn, cjh in details:
                        content_lines.append(
                            f"  ↳ Projet  : {proj_name}\n"
                            f"     WPN    : {wpn}\n"
                            f"     Conso  : {cjh:.4f} jh")
            else:
                content_lines.append("✔  Tous les TJM ont été trouvés dans Histo_TJM.")

            self.anom_box.insert("1.0", "\n".join(content_lines))
            self.anom_box.config(state="disabled")
        self.after(0, _do)


