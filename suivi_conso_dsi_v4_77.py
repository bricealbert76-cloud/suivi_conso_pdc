"""
Suivi Consommation DSI - v4.68
IHM de traitement des fichiers de saisie et plans de charges
  - Suivi Conso : génère un CSV filtré depuis le fichier saisies (.txt)
  - Histo TJM   : calcule et exporte les TJM mensuels par intervenant (Histo_TJM.xlsx)

  v4.63 : prétraitement des fichiers source
    - Génération automatique de CSV filtrés (_avec_filtre) pour Diana et les 2 PDC
    - Si le CSV filtré existe déjà, il est réutilisé directement (pas de relecture du source)
    - Le cache PDC est chargé au clic sur "MAJ Plan de charges" (plus au démarrage)
    - Le fichier intervenants (fichier 4) est présélectionné EN PREMIER au démarrage

  v4.64 : correctif consommés réels fenêtre MAJ PDC
    - Les consommés Diana (JH et EUR) ne prennent en compte que les lignes
      dont "Organization Unit Name" contient "RGU"

  v4.65 : rectification règle de gestion consommés réels fenêtre MAJ PDC
    - Condition élargie : lignes retenues si "Organization Unit Name" contient "RGU"
      OU si le projet est un projet piloté (TARGET_PROJECTS)

  v4.66 : fenêtre Vérif. données Crapull — utilisation des CSV filtrés
    - _open_crapull_window passe les fichiers _avec_filtre à CrapullWindow
      si ceux-ci existent (générés par le prétraitement v4.63)
    - Fallback sur les fichiers source originaux si le CSV filtré est absent

  v4.67 : fenêtre Vérif. données Diana — utilisation du CSV filtré
    - _open_diana_window passe le fichier _avec_filtre à DianaWindow
      si celui-ci existe (généré par le prétraitement v4.63)
    - Fallback sur le fichier source original si le CSV filtré est absent

  v4.68 : fenêtre MAJ PDC — filtre "Périmètre Nucleus uniquement"
    - Cadran bas-droite renommé "Plan de charges TTNR"
    - Case à cocher "Périmètre Nucleus uniquement" (désactivée par défaut)
    - Quand cochée : exclut du détail les intervenants intervenant
      uniquement sur "Surveillances des comm." ou uniquement sur "Analytic"
      (colonnes de intervenants.csv)

  v4.69 : erratum v4.68 — header custom déplacé sur le cadran bas-gauche
    - "Plan de charges TTNR" + checkbox sur le cadran bas-gauche (EUR)
    - Cadran bas-droite restauré en "Détail intervenants" simple

  v4.70 : debug filtre Nucleus — log des exclusions au clic checkbox
    - Au clic sur la checkbox, log dans le journal principal :
      liste des intervenants exclus, projet sélectionné, avertissement si vide

  v4.71 : filtre Nucleus — règle basée sur les projets PDC travaillés
    - SURV_COMM_PROJECTS et ANALYTIC_PROJECTS définis comme constantes
    - _compute_nucleus_exclusions() calcule dynamiquement depuis _cache_jh
      les intervenants dont TOUS les projets sont dans l'un de ces périmètres
    - Remplace la lecture depuis intervenants.csv (colonnes Surv.Comm./Analytic)

  v4.72 : filtre Nucleus — les projets vacances exclus du calcul
    - _compute_nucleus_exclusions() ignore les projets dans _vacance_projects
      (TopAbsence dans Budgets_2026.csv) avant d'évaluer le périmètre

  v4.73 : filtre Nucleus — réutilise _survcomm_intervenants (écarts JH/JO)
    - _compute_nucleus_exclusions() retourne simplement les noms normalisés
      de _survcomm_intervenants (déjà calculé à l'ouverture de la fenêtre)
    - Supprime la logique redondante basée sur SURV_COMM_PROJECTS / ANALYTIC_PROJECTS

  v4.74 : filtre Nucleus appliqué aussi au cadran bas-gauche (EUR)
    - _build_table_eur exclut les intervenants Nucleus des agrégations
      conso_date (diana_eur) et RAF (cache_eur_detail)
    - _on_nucleus_filter appelle _build_table_eur en plus de _show_proj_detail

  v4.75 : refonte filtre Nucleus
    - Suppression des constantes SURV_COMM_PROJECTS / ANALYTIC_PROJECTS (v4.71)
    - Support de TopAnalytic dans Budgets_2026.csv → _analytic_projects/_analytic_intervenants
    - Case à cocher déplacée dans la barre de sélection (même ligne que Rafraîchir)
    - Écarts JH/JO : plus filtrés systématiquement ; filtrés uniquement si case cochée
    - Case cochée : exclusion des intervenants exclusifs SurvComm ET Analytic
      dans les 3 tableaux (écarts, EUR, détail)

  v4.76 : filtre Nucleus — exclusions supplémentaires en dur
    - GUITTARD Raphael et ABDELLI Ahmed : exclusion de leurs consommés (JH et EUR)
      sur les projets topés SurvComm quand la case est cochée (intervenants multiprojets)
    - Intervenants absents de intervenants.csv : exclusion des RAF (PDC EUR) sur
      tous les projets sauf P02508 quand la case est cochée
    - Ces exclusions sont détaillées dans le journal à chaque activation du filtre

  v4.77 : filtre Nucleus — exclusions en dur périmètre Analytic
    - GUITTARD Raphael et TCHIEGAING KAMGUIA Arnaud : exclusion de leurs consommés
      (JH et EUR) sur les projets topés Analytic quand la case est cochée
"""

VERSION = "4.77"

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import csv
import threading
import time
import unicodedata
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ─────────────────────────────────────────────
#  CONSTANTES MÉTIER
# ─────────────────────────────────────────────
PROJECT_CODE_FIELD = "Project Code"
COL_SURV_COMM      = "Surveillances des comm."
COL_ANALYTIC       = "Analytic"
# Colonnes périmètre dans intervenants.csv (toutes sauf "Intervenant" et "Statut")
PERIMETRE_COLS = [
    "Acquisition", "Distribution", "Exploration", "Data Services",
    "Socle&Enabler", "Core Team", COL_SURV_COMM, COL_ANALYTIC, "IA",
]
TARGET_PROJECTS    = {"P02508", "P06628", "P06629", "P04243"}
EXCLUDED_PROJECTS  = {"MAL", "VAC", "NOC", "PH"}
SEP_INPUT          = "~"
SEP_OUTPUT         = ";"
CHECK_LINES        = 1000

COL_INTERVENANT = "Intervenant"
COL_MOIS        = "Mois"
COL_PROJET      = "PRJ_CodeProjet"
COL_NBJOURS     = "NbJours"

MOIS_ORDRE = {
    "janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4,
    "mai": 5, "juin": 6, "juillet": 7, "août": 8, "aout": 8,
    "septembre": 9, "octobre": 10, "novembre": 11,
    "décembre": 12, "decembre": 12,
    1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6,
    7: 7, 8: 8, 9: 9, 10: 10, 11: 11, 12: 12,
}
MOIS_LABELS = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

# Colonnes conservées dans le CSV Diana _avec_filtre
# = SRC_COLS (utilisées dans _process_file1) + "Project Name"
DIANA_COLS_FILTRE = [
    "Project Code", "Project Is Alive", "WorkPackage Name", "Task Name",
    "Organization Unit Name", "User First Name", "User Last Name",
    "User Country Name", "User Contract Type", "TimeSheet Mode Name",
    "Timesheet Entry Details Calendar Date", "Status", "Is Published",
    "Timesheet Entry Details Value", "Application Cluster Name",
    "Project Name",
]

# Colonnes conservées dans les CSV PDC _avec_filtre
# Obligatoires + optionnelles (conservées si présentes dans le fichier source)
PDC_COLS_FILTRE_MANDATORY = [COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS]
PDC_COLS_FILTRE_OPTIONAL  = [
    "PRJ_LibelleProjet", "Year of charge", "PRJ_RespNiv1",
    "Code UT", "Type Intervenant",
]


# ─────────────────────────────────────────────
#  THÈMES  (sélection via --style dark|light|blue)
# ─────────────────────────────────────────────
THEMES = {
    "dark": {
        "BG_MAIN"  : "#0d1117", "BG_PANEL" : "#161b22",
        "BG_CARD"  : "#1c2128", "BG_ENTRY" : "#21262d",
        "ACCENT"   : "#2ea043", "ACCENT2"  : "#388bfd",
        "WARN"     : "#f0883e", "TEXT_PRI" : "#e6edf3",
        "TEXT_SEC" : "#8b949e", "BORDER"   : "#30363d",
        "BTN_ACT"  : "#238636", "BTN_ACT_H": "#2ea043",
        "BTN_NEU"  : "#21262d", "BTN_NEU_H": "#30363d",
    },
    "light": {
        "BG_MAIN"  : "#f6f8fa", "BG_PANEL" : "#ffffff",
        "BG_CARD"  : "#eaeef2", "BG_ENTRY" : "#ffffff",
        "ACCENT"   : "#1a7f37", "ACCENT2"  : "#0969da",
        "WARN"     : "#bc4c00", "TEXT_PRI" : "#1f2328",
        "TEXT_SEC" : "#656d76", "BORDER"   : "#d0d7de",
        "BTN_ACT"  : "#1a7f37", "BTN_ACT_H": "#2ea043",
        "BTN_NEU"  : "#eaeef2", "BTN_NEU_H": "#d0d7de",
    },
    "blue": {
        "BG_MAIN"  : "#0f1923", "BG_PANEL" : "#162032",
        "BG_CARD"  : "#1c2d42", "BG_ENTRY" : "#1e3a5f",
        "ACCENT"   : "#00b4d8", "ACCENT2"  : "#48cae4",
        "WARN"     : "#f77f00", "TEXT_PRI" : "#e0f2fe",
        "TEXT_SEC" : "#90c4e4", "BORDER"   : "#2a4a6b",
        "BTN_ACT"  : "#0077b6", "BTN_ACT_H": "#00b4d8",
        "BTN_NEU"  : "#1e3a5f", "BTN_NEU_H": "#2a4a6b",
    },
}

def _apply_theme(name):
    theme = THEMES.get(name, THEMES["dark"])
    globals().update(theme)

import argparse as _argparse
_parser = _argparse.ArgumentParser(add_help=False)
_parser.add_argument("--style", choices=["dark", "light", "blue"], default="dark")
_args, _ = _parser.parse_known_args()
_apply_theme(_args.style)

FONT_TITLE = ("Consolas", 18, "bold")
FONT_HEAD  = ("Consolas", 11, "bold")
FONT_BODY  = ("Consolas", 10)
FONT_SMALL = ("Consolas", 9)
FONT_MONO  = ("Courier New", 9)


def _calc_jours_ouvres(annee):
    """Calcule les jours ouvrés par mois pour l'année donnée (France métropolitaine).
    Utilise uniquement la bibliothèque standard Python (pas de dépendance externe).
    Retourne une liste de 12 entiers (janvier → décembre).
    """
    import calendar
    from datetime import date, timedelta

    # Calcul de Pâques — algorithme grégorien anonyme
    a = annee % 19
    b = annee // 100
    c = annee % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mois_p = (h + l - 7 * m + 114) // 31
    jour_p = ((h + l - 7 * m + 114) % 31) + 1
    paques = date(annee, mois_p, jour_p)

    feries = {
        date(annee, 1,  1),               # Jour de l'an
        paques + timedelta(days=1),        # Lundi de Pâques
        date(annee, 5,  1),               # Fête du travail
        date(annee, 5,  8),               # Victoire 1945
        paques + timedelta(days=39),       # Ascension
        paques + timedelta(days=50),       # Lundi de Pentecôte
        date(annee, 7,  14),              # Fête nationale
        date(annee, 8,  15),              # Assomption
        date(annee, 11, 1),               # Toussaint
        date(annee, 11, 11),              # Armistice
        date(annee, 12, 25),              # Noël
    }

    result = []
    for mois in range(1, 13):
        nb = sum(
            1 for jour in range(1, calendar.monthrange(annee, mois)[1] + 1)
            if date(annee, mois, jour).weekday() < 5        # lundi–vendredi
            and date(annee, mois, jour) not in feries
        )
        result.append(nb)
    return result


def _norm_name(s):
    """Normalise un nom pour la jointure : supprime les accents, met en majuscules,
    remplace tirets et toutes variantes d'apostrophes par des espaces."""
    if not s:
        return ""
    # Décomposition unicode → suppression des diacritiques
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.upper()
    # Remplacer TOUTES les variantes d'apostrophes/guillemets simples et tirets
    # par un espace (U+0027, U+2019, U+2018, U+02BC, U+2032, U+055A, U+2010, etc.)
    _APOSTROPHES = (
        "\u0027",  # apostrophe ASCII '
        "\u2019",  # apostrophe typographique '
        "\u2018",  # apostrophe réversée '
        "\u02bc",  # apostrophe modificateur ʼ
        "\u2032",  # prime ′
        "\u055a",  # apostrophe arménienne ՚
        "\u0060",  # accent grave `
        "\u00b4",  # accent aigu ´
        "\u2010",  # trait d'union sans espace ‐
        "\u2011",  # trait d'union insécable ‑
        "-",        # tiret standard
    )
    for apo in _APOSTROPHES:
        s = s.replace(apo, " ")
    # Compresser les espaces multiples
    return " ".join(s.split())


# Intervenants multiprojets avec exclusion partielle SurvComm (filtre Nucleus en dur)
_NUCLEUS_HARDCODED_SURVCOMM_ORIG = ["GUITTARD, Raphael", "ABDELLI, Ahmed"]
_NUCLEUS_HARDCODED_SURVCOMM_NORM = frozenset(_norm_name(u) for u in _NUCLEUS_HARDCODED_SURVCOMM_ORIG)
# Intervenants multiprojets avec exclusion partielle Analytic (filtre Nucleus en dur)
_NUCLEUS_HARDCODED_ANALYTIC_ORIG = ["GUITTARD, Raphael", "TCHIEGAING KAMGUIA, Arnaud"]
_NUCLEUS_HARDCODED_ANALYTIC_NORM = frozenset(_norm_name(u) for u in _NUCLEUS_HARDCODED_ANALYTIC_ORIG)
# Union des deux pour les vérifications communes
_NUCLEUS_HARDCODED_NORM = _NUCLEUS_HARDCODED_SURVCOMM_NORM | _NUCLEUS_HARDCODED_ANALYTIC_NORM
# Seul projet autorisé en RAF (PDC EUR) pour les non-intervenants (filtre Nucleus)
_NUCLEUS_P02508_ONLY = "P02508"


def _load_allowed_intervenants(path):
    """Charge le fichier intervenants (path) et retourne (set_norms, list_originals)
    pour tous les intervenants non marqués dans la colonne COL_SURV_COMM."""
    allowed_norms = set()
    allowed_list  = []
    if not path or not os.path.isfile(path):
        return allowed_norms, allowed_list
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                if row.get(COL_SURV_COMM, "").strip() == "1":
                    continue
                name = row.get("Intervenant", "").strip()
                if not name:
                    continue
                allowed_list.append(name)
                allowed_norms.add(_norm_name(name))
        allowed_list.sort()
    except Exception:
        pass
    return allowed_norms, allowed_list


def _load_nucleus_exclusions(path):
    """Retourne un set de noms normalisés des intervenants à exclure du périmètre Nucleus.
    Sont exclus ceux dont la SEULE colonne périmètre cochée est COL_SURV_COMM
    ou la SEULE colonne périmètre cochée est COL_ANALYTIC.
    """
    exclusions = set()
    if not path or not os.path.isfile(path):
        return exclusions
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                name = row.get("Intervenant", "").strip()
                if not name:
                    continue
                active = {col for col in PERIMETRE_COLS
                          if row.get(col, "").strip() == "1"}
                if active == {COL_SURV_COMM} or active == {COL_ANALYTIC}:
                    exclusions.add(_norm_name(name))
    except Exception:
        pass
    return exclusions


def make_button(parent, text, command, style="neutral", width=None):
    bg  = BTN_ACT  if style == "action" else BTN_NEU
    fg  = "#ffffff" if style == "action" else TEXT_PRI
    kw  = dict(text=text, command=command, bg=bg, fg=fg,
               font=FONT_HEAD, bd=0, padx=14, pady=8,
               relief="flat", cursor="hand2", activeforeground="#ffffff")
    if width:
        kw["width"] = width
    btn = tk.Button(parent, **kw)
    hov = BTN_ACT_H if style == "action" else BTN_NEU_H
    btn.bind("<Enter>", lambda e: btn.config(bg=hov))
    btn.bind("<Leave>", lambda e: btn.config(bg=bg))
    return btn


# ─────────────────────────────────────────────
#  APPLICATION
# ─────────────────────────────────────────────
class SuiviConsoApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"Suivi Consommation DSI — v{VERSION}")
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)
        self.minsize(920, 700)

        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.file3_path = tk.StringVar()
        self.file4_path = tk.StringVar()
        self.status_var = tk.StringVar(value="Prêt.")

        self._build_ui()
        self._center_window(980, 780)
        # Intervenants périmètre (hors Surveillances des comm.)
        self._allowed_norms      = set()
        self._allowed_list       = []
        self._nucleus_exclusions = set()  # intervenants uniquement Surv.Comm. ou Analytic

        # ── Point d'entrée unique au démarrage : fichier 4 → intervenants → présélections → prétraitement
        self.after(100, self._preselectfile4)

        # Structures pré-chargées pour la fenêtre MAJ PDC
        # (plus chargées au démarrage — uniquement au clic sur le bouton MAJ PDC)
        self.pdc_usernames    = []
        self.pdc_jh_data      = {}
        self.pdc_eur_data     = {}
        self.pdc_projets_rgu  = set()
        self.pdc_eur_detail   = {}
        self.diana_jh_data    = {}
        self.diana_eur_data   = {}
        self.histo_tjm_data   = {}
        self._pdc_cache_ready = False

    def _center_window(self, w, h):
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    # ── Interface ────────────────────────────────────────────────────────
    def _build_ui(self):
        hdr = tk.Frame(self, bg=BG_MAIN, pady=16)
        hdr.pack(fill="x", padx=24)
        tk.Label(hdr, text="⬡  SUIVI CONSOMMATION DSI",
                 bg=BG_MAIN, fg=ACCENT, font=FONT_TITLE).pack(side="left")
        tk.Label(hdr, text=f"v{VERSION}", bg=BG_MAIN, fg=TEXT_SEC,
                 font=FONT_SMALL).pack(side="left", padx=8, pady=4, anchor="s")
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        body = tk.Frame(self, bg=BG_MAIN)
        body.pack(fill="both", expand=True, padx=24, pady=16)

        left = tk.Frame(body, bg=BG_PANEL, bd=0,
                        highlightbackground=BORDER, highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0, 12), ipadx=14, ipady=14)
        self._build_file_panel(left)

        right = tk.Frame(body, bg=BG_PANEL, bd=0,
                         highlightbackground=BORDER, highlightthickness=1)
        right.pack(side="left", fill="both", expand=True, ipadx=14, ipady=14)
        self._build_result_panel(right)

        bar = tk.Frame(self, bg=BG_CARD, height=30)
        bar.pack(fill="x", side="bottom")
        tk.Label(bar, textvariable=self.status_var, bg=BG_CARD,
                 fg=TEXT_SEC, font=FONT_SMALL, anchor="w", padx=12).pack(fill="x")

    def _build_file_panel(self, parent):
        tk.Label(parent, text="SOURCES DE DONNÉES", bg=BG_PANEL,
                 fg=ACCENT2, font=FONT_HEAD).pack(anchor="w", padx=10, pady=(10, 14))

        specs = [
            ("1", "Saisies collaborateurs (.txt)",
             "Fichier texte séparé par '~'",
             self.file1_path, self._browse_file1),
            ("2", "Plan de charges — jours/hommes (.xlsx/.csv)",
             "Consommés réels & prévisionnels JH",
             self.file2_path, self._browse_file2),
            ("3", "Plan de charges — euros (.xlsx/.csv)",
             "Consommés réels & prévisionnels €",
             self.file3_path, self._browse_file3),
            ("4", "Intervenants périmètre (.csv)",
             "Référentiel intervenants — filtre hors Surveillances des comm.",
             self.file4_path, self._browse_file4),
        ]
        for num, label, hint, var, cmd in specs:
            self._build_file_row(parent, num, label, hint, var, cmd)

        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=10, pady=14)

        bf = tk.Frame(parent, bg=BG_PANEL)
        bf.pack(padx=10, pady=4, fill="x")

        self.btn_suivi = make_button(bf, "▶  Suivi Conso",
                                     self._run_suivi_conso,
                                     style="action", width=22)
        self.btn_suivi.pack(fill="x")

        self.btn_tjm = make_button(bf, "📊  Générer Histo TJM",
                                   self._run_histo_tjm,
                                   style="action", width=22)
        self.btn_tjm.pack(fill="x", pady=(8, 0))

        self.btn_crapull = make_button(bf, "🔍  Vérif. données Crapull",
                                       self._open_crapull_window,
                                       width=22)
        self.btn_crapull.pack(fill="x", pady=(8, 0))
        self.btn_crapull.config(state="disabled")

        self.btn_diana = make_button(bf, "📋  Vérif. données Diana",
                                     self._open_diana_window,
                                     width=22)
        self.btn_diana.pack(fill="x", pady=(8, 0))
        self.btn_diana.config(state="disabled")

        self.btn_pdc_maj = make_button(bf, "📝  MAJ Plan de charges",
                                       self._open_pdc_maj_window,
                                       width=22)
        self.btn_pdc_maj.pack(fill="x", pady=(8, 0))
        self.btn_pdc_maj.config(state="disabled")

        make_button(bf, "✕  Effacer journal",
                    self._clear_log, width=22).pack(fill="x", pady=(8, 0))

    def _build_file_row(self, parent, num, label, hint, var, cmd):
        card = tk.Frame(parent, bg=BG_CARD,
                        highlightbackground=BORDER, highlightthickness=1)
        card.pack(fill="x", padx=10, pady=4, ipady=6, ipadx=8)

        top = tk.Frame(card, bg=BG_CARD)
        top.pack(fill="x")
        tk.Label(top, text=f" {num} ", bg=ACCENT2, fg="#ffffff",
                 font=("Consolas", 9, "bold")).pack(side="left", padx=(0, 6))
        tk.Label(top, text=label, bg=BG_CARD, fg=TEXT_PRI,
                 font=FONT_BODY).pack(side="left")
        tk.Label(card, text=hint, bg=BG_CARD, fg=TEXT_SEC,
                 font=FONT_SMALL).pack(anchor="w", padx=2)

        row = tk.Frame(card, bg=BG_CARD)
        row.pack(fill="x", pady=(4, 0))
        tk.Entry(row, textvariable=var, bg=BG_ENTRY, fg=TEXT_SEC,
                 font=FONT_SMALL, insertbackground=TEXT_PRI, relief="flat", bd=0,
                 highlightbackground=BORDER, highlightthickness=1,
                 width=28).pack(side="left", fill="x", expand=True, ipady=4)
        make_button(row, "…", cmd, width=3).pack(side="left", padx=(6, 0))

    def _build_result_panel(self, parent):
        tk.Label(parent, text="RÉSULTATS & JOURNAL",
                 bg=BG_PANEL, fg=ACCENT2, font=FONT_HEAD).pack(
            anchor="w", padx=10, pady=(10, 8))

        sty = ttk.Style()
        sty.theme_use("default")
        sty.configure("DSI.TNotebook", background=BG_PANEL, borderwidth=0)
        sty.configure("DSI.TNotebook.Tab", background=BG_ENTRY,
                      foreground=TEXT_SEC, font=FONT_SMALL,
                      padding=[10, 4], borderwidth=0)
        sty.map("DSI.TNotebook.Tab",
                background=[("selected", BG_CARD)],
                foreground=[("selected", TEXT_PRI)])

        nb = ttk.Notebook(parent, style="DSI.TNotebook")
        nb.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Journal
        t1 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t1, text=" Journal ")
        self.log_box = scrolledtext.ScrolledText(
            t1, bg=BG_CARD, fg=TEXT_PRI, font=FONT_MONO,
            bd=0, relief="flat", insertbackground=TEXT_PRI,
            state="disabled", wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=4, pady=4)
        self._configure_log_tags()

        # Intervenants périmètre
        t2 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t2, text=" Intervenants périmètre ")
        tk.Label(t2, text="Intervenants retenus (hors Surveillances des comm.)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))
        self.rgu_box = scrolledtext.ScrolledText(
            t2, bg=BG_CARD, fg=ACCENT, font=FONT_MONO,
            bd=0, relief="flat", state="disabled", wrap="word")
        self.rgu_box.pack(fill="both", expand=True, padx=4, pady=4)

        # Codes Projet
        t3 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t3, text=" Codes Projet ciblés ")
        tk.Label(t3, text="Codes projets filtrés (liste fixe)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))
        pb = scrolledtext.ScrolledText(t3, bg=BG_CARD, fg=WARN,
                                        font=FONT_MONO, bd=0, relief="flat")
        pb.pack(fill="both", expand=True, padx=4, pady=4)
        pb.insert("1.0", "\n".join(sorted(TARGET_PROJECTS)))
        pb.config(state="disabled")

        # Anomalies TJM
        t4 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t4, text=" Anomalies TJM ")
        tk.Label(t4,
                 text="Conflits TJM (plusieurs valeurs) + TJM manquants (conso_diana_ttnr non calculable)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))
        self.anom_box = scrolledtext.ScrolledText(
            t4, bg=BG_CARD, fg=WARN, font=FONT_MONO,
            bd=0, relief="flat", state="disabled", wrap="word")
        self.anom_box.pack(fill="both", expand=True, padx=4, pady=4)

        # Incohérence des noms
        t5 = tk.Frame(nb, bg=BG_CARD)
        nb.add(t5, text=" Incohérence des noms ")
        tk.Label(t5,
                 text="Noms différents entre Diana et Histo_TJM (avant normalisation)",
                 bg=BG_CARD, fg=TEXT_SEC, font=FONT_SMALL).pack(
            anchor="w", padx=8, pady=(6, 2))

        inco_frame = tk.Frame(t5, bg=BG_CARD)
        inco_frame.pack(fill="both", expand=True, padx=4, pady=4)
        vsb_i = tk.Scrollbar(inco_frame, orient="vertical")
        hsb_i = tk.Scrollbar(inco_frame, orient="horizontal")
        vsb_i.pack(side="right", fill="y")
        hsb_i.pack(side="bottom", fill="x")
        inco_cols = ("nom_diana", "nom_histo_tjm", "norm_diana", "norm_histo_tjm")
        self.inco_tree = ttk.Treeview(
            inco_frame, columns=inco_cols, show="headings",
            yscrollcommand=vsb_i.set, xscrollcommand=hsb_i.set)
        vsb_i.config(command=self.inco_tree.yview)
        hsb_i.config(command=self.inco_tree.xview)
        headers_inco = {
            "nom_diana":      ("Nom Diana (original)",    200),
            "nom_histo_tjm":  ("Nom Histo_TJM (original)", 200),
            "norm_diana":     ("Nom Diana normalisé",      200),
            "norm_histo_tjm": ("Nom Histo_TJM normalisé",  200),
        }
        style_inco = ttk.Style()
        style_inco.configure("Inco.Treeview",
                              background=BG_CARD, foreground=TEXT_PRI,
                              fieldbackground=BG_CARD, rowheight=22,
                              font=FONT_SMALL)
        style_inco.configure("Inco.Treeview.Heading",
                              background=BG_ENTRY, foreground=ACCENT2,
                              font=("Consolas", 9, "bold"))
        self.inco_tree.config(style="Inco.Treeview")
        for col, (head, width) in headers_inco.items():
            self.inco_tree.heading(col, text=head)
            self.inco_tree.column(col, width=width, anchor="w", stretch=True)
        self.inco_tree.tag_configure("odd",  background=BG_CARD)
        self.inco_tree.tag_configure("even", background=BG_ENTRY)
        self.inco_tree.pack(fill="both", expand=True)

    def _configure_log_tags(self):
        self.log_box.tag_config("info",    foreground=TEXT_PRI)
        self.log_box.tag_config("ok",      foreground=ACCENT)
        self.log_box.tag_config("warn",    foreground=WARN)
        self.log_box.tag_config("error",   foreground="#f85149")
        self.log_box.tag_config("section", foreground=ACCENT2,
                                font=("Consolas", 10, "bold"))

    # ── Navigateurs fichiers ──────────────────────────────────────────────
    def _preselectfile1(self):
        import glob, re
        script_dir = os.path.dirname(os.path.abspath(__file__))
        diana_dir  = os.path.join(script_dir, "diana_files")
        pattern    = os.path.join(diana_dir, "YTDTimesheet_*.txt")
        date_re    = re.compile(r"YTDTimesheet_(\d{2}-\d{2}-\d{4})\.txt$", re.IGNORECASE)

        self._log(f"  Recherche fichiers Diana : YTDTimesheet_DD-MM-YYYY.txt"
                  f" dans {diana_dir}", "info")

        if not os.path.isdir(diana_dir):
            msg = f"Répertoire ./diana_files absent ({diana_dir}) — sélection manuelle."
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)
            return

        candidates = glob.glob(pattern)
        self._log(f"  {len(candidates)} fichier(s) YTDTimesheet_*.txt trouvé(s) :", "info")
        for p in sorted(candidates):
            self._log(f"      • {os.path.basename(p)}", "info")

        best_path = None
        best_sort = ""
        for p in candidates:
            m = date_re.search(os.path.basename(p))
            if m:
                dd, mm, yyyy = m.group(1).split("-")
                sort_key = f"{yyyy}-{mm}-{dd}"
                if sort_key > best_sort:
                    best_sort = sort_key
                    best_path = p

        if best_path:
            self.file1_path.set(best_path)
            self._log(f"  ✔  Fichier 1 présélectionné : {os.path.basename(best_path)}", "ok")
            self._set_status(f"Fichier Diana présélectionné : {os.path.basename(best_path)}")
            self.after(0, self._update_diana_btn)
        else:
            msg = ("Aucun fichier YTDTimesheet_DD-MM-YYYY.txt trouvé dans ./diana_files "
                   "— vérifiez le nom des fichiers.")
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)

    def _browse_file1(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier saisies collaborateurs",
            filetypes=[("Fichiers texte", "*.txt"), ("Tous", "*.*")])
        if not path:
            return
        self._set_status("Vérification du séparateur…")
        ok, msg = self._check_separator(path)
        if ok:
            self.file1_path.set(path)
            self._log(f"✔  Fichier 1 : {os.path.basename(path)}", "ok")
            self._log(f"   Séparateur '~' confirmé sur ≤{CHECK_LINES} lignes.", "ok")
            self._update_diana_btn()
        else:
            messagebox.showerror("Contrôle séparateur",
                                 f"Le fichier sélectionné ne semble pas utiliser\n"
                                 f"le caractère '~' comme séparateur.\n\n{msg}")
            self._log(f"✘  Fichier 1 refusé : {os.path.basename(path)}", "error")
            self._log(f"   {msg}", "warn")
        self._set_status("Prêt.")

    def _preselectfiles23(self):
        import glob, re
        script_dir = os.path.dirname(os.path.abspath(__file__))
        pdc_dir    = os.path.join(script_dir, "pdc_files")

        DATE_PAT = r"(\d{4}-\d{2}-\d{2})"

        self._log(f"  Recherche fichiers PDC dans {pdc_dir}", "info")
        self._log(f"  Patterns : SynthesePrevIntervenant_Pere_JH_YYYY-MM-DD.csv/.xlsx", "info")
        self._log(f"           & SynthesePrevIntervenant_Pere_TTNR_YYYY-MM-DD.csv/.xlsx", "info")

        if not os.path.isdir(pdc_dir):
            msg = f"Répertoire ./pdc_files absent ({pdc_dir}) — sélection manuelle."
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)
            return

        def _scan_pairs(ext):
            e = ext.lstrip(".")
            pat_jh   = re.compile(
                rf"SynthesePrevIntervenant_Pere_JH_{DATE_PAT}\.{e}$", re.IGNORECASE)
            pat_ttnr = re.compile(
                rf"SynthesePrevIntervenant_Pere_TTNR_{DATE_PAT}\.{e}$", re.IGNORECASE)
            jh_files, ttnr_files = {}, {}
            for f in glob.glob(os.path.join(pdc_dir, f"*.{e}")):
                bn = os.path.basename(f)
                m = pat_jh.match(bn)
                if m: jh_files[m.group(1)] = f
                m = pat_ttnr.match(bn)
                if m: ttnr_files[m.group(1)] = f
            self._log(f"  Fichiers JH   [{ext.upper()}] : {len(jh_files)} → "
                      f"{[os.path.basename(v) for v in jh_files.values()]}", "info")
            self._log(f"  Fichiers TTNR [{ext.upper()}] : {len(ttnr_files)} → "
                      f"{[os.path.basename(v) for v in ttnr_files.values()]}", "info")
            common = sorted(set(jh_files) & set(ttnr_files), reverse=True)
            if common:
                d = common[0]
                return jh_files[d], ttnr_files[d], d, ext.upper()
            return None, None, None, None

        path_jh, path_eur, date_sel, fmt = _scan_pairs(".csv")
        if not path_jh:
            self._log("  Aucune paire CSV — recherche en XLSX...", "info")
            path_jh, path_eur, date_sel, fmt = _scan_pairs(".xlsx")

        if path_jh and path_eur:
            self.file2_path.set(path_jh)
            self.file3_path.set(path_eur)
            self._log(f"  ✔  Fichier 2 [{fmt}] présélectionné : {os.path.basename(path_jh)}", "ok")
            self._log(f"  ✔  Fichier 3 [{fmt}] présélectionné : {os.path.basename(path_eur)}", "ok")
            self._set_status(f"PDC présélectionnés ({fmt}, date {date_sel}).")
            self.after(0, self._update_crapull_btn)
        else:
            msg = ("Aucune paire PDC (JH + TTNR, même date) trouvée dans ./pdc_files "
                   "— vérifiez les noms de fichiers.")
            self._log(f"  ⚠  {msg}", "warn")
            self._set_status(msg)

    def _browse_file2(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier plan de charges JH",
            filetypes=[("Excel ou CSV", "*.xlsx *.csv"),
                       ("Fichiers Excel", "*.xlsx"),
                       ("Fichiers CSV",   "*.csv"),
                       ("Tous", "*.*")])
        if not path:
            return
        other = self.file3_path.get()
        if other and (os.path.splitext(path)[1].lower() !=
                      os.path.splitext(other)[1].lower()):
            messagebox.showerror(
                "Formats incompatibles",
                "Les fichiers 2 et 3 doivent être dans le même format\n"
                "(tous les deux .xlsx  OU  tous les deux .csv).")
            return
        self.file2_path.set(path)
        ext = os.path.splitext(path)[1].upper()
        self._log(f"✔  Fichier 2 [{ext}] : {os.path.basename(path)}", "ok")
        self._update_crapull_btn()

    def _browse_file3(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier plan de charges €",
            filetypes=[("Excel ou CSV", "*.xlsx *.csv"),
                       ("Fichiers Excel", "*.xlsx"),
                       ("Fichiers CSV",   "*.csv"),
                       ("Tous", "*.*")])
        if not path:
            return
        other = self.file2_path.get()
        if other and (os.path.splitext(path)[1].lower() !=
                      os.path.splitext(other)[1].lower()):
            messagebox.showerror(
                "Formats incompatibles",
                "Les fichiers 2 et 3 doivent être dans le même format\n"
                "(tous les deux .xlsx  OU  tous les deux .csv).")
            return
        self.file3_path.set(path)
        ext = os.path.splitext(path)[1].upper()
        self._log(f"✔  Fichier 3 [{ext}] : {os.path.basename(path)}", "ok")
        self._update_crapull_btn()

    def _browse_file4(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier intervenants périmètre",
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous", "*.*")])
        if not path:
            return
        self.file4_path.set(path)
        self._log(f"✔  Fichier 4 [CSV] : {os.path.basename(path)}", "ok")
        self._reload_intervenants()
        # Si les fichiers source sont déjà sélectionnés → lancer le prétraitement
        if self.file1_path.get() and self.file2_path.get() and self.file3_path.get():
            self._preprocess_sources_async()

    def _preselectfile4(self):
        """Chef d'orchestre au démarrage :
        1. Cherche intervenants.csv dans le répertoire du script
        2. Charge _allowed_norms (synchrone — rapide)
        3. Présélectionne les fichiers 1, 2, 3
        4. Lance _preprocess_sources() en arrière-plan (avec spinner)
        Si intervenants.csv est absent → log + attente sélection manuelle.
        """
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidate  = os.path.join(script_dir, "intervenants.csv")
        if not os.path.isfile(candidate):
            self._log("  ⚠  intervenants.csv absent du répertoire script — sélection manuelle.", "warn")
            self._log("  ℹ  Le prétraitement sera déclenché après la sélection du fichier 4.", "info")
            return

        # 1 — Charger les intervenants (synchrone, fichier CSV léger)
        self.file4_path.set(candidate)
        self._log(f"  ✔  Fichier 4 présélectionné : intervenants.csv", "ok")
        self._reload_intervenants()   # _allowed_norms garanti chargé ici

        # 2 — Présélectionner les fichiers source 1, 2, 3 (synchrone)
        self._preselectfile1()
        self._preselectfiles23()

        # 3 — Lancer le prétraitement en arrière-plan si les fichiers sont présents
        if self.file1_path.get() and self.file2_path.get() and self.file3_path.get():
            self._preprocess_sources_async()
        else:
            self._log("  ℹ  Prétraitement différé (fichiers source incomplets).", "info")

    def _reload_intervenants(self):
        path = self.file4_path.get()
        self._allowed_norms, self._allowed_list = _load_allowed_intervenants(path)
        self._nucleus_exclusions = _load_nucleus_exclusions(path)
        self._update_rgu_box(self._allowed_list)
        self._log(f"  Intervenants périmètre chargés : {len(self._allowed_list)}", "info")

    # ── Spinner dans le journal ───────────────────────────────────────────
    _SPINNER_FRAMES = ("◐", "◓", "◑", "◒")

    def _spinner_start(self, label):
        """Insère une ligne 'label ◐' dans le journal et la fait tourner sur place.
        Retourne un token (index de ligne) pour l'arrêter via _spinner_stop(token).
        """
        ts   = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] ⏳  {label}  {self._SPINNER_FRAMES[0]}\n"

        self.log_box.config(state="normal")
        self.log_box.insert("end", line, "warn")
        self.log_box.see("end")
        # Mémoriser l'indice de début de cette ligne pour la mettre à jour sur place
        line_index = self.log_box.index("end - 1 line linestart")
        self.log_box.config(state="disabled")

        self._spinner_active = True
        self._spinner_label  = label
        self._spinner_ts     = ts
        self._spinner_frame  = 0
        self._spinner_index  = line_index
        self._spinner_tick()
        return line_index

    def _spinner_tick(self):
        if not getattr(self, "_spinner_active", False):
            return
        self._spinner_frame = (self._spinner_frame + 1) % len(self._SPINNER_FRAMES)
        frame = self._SPINNER_FRAMES[self._spinner_frame]
        new_line = (f"[{self._spinner_ts}] ⏳  {self._spinner_label}  {frame}\n")
        self.log_box.config(state="normal")
        start = self._spinner_index
        end   = f"{start} lineend + 1 char"
        self.log_box.delete(start, end)
        self.log_box.insert(start, new_line, "warn")
        self.log_box.see("end")
        self.log_box.config(state="disabled")
        self._spinner_after_id = self.after(120, self._spinner_tick)

    def _spinner_stop(self, label_done, duration_str=""):
        """Arrête le spinner et remplace la ligne par un message de fin."""
        self._spinner_active = False
        if hasattr(self, "_spinner_after_id"):
            self.after_cancel(self._spinner_after_id)
        suffix = f"  ({duration_str})" if duration_str else ""
        done_line = f"[{self._spinner_ts}] ✔  {label_done}{suffix}\n"
        self.log_box.config(state="normal")
        start = self._spinner_index
        end   = f"{start} lineend + 1 char"
        self.log_box.delete(start, end)
        self.log_box.insert(start, done_line, "ok")
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    # ── Prétraitement asynchrone ──────────────────────────────────────────
    def _disable_all_buttons(self):
        """Désactive tous les boutons d'action pendant le prétraitement."""
        for btn in (self.btn_suivi, self.btn_tjm,
                    self.btn_crapull, self.btn_diana, self.btn_pdc_maj):
            btn.config(state="disabled")

    def _restore_buttons(self):
        """Restaure l'état des boutons selon les fichiers sélectionnés."""
        self.btn_suivi.config(state="normal")
        self.btn_tjm.config(state="normal")
        self._update_diana_btn()
        self._update_crapull_btn()

    def _preprocess_sources_async(self, on_done=None):
        """Lance _preprocess_sources() dans un thread avec spinner animé.
        Désactive tous les boutons pendant l'exécution, les restaure à la fin.
        """
        self._disable_all_buttons()
        self._set_status("Prétraitement des CSV filtrés en cours…")
        self._spinner_start("Génération des CSV filtrés")

        def _run():
            t0 = time.perf_counter()
            try:
                path_diana, path_jh, path_eur = self._preprocess_sources()
                elapsed = time.perf_counter() - t0
                dur_str = (f"{elapsed:.1f}s" if elapsed < 60
                           else f"{int(elapsed)//60}m {int(elapsed)%60:02d}s")
                self.after(0, lambda: self._spinner_stop(
                    "CSV filtrés prêts", dur_str))
                self.after(0, lambda: self._set_status(
                    f"Prétraitement terminé en {dur_str} — prêt."))
                self.after(0, self._restore_buttons)
                if on_done:
                    self.after(0, on_done)
            except Exception as exc:
                import traceback
                self.after(0, lambda: self._spinner_stop("Erreur prétraitement"))
                self.after(0, lambda: self._log(f"  ✘  Prétraitement : {exc}", "error"))
                self.after(0, lambda: self._log(traceback.format_exc(), "error"))
                self.after(0, lambda: self._set_status(f"Erreur prétraitement : {exc}"))
                self.after(0, self._restore_buttons)

        threading.Thread(target=_run, daemon=True).start()

    # ── Contrôle séparateur ───────────────────────────────────────────────
    def _check_separator(self, path):
        try:
            found = checked = 0
            with open(path, "r", encoding="utf-8", errors="replace") as fh:
                for i, line in enumerate(fh):
                    if i >= CHECK_LINES:
                        break
                    if line.strip():
                        checked += 1
                        if SEP_INPUT in line:
                            found += 1
            if checked == 0:
                return False, "Fichier vide ou non lisible."
            if found / checked >= 0.5:
                return True, f"{found}/{checked} lignes contiennent '~'."
            return False, (f"Seulement {found}/{checked} lignes contiennent '~'. "
                           f"Séparateur incorrect ou fichier mal formaté.")
        except Exception as exc:
            return False, str(exc)

    # ══════════════════════════════════════════════════════════════════════
    #  PRÉTRAITEMENT DES FICHIERS SOURCE  (v4.63)
    # ══════════════════════════════════════════════════════════════════════
    def _preprocess_sources(self):
        """Génère les CSV filtrés pour Diana et les 2 PDC si nécessaire,
        puis retourne les chemins effectifs à utiliser pour la suite du traitement.

        Logique par fichier :
          • Si le CSV _avec_filtre existe déjà → l'utiliser directement.
          • Sinon :
              - Diana .txt   → générer CSV sans filtre + CSV _avec_filtre
              - PDC .xlsx    → générer CSV sans filtre + CSV _avec_filtre
              - PDC .csv     → générer CSV _avec_filtre uniquement
                               (le source est déjà un CSV brut)

        Retourne (path_diana_eff, path_jh_eff, path_eur_eff).
        """
        self._log("━" * 54, "section")
        self._log(" PRÉTRAITEMENT — génération CSV filtrés", "section")
        self._log("━" * 54, "section")

        path_diana_eff = self._preprocess_diana()
        path_jh_eff    = self._preprocess_pdc(self.file2_path.get(), label="JH")
        path_eur_eff   = self._preprocess_pdc(self.file3_path.get(), label="EUR")

        self._log("━" * 54, "section")
        self._log(" PRÉTRAITEMENT TERMINÉ", "section")
        self._log(f"  Diana : {os.path.basename(path_diana_eff)}", "ok")
        self._log(f"  JH    : {os.path.basename(path_jh_eff)}", "ok")
        self._log(f"  EUR   : {os.path.basename(path_eur_eff)}", "ok")
        self._log("━" * 54, "section")

        return path_diana_eff, path_jh_eff, path_eur_eff

    # ── Prétraitement Diana ───────────────────────────────────────────────
    def _preprocess_diana(self):
        """Génère (si besoin) les CSV Diana sans filtre et avec filtre.
        Retourne le chemin du CSV _avec_filtre.
        """
        path_src = self.file1_path.get()
        if not path_src or not os.path.isfile(path_src):
            raise FileNotFoundError(f"Fichier Diana introuvable : {path_src}")

        base_no_ext  = os.path.splitext(path_src)[0]
        path_csv_all = base_no_ext + ".csv"
        path_csv_flt = base_no_ext + "_avec_filtre.csv"

        # Si le CSV filtré existe déjà → réutilisation directe
        if os.path.isfile(path_csv_flt):
            self._log(f"  ✔  Diana CSV filtré existant réutilisé : "
                      f"{os.path.basename(path_csv_flt)}", "ok")
            return path_csv_flt

        # Lecture du fichier .txt source
        self._log(f"  ⏳  Lecture Diana source : {os.path.basename(path_src)}", "info")
        t0 = time.perf_counter()

        enc = "utf-8"
        for candidate in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
            try:
                with open(path_src, "r", encoding=candidate) as f:
                    f.read(8192)
                enc = candidate
                break
            except UnicodeDecodeError:
                continue

        with open(path_src, "r", encoding=enc, errors="replace") as fh:
            raw_lines = fh.readlines()

        if not raw_lines:
            raise ValueError("Le fichier Diana source est vide.")

        # En-tête : suppression du 1er caractère s'il n'est pas alphabétique
        header_raw = raw_lines[0].rstrip("\n\r")
        if header_raw and not header_raw[0].isalpha():
            header_raw = header_raw[1:]
        headers_src = [h.strip().strip('"') for h in header_raw.split(SEP_INPUT)]

        def _ci(name):
            """Index d'une colonne (insensible à la casse)."""
            nl = name.lower()
            for i, h in enumerate(headers_src):
                if h.lower() == nl:
                    return i
            return -1

        idx_pc   = _ci(PROJECT_CODE_FIELD)
        idx_uf   = _ci("User First Name")
        idx_ul   = _ci("User Last Name")

        # Indices des colonnes à conserver dans le CSV filtré
        # (DIANA_COLS_FILTRE, insensible à la casse)
        flt_indices = []
        flt_headers = []
        for col in DIANA_COLS_FILTRE:
            i = _ci(col)
            if i >= 0:
                flt_indices.append(i)
                flt_headers.append(headers_src[i])   # nom original du fichier source
            else:
                self._log(f"  ⚠  Diana : colonne '{col}' absente du fichier source — ignorée.", "warn")

        # ── Génération CSV sans filtre (toutes colonnes, toutes lignes) ──
        self._log(f"  ⏳  Génération CSV sans filtre : {os.path.basename(path_csv_all)}", "info")
        nb_all = 0
        with open(path_csv_all, "w", newline="", encoding="utf-8-sig") as fout:
            fout.write(SEP_OUTPUT.join(headers_src) + "\n")
            for raw in raw_lines[1:]:
                line = raw.rstrip("\n\r")
                if not line.strip():
                    continue
                fields = [f.strip().strip('"') for f in line.split(SEP_INPUT)]
                fout.write(SEP_OUTPUT.join(fields) + "\n")
                nb_all += 1
        self._log(f"  ✔  CSV sans filtre : {nb_all} lignes → {os.path.basename(path_csv_all)}", "ok")

        # ── Génération CSV avec filtre ────────────────────────────────────
        self._log(f"  ⏳  Génération CSV avec filtre : {os.path.basename(path_csv_flt)}", "info")
        nb_flt = 0
        with open(path_csv_flt, "w", newline="", encoding="utf-8-sig") as fout:
            fout.write(SEP_OUTPUT.join(flt_headers) + "\n")
            for raw in raw_lines[1:]:
                line = raw.rstrip("\n\r")
                if not line.strip():
                    continue
                fields = [f.strip().strip('"') for f in line.split(SEP_INPUT)]
                while len(fields) < len(headers_src):
                    fields.append("")

                pc_val     = fields[idx_pc]  if idx_pc >= 0  else ""
                u_orig     = (f"{fields[idx_ul]}, {fields[idx_uf]}"
                              if idx_ul >= 0 and idx_uf >= 0 else "")
                match_interv = _norm_name(u_orig) in self._allowed_norms if self._allowed_norms else False
                match_proj   = pc_val in TARGET_PROJECTS

                if not (match_interv or match_proj):
                    continue

                row_out = [fields[i] if i < len(fields) else "" for i in flt_indices]
                fout.write(SEP_OUTPUT.join(row_out) + "\n")
                nb_flt += 1

        self._log_duration("Prétraitement Diana", t0)
        self._log(f"  ✔  CSV filtré : {nb_flt} lignes → {os.path.basename(path_csv_flt)}", "ok")
        return path_csv_flt

    # ── Prétraitement PDC (JH ou EUR) ────────────────────────────────────
    def _preprocess_pdc(self, path_src, label):
        """Génère (si besoin) les CSV PDC sans filtre (si source XLSX) et avec filtre.
        Retourne le chemin du CSV _avec_filtre.
        """
        if not path_src or not os.path.isfile(path_src):
            raise FileNotFoundError(f"Fichier PDC {label} introuvable : {path_src}")

        ext_src      = os.path.splitext(path_src)[1].lower()
        base_no_ext  = os.path.splitext(path_src)[0]
        path_csv_all = base_no_ext + ".csv"          # sans filtre (XLSX seulement)
        path_csv_flt = base_no_ext + "_avec_filtre.csv"

        # Si le CSV filtré existe déjà → réutilisation directe
        if os.path.isfile(path_csv_flt):
            self._log(f"  ✔  PDC {label} CSV filtré existant réutilisé : "
                      f"{os.path.basename(path_csv_flt)}", "ok")
            return path_csv_flt

        if not HAS_PANDAS:
            raise RuntimeError("pandas est requis pour le prétraitement PDC : pip install pandas openpyxl")

        self._log(f"  ⏳  Lecture PDC {label} source : {os.path.basename(path_src)}", "info")
        t0 = time.perf_counter()

        # ── Lecture du fichier source ─────────────────────────────────────
        if ext_src == ".xlsx":
            df = pd.read_excel(path_src, dtype=str, header=3)
            df.columns = [c.strip() for c in df.columns]

            # Génération CSV sans filtre (toutes colonnes, toutes lignes)
            self._log(f"  ⏳  Génération CSV sans filtre PDC {label} : "
                      f"{os.path.basename(path_csv_all)}", "info")
            df.to_csv(path_csv_all, index=False, sep=SEP_OUTPUT, encoding="utf-8-sig")
            self._log(f"  ✔  CSV sans filtre PDC {label} : {len(df)} lignes → "
                      f"{os.path.basename(path_csv_all)}", "ok")

        else:  # .csv
            enc = "utf-8-sig"
            for candidate in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
                try:
                    with open(path_src, "r", encoding=candidate) as f:
                        sample = f.read(8192)
                    enc = candidate
                    break
                except UnicodeDecodeError:
                    continue
            else:
                with open(path_src, "r", encoding="latin-1", errors="replace") as f:
                    sample = f.read(8192)
                enc = "latin-1"

            sep_src = ";" if sample.count(";") >= sample.count(",") else ","

            # Détection position du header (ligne 1 ou 4)
            df_probe = pd.read_csv(path_src, dtype=str, sep=sep_src,
                                   encoding=enc, encoding_errors="replace",
                                   on_bad_lines="skip", nrows=5)
            df_probe.columns = [c.strip() for c in df_probe.columns]
            cols_needed = {COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS}
            skiprows = None if cols_needed.issubset(set(df_probe.columns)) else 3

            df = pd.read_csv(path_src, dtype=str, sep=sep_src,
                             encoding=enc, encoding_errors="replace",
                             skiprows=skiprows, on_bad_lines="skip")
            df.columns = [c.strip() for c in df.columns]
            # Pas de CSV "sans filtre" pour les sources déjà en CSV

        df = df.fillna("")

        # ── Normalisation des noms de colonnes pour le filtre ─────────────
        def _norm_col(s):
            return s.lower().replace(" ", "").replace("_", "")

        col_map = {_norm_col(c): c for c in df.columns}

        def _find(name):
            if name in df.columns:
                return name
            return col_map.get(_norm_col(name))

        interv_col = _find(COL_INTERVENANT)
        proj_col   = _find(COL_PROJET)

        if not interv_col:
            self._log(f"  ⚠  PDC {label} : colonne '{COL_INTERVENANT}' introuvable — "
                      f"filtre intervenant désactivé.", "warn")
        if not proj_col:
            self._log(f"  ⚠  PDC {label} : colonne '{COL_PROJET}' introuvable — "
                      f"filtre projet désactivé.", "warn")

        # ── Colonnes à conserver dans le CSV filtré ───────────────────────
        keep_cols = []
        for col in PDC_COLS_FILTRE_MANDATORY:
            real = _find(col)
            if real:
                keep_cols.append(real)
            else:
                self._log(f"  ⚠  PDC {label} : colonne obligatoire '{col}' absente.", "warn")
        for col in PDC_COLS_FILTRE_OPTIONAL:
            real = _find(col)
            if real and real not in keep_cols:
                keep_cols.append(real)

        # ── Application du filtre ─────────────────────────────────────────
        mask = pd.Series([False] * len(df), index=df.index)
        if interv_col:
            mask |= df[interv_col].apply(
                lambda v: _norm_name(str(v).strip()) in self._allowed_norms
            )
        if proj_col:
            mask |= df[proj_col].astype(str).str.strip().isin(TARGET_PROJECTS)

        df_flt = df[mask][keep_cols]

        # ── Écriture CSV filtré ───────────────────────────────────────────
        df_flt.to_csv(path_csv_flt, index=False, sep=SEP_OUTPUT, encoding="utf-8-sig")
        self._log_duration(f"Prétraitement PDC {label}", t0)
        self._log(f"  ✔  CSV filtré PDC {label} : {len(df_flt)} lignes / "
                  f"{len(df)} total → {os.path.basename(path_csv_flt)}", "ok")

        return path_csv_flt

    # ══════════════════════════════════════════════════════════════════════
    #  TRAITEMENT 1 : Suivi Conso
    # ══════════════════════════════════════════════════════════════════════
    def _run_suivi_conso(self):
        if not self.file1_path.get():
            messagebox.showwarning("Fichier manquant",
                                   "Veuillez sélectionner le fichier n°1 (saisies .txt).")
            return
        p2 = self.file2_path.get()
        p3 = self.file3_path.get()
        if p2 and p3:
            ext2 = os.path.splitext(p2)[1].lower()
            ext3 = os.path.splitext(p3)[1].lower()
            if ext2 != ext3:
                messagebox.showerror(
                    "Formats incompatibles",
                    "Les fichiers 2 et 3 doivent être dans le même format\n"
                    "(tous les deux .xlsx  OU  tous les deux .csv).\n\n"
                    f"Fichier 2 : {ext2.upper()}\n"
                    f"Fichier 3 : {ext3.upper()}")
                return
        self.btn_suivi.config(state="disabled")
        threading.Thread(target=self._process_file1, daemon=True).start()

    def _process_file1(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))

        # Fichiers de sortie
        ts_run    = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_name = f"suivi_conso_nucleus_{ts_run}.xlsx"
        path_xlsx = os.path.join(script_dir, xlsx_name)

        # Fichier de transcodification Stream
        path_transco   = os.path.join(script_dir, "transco_Stream.xlsx")
        # Fichier Histo_TJM
        path_histo_tjm = os.path.join(script_dir, "Histo_TJM.xlsx")

        self._log("━" * 54, "section")
        self._log(" TRAITEMENT SUIVI CONSO — démarrage", "section")
        self._log("━" * 54, "section")
        self._set_status("Prétraitement des fichiers source…")
        t_global = time.perf_counter()

        try:
            if not HAS_OPENPYXL:
                raise RuntimeError("openpyxl est requis. Installez-le : pip install openpyxl")

            # ══════════════════════════════════════════════════════════════
            # ÉTAPE 0 : Prétraitement — génération / réutilisation CSV filtrés
            # ══════════════════════════════════════════════════════════════
            path_in, path_pdc_jh, path_pdc_eur = self._preprocess_sources()

            # Recalcul du nom CSV de sortie basé sur le fichier source original
            path_in_orig = self.file1_path.get()
            base_name    = os.path.splitext(os.path.basename(path_in_orig))[0]
            path_csv     = os.path.join(script_dir, base_name + "_suivi_conso.csv")

            self._log(f"  Source Diana  : {os.path.basename(path_in)}", "info")
            self._log(f"  Source PDC JH : {os.path.basename(path_pdc_jh)}", "info")
            self._log(f"  Source PDC EUR: {os.path.basename(path_pdc_eur)}", "info")
            self._log(f"  Sortie XLSX   : {xlsx_name}", "info")
            self._log(f"  Sortie CSV    : {os.path.basename(path_csv)}", "info")
            self._set_status("Traitement en cours…")

            # ── Chargement de la table de transcodification Stream ──────────
            t_step = time.perf_counter()
            transco_stream = {}
            if os.path.isfile(path_transco):
                wb_tr = openpyxl.load_workbook(path_transco, read_only=True, data_only=True)
                if "Transco_Stream" in wb_tr.sheetnames:
                    ws_tr = wb_tr["Transco_Stream"]
                    for i, row in enumerate(ws_tr.iter_rows(values_only=True)):
                        if i == 0:
                            continue
                        if row[0] is not None:
                            wpn_key    = str(row[0]).strip()
                            task_key   = str(row[1]).strip() if row[1] is not None else ""
                            stream_val = str(row[2]).strip() if row[2] is not None else "NT"
                            transco_stream[(wpn_key, task_key)] = stream_val
                    wb_tr.close()
                    self._log(f"  Transco Stream chargé : {len(transco_stream)} entrée(s).", "ok")
                    self._log_duration("Chargement Transco Stream", t_step)
                else:
                    self._log("  ⚠  Onglet 'Transco_Stream' introuvable dans transco_Stream.xlsx.", "warn")
            else:
                self._log(f"  ⚠  Fichier transco_Stream.xlsx absent ({path_transco}). "
                          f"Stream = 'NT' pour toutes les lignes.", "warn")

            # ── Chargement de Histo_TJM.xlsx ────────────────────────────────
            t_step = time.perf_counter()
            histo_tjm     = {}
            histo_tjm_raw = {}
            if os.path.isfile(path_histo_tjm):
                wb_tjm = openpyxl.load_workbook(path_histo_tjm, read_only=True, data_only=True)
                ws_tjm = wb_tjm.active
                nb_tjm = 0
                for i, row in enumerate(ws_tjm.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    if row[0] is None:
                        continue
                    raw_tjm  = str(row[0]).strip()
                    username = _norm_name(raw_tjm)
                    histo_tjm_raw[username] = raw_tjm
                    for mois_num in range(1, 13):
                        val = row[mois_num] if len(row) > mois_num else None
                        if val is not None:
                            try:
                                histo_tjm[(username, mois_num)] = float(val)
                                nb_tjm += 1
                            except (ValueError, TypeError):
                                pass
                wb_tjm.close()
                self._log(f"  Histo_TJM chargé : {nb_tjm} valeur(s) TJM.", "ok")
                self._log_duration("Chargement Histo_TJM", t_step)
            else:
                self._log(f"  ⚠  Histo_TJM.xlsx absent ({path_histo_tjm}). "
                          f"conso_diana_ttnr ne pourra pas être calculé.", "warn")

            # ── Chargement des plans de charge JH et EUR ────────────────────
            # Désormais on passe les chemins CSV filtrés retournés par _preprocess_sources
            def _load_pdc(path, label):
                """Lit un fichier plan de charge CSV filtré,
                   retourne dict (intervenant, mois_num, project_code)→float.
                   Le CSV filtré a toujours le header en ligne 1 (sep=;, utf-8-sig).
                """
                result = {}
                if not path:
                    self._log(f"  ⚠  Fichier {label} non sélectionné — conso PDC {label} absente.", "warn")
                    return result
                if not HAS_PANDAS:
                    self._log(f"  ⚠  pandas requis pour lire {label}.", "warn")
                    return result
                try:
                    df = pd.read_csv(path, dtype=str, sep=SEP_OUTPUT,
                                     encoding="utf-8-sig", encoding_errors="replace",
                                     on_bad_lines="skip")
                    df.columns = [c.strip() for c in df.columns]

                    for col in [COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS]:
                        if col not in df.columns:
                            self._log(f"  ⚠  Colonne '{col}' absente dans {label}.", "warn")
                            return result
                    df[COL_INTERVENANT] = df[COL_INTERVENANT].fillna("").str.strip()
                    df[COL_MOIS]        = df[COL_MOIS].fillna("").str.strip()
                    df[COL_PROJET]      = df[COL_PROJET].fillna("").str.strip()
                    df[COL_NBJOURS]     = (df[COL_NBJOURS].astype(str)
                                           .str.replace(",", ".", regex=False).str.strip())
                    df = df[df[COL_INTERVENANT] != ""]
                    df[COL_NBJOURS] = pd.to_numeric(df[COL_NBJOURS], errors="coerce").fillna(0)

                    def _norm_mois(val):
                        v = str(val).strip().lower()
                        try:
                            return MOIS_ORDRE.get(int(float(v)))
                        except (ValueError, TypeError):
                            pass
                        return MOIS_ORDRE.get(v)

                    df["mois_num"] = df[COL_MOIS].apply(_norm_mois)
                    df = df.dropna(subset=["mois_num"])
                    df["mois_num"] = df["mois_num"].astype(int)

                    for _, row in df.iterrows():
                        key = (_norm_name(str(row[COL_INTERVENANT]).strip()),
                               int(row["mois_num"]),
                               str(row[COL_PROJET]).strip())
                        result[key] = result.get(key, 0.0) + float(row[COL_NBJOURS])

                    self._log(f"  PDC {label} chargé : {len(result)} combinaison(s) "
                              f"(intervenant, mois, projet).", "ok")
                except Exception as e:
                    self._log(f"  ⚠  Erreur lecture {label} : {e}", "warn")
                return result

            self._log(f"  ⏳  Chargement PDC JH en cours...", "info")
            self._set_status("Chargement PDC JH en cours...")
            t_step = time.perf_counter()
            pdc_jh  = _load_pdc(path_pdc_jh,  "JH")
            self._log_duration("Chargement PDC JH", t_step)
            self._log(f"  ⏳  Chargement PDC EUR en cours...", "info")
            self._set_status("Chargement PDC EUR en cours...")
            t_step = time.perf_counter()
            pdc_eur = _load_pdc(path_pdc_eur, "EUR")
            self._log_duration("Chargement PDC EUR", t_step)

            # ── Lecture du fichier Diana CSV filtré ──────────────────────────
            # Le CSV filtré a sep=; et encodage utf-8-sig — lecture simple
            t_step = time.perf_counter()
            with open(path_in, "r", encoding="utf-8-sig", errors="replace") as fh:
                raw_lines = fh.readlines()

            if not raw_lines:
                raise ValueError("Le fichier Diana CSV filtré est vide.")

            headers = [h.strip().strip('"') for h in raw_lines[0].rstrip("\n\r").split(SEP_OUTPUT)]
            self._log(f"  En-tête Diana CSV filtré : {len(headers)} colonnes détectées.", "info")

            def col_index(name):
                nl = name.lower()
                for i, h in enumerate(headers):
                    if h.lower() == nl:
                        return i
                return -1

            idx_pc = col_index(PROJECT_CODE_FIELD)
            if idx_pc < 0:
                self._log(f"  ⚠  Colonne '{PROJECT_CODE_FIELD}' introuvable — filtre codes projet désactivé.", "warn")
            if not self._allowed_norms:
                self._log("  ⚠  Fichier intervenants non chargé (fichier 4) — filtre périmètre désactivé.", "warn")

            # Index des colonnes sources
            SRC_COLS = [
                "Project Code", "Project Is Alive", "WorkPackage Name", "Task Name",
                "Organization Unit Name", "User First Name", "User Last Name",
                "User Country Name", "User Contract Type", "TimeSheet Mode Name",
                "Timesheet Entry Details Calendar Date", "Status", "Is Published",
                "Timesheet Entry Details Value", "Application Cluster Name",
            ]
            IDX_PROJECT_CODE     = col_index("Project Code")
            IDX_PROJECT_NAME_SRC = col_index("Project Name")
            IDX_WPN              = col_index("WorkPackage Name")
            IDX_TASK             = col_index("Task Name")
            IDX_UFIRST           = col_index("User First Name")
            IDX_ULAST            = col_index("User Last Name")
            IDX_DATE             = col_index("Timesheet Entry Details Calendar Date")

            src_indices = [col_index(c) for c in SRC_COLS]
            missing_src = [SRC_COLS[i] for i, idx in enumerate(src_indices) if idx < 0]
            if missing_src:
                self._log(f"  ⚠  Colonnes sources introuvables : {missing_src}", "warn")

            SRC_DIMS = [c for c in SRC_COLS
                        if c not in ("Timesheet Entry Details Calendar Date",
                                     "Timesheet Entry Details Value",
                                     "User First Name", "User Last Name")]
            OUT_HEADERS = SRC_DIMS + [
                "ProjectName", "UserName", "annee", "dianamonth", "stream",
                "conso_diana_jh", "TJM", "conso_diana_ttnr",
            ]
            NB_MEASURES = 3

            # ── Parcours des lignes Diana CSV filtré ─────────────────────────
            # Le fichier est déjà filtré (intervenant périmètre OU projet ciblé)
            # → on parcourt toutes les lignes sans re-filtrer
            kept_rows  = []
            total_data = 0
            nt_wpn     = set()

            for raw in raw_lines[1:]:
                line = raw.rstrip("\n\r")
                if not line.strip():
                    continue
                total_data += 1
                fields = [f.strip().strip('"') for f in line.split(SEP_OUTPUT)]

                while len(fields) < len(headers):
                    fields.append("")

                def fv(idx):
                    return fields[idx] if 0 <= idx < len(fields) else ""

                idx_date_src   = SRC_COLS.index("Timesheet Entry Details Calendar Date")
                idx_value_src  = SRC_COLS.index("Timesheet Entry Details Value")
                idx_ufirst_src = SRC_COLS.index("User First Name")
                idx_ulast_src  = SRC_COLS.index("User Last Name")
                skip_src = {idx_date_src, idx_value_src, idx_ufirst_src, idx_ulast_src}

                out_row = []
                for pos, idx in enumerate(src_indices):
                    if pos in skip_src:
                        continue
                    out_row.append(fv(idx) if idx >= 0 else "")

                proj_code_val     = fv(IDX_PROJECT_CODE)
                proj_name_src     = fv(IDX_PROJECT_NAME_SRC)
                calc_project_name = (f"{proj_code_val} - {proj_name_src}"
                                     if proj_name_src else proj_code_val)

                calc_username     = f"{fv(IDX_ULAST)}, {fv(IDX_UFIRST)}"
                calc_username_key = _norm_name(calc_username)

                date_val   = fv(IDX_DATE)
                calc_annee = ""
                calc_mois  = ""
                if date_val and len(date_val) >= 7:
                    parts = date_val.split("-")
                    if len(parts) >= 2:
                        calc_annee = parts[0]
                        calc_mois  = parts[1]

                calc_dianamonth = (f"{calc_annee} - {calc_mois}"
                                   if calc_annee and calc_mois else "")

                proj_code_for_stream = fv(IDX_PROJECT_CODE)
                if proj_code_for_stream != "P02508":
                    calc_stream = "NA"
                else:
                    wpn_val  = fv(IDX_WPN).strip()
                    task_val = fv(IDX_TASK).strip()
                    calc_stream = transco_stream.get((wpn_val, task_val), "NT")
                    if calc_stream == "NT":
                        nt_wpn.add(f"{wpn_val} | {task_val}")

                raw_value = fv(src_indices[idx_value_src] if src_indices[idx_value_src] >= 0 else -1)
                try:
                    num_value = float(raw_value.replace(",", ".")) if raw_value else 0.0
                except ValueError:
                    num_value = 0.0

                out_row += [calc_project_name, calc_username,
                            calc_annee, calc_dianamonth, calc_stream]

                key_row = out_row[:-5] + [calc_project_name, calc_username_key,
                                          calc_annee, calc_dianamonth, calc_stream]
                dim_key     = tuple(key_row)
                wpn_for_anom = fv(IDX_WPN)
                kept_rows.append((dim_key, num_value, calc_username,
                                   calc_project_name, wpn_for_anom))

            self._log(f"  Lignes lues    : {total_data}", "info")
            self._log(f"  Lignes détail retenues : {len(kept_rows)}", "ok")
            self._log_duration("Lecture & filtrage fichier source", t_step)

            if nt_wpn:
                self._log(f"  ⚠  {len(nt_wpn)} paire(s) (WPN, Task) non transcodée(s) → Stream='NT' :", "warn")
                for w in sorted(nt_wpn):
                    self._log(f"      • {w}", "warn")

            # ── Agrégation ───────────────────────────────────────────────────
            t_step = time.perf_counter()
            from collections import OrderedDict
            agg          = OrderedDict()
            agg_dispname = {}
            agg_proj_wpn = {}
            for dim_key, num_value, username_display, proj_name, wpn in kept_rows:
                if dim_key in agg:
                    agg[dim_key] += num_value
                else:
                    agg[dim_key] = num_value
                    agg_dispname[dim_key] = username_display
                if dim_key not in agg_proj_wpn:
                    agg_proj_wpn[dim_key] = set()
                agg_proj_wpn[dim_key].add((proj_name, wpn))

            IDX_USERNAME_IN_KEY  = len(SRC_DIMS) + 1
            IDX_DIANA_IN_KEY_AGG = len(SRC_DIMS) + 3

            tjm_not_found = {}
            agg_rows = []

            for dim_key, conso_jh in agg.items():
                username_norm = dim_key[IDX_USERNAME_IN_KEY]
                username_disp = agg_dispname[dim_key]
                diana_str     = dim_key[IDX_DIANA_IN_KEY_AGG]
                try:
                    mois_num = int(str(diana_str).split("-")[-1].strip())
                except (ValueError, TypeError, IndexError):
                    mois_num = None

                tjm = histo_tjm.get((username_norm, mois_num)) if mois_num else None
                if tjm is None and mois_num is not None:
                    anom_key     = (username_disp, mois_num)
                    proj_wpn_set = agg_proj_wpn.get(dim_key, set())
                    if anom_key not in tjm_not_found:
                        tjm_not_found[anom_key] = []
                    for pn, wpn in sorted(proj_wpn_set):
                        tjm_not_found[anom_key].append((pn, wpn, conso_jh))

                conso_ttnr = round(conso_jh * tjm, 2) if tjm is not None else ""

                disp_row = list(dim_key)
                disp_row[IDX_USERNAME_IN_KEY] = username_disp
                agg_rows.append(
                    disp_row + [
                        conso_jh,
                        round(tjm, 4) if tjm is not None else "",
                        conso_ttnr,
                    ]
                )

            inco_names = {}
            for dim_key in agg.keys():
                u_norm = dim_key[IDX_USERNAME_IN_KEY]
                u_disp = agg_dispname[dim_key]
                if u_norm in histo_tjm_raw:
                    raw_tjm = histo_tjm_raw[u_norm]
                    if u_disp != raw_tjm and u_norm not in inco_names:
                        inco_names[u_norm] = (u_disp, raw_tjm)
            self._update_inco_box(inco_names)

            self._log(f"  Lignes agrégées (dianamonth) : {len(agg_rows)}", "ok")
            self._log_duration("Agrégation onglet Diana_Task", t_step)
            if tjm_not_found:
                self._log(f"  ⚠  {len(tjm_not_found)} combinaison(s) (UserName, mois) "
                          f"sans TJM dans Histo_TJM → conso_diana_ttnr vide :", "warn")
                for (username, mois_num), details in sorted(tjm_not_found.items()):
                    self._log(f"      • {username} / mois {mois_num:02d}", "warn")
                    for pn, wpn, cjh in details:
                        self._log(f"          ↳ {pn} | {wpn} ({cjh:.4f} jh)", "warn")

            # ── Écriture XLSX ────────────────────────────────────────────────
            t_step = time.perf_counter()
            from openpyxl.utils import get_column_letter

            wb_out = openpyxl.Workbook()

            hdr_font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            hdr_fill_src  = PatternFill("solid", start_color="1F4E79")
            hdr_fill_calc = PatternFill("solid", start_color="375623")
            alt_fill      = PatternFill("solid", start_color="EBF3FB")
            alt_calc      = PatternFill("solid", start_color="EBF7EB")
            thin          = Side(style="thin", color="D0D0D0")
            brd           = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr           = Alignment(horizontal="center", vertical="center")
            lft           = Alignment(horizontal="left",   vertical="center")
            rgt           = Alignment(horizontal="right",  vertical="center")

            col_widths = {
                "Project Code": 14, "Project Is Alive": 16, "WorkPackage Name": 45,
                "Task Name": 35, "Organization Unit Name": 30, "User First Name": 18,
                "User Last Name": 20, "User Country Name": 16, "User Contract Type": 18,
                "TimeSheet Mode Name": 24, "Status": 12, "Is Published": 14,
                "Application Cluster Name": 28,
                "ProjectName": 40, "UserName": 30,
                "annee": 8, "mois": 6, "dianamonth": 14, "stream": 18,
                "conso_diana_jh": 22, "TJM": 14, "conso_diana_ttnr": 22,
            }

            def _write_tcd(ws, data, all_cols, row_label, measure_label):
                from openpyxl.utils import get_column_letter as gcl
                hf  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                hfb = PatternFill("solid", start_color="1F4E79")
                hfc = PatternFill("solid", start_color="375623")
                af  = PatternFill("solid", start_color="EBF3FB")
                ac  = PatternFill("solid", start_color="EBF7EB")
                thn = Side(style="thin", color="D0D0D0")
                b   = Border(left=thn, right=thn, top=thn, bottom=thn)
                ctr = Alignment(horizontal="center", vertical="center")
                lft = Alignment(horizontal="left",   vertical="center")
                rgt = Alignment(horizontal="right",  vertical="center")
                num_fmt = "#,##0.00"

                c = ws.cell(1, 1, f"{row_label} / {measure_label}")
                c.font = hf; c.fill = hfb; c.border = b; c.alignment = ctr
                ws.column_dimensions["A"].width = 42
                for ci, dm in enumerate(all_cols, start=2):
                    c = ws.cell(1, ci, dm)
                    c.font = hf; c.fill = hfc; c.border = b; c.alignment = ctr
                    ws.column_dimensions[gcl(ci)].width = 14
                tot_ci = len(all_cols) + 2
                c = ws.cell(1, tot_ci, "TOTAL")
                c.font = hf; c.fill = hfb; c.border = b; c.alignment = ctr
                ws.column_dimensions[gcl(tot_ci)].width = 14

                row_totals = {dm: 0.0 for dm in all_cols}
                for ri, (row_key, dm_dict) in enumerate(data.items(), start=2):
                    use_alt = (ri % 2 == 0)
                    c = ws.cell(ri, 1, row_key)
                    c.font = Font(name="Arial", size=10); c.border = b; c.alignment = lft
                    if use_alt: c.fill = af
                    line_tot = 0.0
                    for ci, dm in enumerate(all_cols, start=2):
                        val = dm_dict.get(dm, 0.0)
                        cell = ws.cell(ri, ci, round(val, 2) if val else "")
                        cell.font = Font(name="Arial", size=10); cell.border = b
                        cell.alignment = rgt
                        if use_alt: cell.fill = ac
                        if val:
                            cell.number_format = num_fmt
                            line_tot += val
                            row_totals[dm] += val
                    tc = ws.cell(ri, tot_ci, round(line_tot, 2) if line_tot else "")
                    tc.font = Font(name="Arial", bold=True, size=10); tc.border = b
                    tc.alignment = rgt
                    if line_tot: tc.number_format = num_fmt
                    if use_alt: tc.fill = af

                tot_ri = len(data) + 2
                c = ws.cell(tot_ri, 1, "TOTAL")
                c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill = hfb; c.border = b; c.alignment = lft
                grand_tot = 0.0
                for ci, dm in enumerate(all_cols, start=2):
                    val = row_totals[dm]
                    cell = ws.cell(tot_ri, ci, round(val, 2) if val else "")
                    cell.font = Font(name="Arial", bold=True, size=10)
                    cell.fill = hfb; cell.border = b; cell.alignment = rgt
                    if val:
                        cell.number_format = num_fmt
                        grand_tot += val
                gc = ws.cell(tot_ri, tot_ci, round(grand_tot, 2) if grand_tot else "")
                gc.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                gc.fill = hfb; gc.border = b; gc.alignment = rgt
                if grand_tot: gc.number_format = num_fmt
                ws.freeze_panes = "B2"

            def _write_sheet(ws, headers, rows, nb_src_cols, nb_measures):
                idx_first_measure = len(headers) - nb_measures + 1
                for ci, h in enumerate(headers, start=1):
                    c = ws.cell(1, ci, h)
                    c.font      = hdr_font
                    c.fill      = hdr_fill_calc if ci > nb_src_cols else hdr_fill_src
                    c.border    = brd
                    c.alignment = ctr
                    ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 16)
                for ri, row_data in enumerate(rows, start=2):
                    use_alt = (ri % 2 == 0)
                    for ci, val in enumerate(row_data, start=1):
                        is_measure = (ci >= idx_first_measure)
                        if is_measure and val != "":
                            c = ws.cell(ri, ci,
                                        round(val, 4) if isinstance(val, float) else val)
                            c.number_format = "#,##0.0000"
                            c.alignment = rgt
                        else:
                            c = ws.cell(ri, ci, val)
                            c.alignment = lft
                        c.font   = Font(name="Arial", size=10)
                        c.border = brd
                        if use_alt:
                            c.fill = alt_calc if ci > (len(headers) - nb_measures) else alt_fill
                ws.freeze_panes = "A2"

            # Onglet 1 : Diana_Task
            ws1 = wb_out.active
            ws1.title = "Diana_Task"
            _write_sheet(ws1, OUT_HEADERS, agg_rows,
                         nb_src_cols=len(SRC_DIMS), nb_measures=NB_MEASURES)
            self._log(f"  ✔  Onglet 'Diana_Task' : {len(agg_rows)} lignes", "ok")

            # Onglet 2 : Diana_Projet
            EXCL_ONG2 = {"WorkPackage Name", "Task Name", "stream", "Application Cluster Name", "Is Published"}
            keep_idx2 = [i for i, h in enumerate(OUT_HEADERS) if h not in EXCL_ONG2]

            idx_val_oh   = OUT_HEADERS.index("conso_diana_jh")
            idx_ttnr_oh  = OUT_HEADERS.index("conso_diana_ttnr")
            idx_tjm_oh   = OUT_HEADERS.index("TJM")
            IDX_PC_IN_KEY    = OUT_HEADERS.index("Project Code")
            IDX_USER_IN_KEY  = OUT_HEADERS.index("UserName")
            IDX_DIANA_IN_KEY = OUT_HEADERS.index("dianamonth")
            measure_idx_set  = {idx_val_oh, idx_ttnr_oh, idx_tjm_oh}

            from collections import OrderedDict as OD2
            agg2_diana = OD2()
            for row in agg_rows:
                dim2 = tuple(row[i] for i in keep_idx2 if i not in measure_idx_set)
                val_jh   = row[idx_val_oh]  if row[idx_val_oh]  != "" else 0.0
                val_ttnr = row[idx_ttnr_oh] if row[idx_ttnr_oh] != "" else 0.0
                tjm_row  = row[idx_tjm_oh]  if row[idx_tjm_oh]  != "" else None
                if dim2 in agg2_diana:
                    agg2_diana[dim2][0] += val_jh
                    agg2_diana[dim2][1] += val_ttnr
                    if tjm_row is not None:
                        agg2_diana[dim2][2] = max(agg2_diana[dim2][2] or 0, tjm_row)
                else:
                    agg2_diana[dim2] = [val_jh, val_ttnr, tjm_row]

            keep_idx2_dims = [i for i in keep_idx2 if i not in measure_idx_set]
            pos_pc    = keep_idx2_dims.index(IDX_PC_IN_KEY)
            pos_user  = keep_idx2_dims.index(IDX_USER_IN_KEY)
            pos_diana = keep_idx2_dims.index(IDX_DIANA_IN_KEY)

            def _fmt(v):
                if v is None or v == 0.0:
                    return ""
                return round(v, 4) if isinstance(v, float) else v

            agg_rows2 = []
            for dim2, (val_jh, val_ttnr, tjm_val) in agg2_diana.items():
                username_disp2 = dim2[pos_user]
                username_key2  = _norm_name(username_disp2)
                proj_code_r    = dim2[pos_pc]
                diana_val      = dim2[pos_diana]
                try:
                    mois_num_r = int(str(diana_val).split("-")[-1].strip())
                except (ValueError, TypeError, IndexError):
                    mois_num_r = None
                pdc_key2    = (username_key2, mois_num_r, proj_code_r) if mois_num_r else None
                val_pdc_jh  = pdc_jh.get(pdc_key2,  0.0) if pdc_key2 else 0.0
                val_pdc_eur = pdc_eur.get(pdc_key2, 0.0) if pdc_key2 else 0.0

                agg_rows2.append(
                    list(dim2) + [
                        _fmt(val_jh), _fmt(val_ttnr),
                        _fmt(val_pdc_jh), _fmt(val_pdc_eur),
                    ]
                )

            NB_MEAS2   = 5
            HDR2_ALL   = [OUT_HEADERS[i] for i in keep_idx2]
            HDR2_DIMS  = [h for h in HDR2_ALL if h not in ("conso_diana_jh",
                                                             "conso_diana_ttnr", "TJM")]
            HDR2_FINAL = HDR2_DIMS + [
                "conso_diana_jh", "conso_diana_ttnr",
                "conso_pdc_jh", "conso_pdc_eur", "ecart_jh",
            ]

            agg_rows2 = [
                row + [_fmt(
                    (row[-2] if row[-2] != "" else 0.0) - (row[-4] if row[-4] != "" else 0.0)
                )]
                for row in agg_rows2
            ]

            col_widths.update({"conso_pdc_jh": 18, "conso_pdc_eur": 18, "ecart_jh": 14})

            ws2 = wb_out.create_sheet("Diana_Projet")
            _write_sheet(ws2, HDR2_FINAL, agg_rows2,
                         nb_src_cols=len([h for h in HDR2_DIMS if h in SRC_DIMS]),
                         nb_measures=NB_MEAS2)
            self._log(f"  ✔  Onglet 'Diana_Projet' : {len(agg_rows2)} lignes", "ok")

            saved_path = path_xlsx
            wb_out.save(saved_path)
            self._log(f"  ✔  {xlsx_name} enregistré.", "ok")
            self._log_duration("Écriture fichier nucleus", t_step)

            # TCD natifs via win32com
            self._log("  ⏳  Création des TCD natifs Excel en cours...", "info")
            self._set_status("Création des TCD Excel...")
            t_step = time.perf_counter()
            tcd_ok = self._create_pivot_tables(saved_path, HDR2_FINAL, OUT_HEADERS)
            if tcd_ok:
                self._log_duration("Création TCD natifs", t_step)
            else:
                self._log("  ⚠  TCD natifs non créés (win32com indisponible). "
                          "Installez pywin32 : pip install pywin32", "warn")

            # CSV brut complet (toutes lignes du fichier Diana original)
            # On relit le fichier .txt original pour ce CSV exhaustif
            t_step = time.perf_counter()
            path_in_orig = self.file1_path.get()
            with open(path_in_orig, "r", encoding="utf-8", errors="replace") as fh:
                raw_lines_orig = fh.readlines()
            header_raw_orig = raw_lines_orig[0].rstrip("\n\r")
            if header_raw_orig and not header_raw_orig[0].isalpha():
                header_raw_orig = header_raw_orig[1:]
            headers_orig = [h.strip().strip('"') for h in header_raw_orig.split(SEP_INPUT)]
            with open(path_csv, "w", newline="", encoding="utf-8-sig") as fh:
                fh.write(SEP_OUTPUT.join(headers_orig) + "\n")
                nb_csv = 0
                for raw in raw_lines_orig[1:]:
                    line = raw.rstrip("\n\r")
                    if not line.strip():
                        continue
                    fields_all = [f.strip().strip('"') for f in line.split(SEP_INPUT)]
                    fh.write(SEP_OUTPUT.join(fields_all) + "\n")
                    nb_csv += 1
            self._log(f"  ✔  {os.path.basename(path_csv)} créé ({nb_csv} lignes)", "ok")
            self._log_duration("Écriture CSV brut complet", t_step)

            self._log(f"  Intervenants périmètre chargés : {len(self._allowed_list)}", "info")

            self._log_duration("Traitement total Suivi Conso", t_global, tag="section")
            self._log("━" * 54, "section")
            self._log(" TRAITEMENT TERMINÉ AVEC SUCCÈS ✔", "ok")
            self._log("━" * 54, "section")
            self._set_status(
                f"Terminé — {len(agg_rows)} lignes (Task) | {len(agg_rows2)} lignes (Projet) | "
                f"{total_data} lignes CSV")
            self.after(100, lambda: messagebox.showinfo(
                "Traitement terminé",
                f"Fichiers générés avec succès !\n\n"
                f"XLSX : {xlsx_name}\n"
                f"  Onglet 'Diana_Task'   : {len(agg_rows)} lignes\n"
                f"  Onglet 'Diana_Projet' : {len(agg_rows2)} lignes\n"
                f"  {len(nt_wpn)} paire(s) (WPN, Task) non transcodée(s)\n\n"
                f"CSV brut : {os.path.basename(path_csv)}\n"
                f"  → {total_data} lignes\n\n"
                f"Répertoire : {script_dir}"))

        except Exception as exc:
            import traceback
            self._log(f"✘ ERREUR : {exc}", "error")
            self._log(traceback.format_exc(), "error")
            self._set_status(f"Erreur : {exc}")
            self.after(100, lambda: messagebox.showerror(
                "Erreur de traitement", str(exc)))
        finally:
            self.after(100, lambda: self.btn_suivi.config(state="normal"))

    # ── TRAITEMENT 2 : Histo TJM ──────────────────────────────────────────
    def _run_histo_tjm(self):
        if not self.file2_path.get() or not self.file3_path.get():
            messagebox.showwarning(
                "Fichiers manquants",
                "Veuillez sélectionner les fichiers n°2 (JH) et n°3 (€).")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("Dépendance manquante",
                                  "Installez openpyxl :  pip install openpyxl")
            return
        if not HAS_PANDAS:
            messagebox.showerror("Dépendance manquante",
                                  "Installez pandas :  pip install pandas openpyxl")
            return
        self.btn_tjm.config(state="disabled")
        threading.Thread(target=self._process_histo_tjm, daemon=True).start()

    def _process_histo_tjm(self):
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()

        script_dir = os.path.dirname(os.path.abspath(__file__))
        path_out   = os.path.join(script_dir, "Histo_TJM.xlsx")

        t_global_tjm = time.perf_counter()
        self._log("━" * 54, "section")
        self._log(" GÉNÉRATION HISTO_TJM — démarrage", "section")
        self._log("━" * 54, "section")
        self._log(f"  Fichier JH  : {os.path.basename(path_jh)}", "info")
        self._log(f"  Fichier EUR : {os.path.basename(path_eur)}", "info")
        self._log(f"  Sortie      : {path_out}", "info")
        self._set_status("Calcul des TJM en cours…")

        try:
            t_step = time.perf_counter()
            df_jh  = pd.read_excel(path_jh,  dtype=str, header=3)
            self._log_duration("Lecture fichier JH", t_step)
            t_step = time.perf_counter()
            df_eur = pd.read_excel(path_eur, dtype=str, header=3)
            self._log_duration("Lecture fichier EUR", t_step)
            df_jh.columns  = [c.strip() for c in df_jh.columns]
            df_eur.columns = [c.strip() for c in df_eur.columns]

            self._log(f"  JH  : {len(df_jh)} lignes | "
                      f"colonnes : {list(df_jh.columns)}", "info")
            self._log(f"  EUR : {len(df_eur)} lignes | "
                      f"colonnes : {list(df_eur.columns)}", "info")

            required = [COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS]
            for label, df in [("JH (fichier 2)", df_jh), ("EUR (fichier 3)", df_eur)]:
                missing = [c for c in required if c not in df.columns]
                if missing:
                    raise ValueError(
                        f"Fichier {label} — colonnes manquantes : {missing}\n"
                        f"Colonnes présentes : {list(df.columns)}")

            def clean_df(df):
                df = df[[COL_INTERVENANT, COL_MOIS, COL_PROJET, COL_NBJOURS]].copy()
                df[COL_INTERVENANT] = df[COL_INTERVENANT].fillna("").str.strip()
                df[COL_MOIS]        = df[COL_MOIS].fillna("").str.strip()
                df[COL_PROJET]      = df[COL_PROJET].fillna("").str.strip()
                df[COL_NBJOURS]     = (df[COL_NBJOURS].astype(str)
                                       .str.replace(",", ".", regex=False)
                                       .str.strip())
                df = df[df[COL_INTERVENANT] != ""]
                df = df[df[COL_NBJOURS]     != ""]
                df[COL_NBJOURS] = pd.to_numeric(df[COL_NBJOURS], errors="coerce")
                return df.dropna(subset=[COL_NBJOURS])

            df_jh  = clean_df(df_jh)
            df_eur = clean_df(df_eur)

            df_jh[COL_INTERVENANT]  = df_jh[COL_INTERVENANT].apply(_norm_name)
            df_eur[COL_INTERVENANT] = df_eur[COL_INTERVENANT].apply(_norm_name)
            self._log("  Intervenants normalisés dans JH et EUR.", "info")

            def normalize_mois(val):
                v = str(val).strip().lower()
                try:
                    return MOIS_ORDRE.get(int(float(v)))
                except (ValueError, TypeError):
                    pass
                return MOIS_ORDRE.get(v)

            df_jh["mois_num"]  = df_jh[COL_MOIS].apply(normalize_mois)
            df_eur["mois_num"] = df_eur[COL_MOIS].apply(normalize_mois)

            for label, df in [("JH", df_jh), ("EUR", df_eur)]:
                n = df["mois_num"].isna().sum()
                if n:
                    self._log(f"  ⚠  {n} ligne(s) {label} avec mois "
                              f"non reconnu — ignorées.", "warn")

            df_jh  = df_jh.dropna(subset=["mois_num"])
            df_eur = df_eur.dropna(subset=["mois_num"])
            df_jh["mois_num"]  = df_jh["mois_num"].astype(int)
            df_eur["mois_num"] = df_eur["mois_num"].astype(int)

            key = [COL_INTERVENANT, COL_PROJET, "mois_num"]
            agg_jh  = (df_jh.groupby(key, as_index=False)[COL_NBJOURS]
                       .sum().rename(columns={COL_NBJOURS: "jh"}))
            agg_eur = (df_eur.groupby(key, as_index=False)[COL_NBJOURS]
                       .sum().rename(columns={COL_NBJOURS: "eur"}))

            merged = pd.merge(agg_jh, agg_eur, on=key, how="inner")
            merged = merged[merged["jh"] > 0].copy()
            merged["tjm"] = merged["eur"] / merged["jh"]

            self._log(f"  Couples (intervenant, projet, mois) appariés : "
                      f"{len(merged)}", "info")

            key2      = [COL_INTERVENANT, "mois_num"]
            anomalies = []
            tjm_final = {}

            for (interv, mois), grp in merged.groupby(key2):
                tjms    = grp["tjm"].tolist()
                projets = grp[COL_PROJET].tolist()
                tjm_max = max(tjms)
                tjm_final[(interv, mois)] = tjm_max

                if max(tjms) - min(tjms) > 0.01:
                    detail = ", ".join(
                        f"{p}={t:.2f}€" for p, t in zip(projets, tjms))
                    anomalies.append((interv, mois, tjm_max, detail))
                    self._log(
                        f"  ⚠  CONFLIT TJM | Intervenant : {interv} | "
                        f"Mois : {mois:02d} | "
                        f"TJM/projet : {detail} "
                        f"→ retenu (max) : {tjm_max:.2f} €",
                        "warn")

            self._log(f"  Anomalies TJM : {len(anomalies)}",
                      "warn" if anomalies else "ok")
            self._update_anom_box(anomalies, {})

            intervenants = sorted(merged[COL_INTERVENANT].unique())
            self._log(f"  Intervenants distincts : {len(intervenants)}", "info")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Histo_TJM"

            hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            hdr_fill   = PatternFill("solid", start_color="1F4E79")
            center_aln = Alignment(horizontal="center", vertical="center")
            left_aln   = Alignment(horizontal="left",   vertical="center")
            num_fmt    = '#,##0.00 "€"'
            thin       = Side(style="thin", color="D0D0D0")
            brd        = Border(left=thin, right=thin, top=thin, bottom=thin)
            alt_fill   = PatternFill("solid", start_color="EBF3FB")
            data_font  = Font(name="Arial", size=10)

            c = ws.cell(1, 1, "Intervenant")
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = left_aln; c.border = brd
            ws.column_dimensions["A"].width = 34

            col_letters = list("BCDEFGHIJKLM")
            for idx, (letter, label) in enumerate(
                    zip(col_letters, MOIS_LABELS), start=1):
                c = ws.cell(1, idx + 1, label)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = center_aln; c.border = brd
                ws.column_dimensions[letter].width = 14

            for row_idx, interv in enumerate(intervenants, start=2):
                use_alt  = (row_idx % 2 == 0)
                row_fill = alt_fill if use_alt else None

                c = ws.cell(row_idx, 1, interv)
                c.font = data_font; c.alignment = left_aln; c.border = brd
                if row_fill:
                    c.fill = row_fill

                for m in range(1, 13):
                    tjm_val = tjm_final.get((interv, m))
                    cell    = ws.cell(row_idx, m + 1,
                                      round(tjm_val, 2) if tjm_val is not None else "")
                    cell.font      = data_font
                    cell.alignment = center_aln
                    cell.border    = brd
                    if row_fill:
                        cell.fill = row_fill
                    if tjm_val is not None:
                        cell.number_format = num_fmt

            ws.freeze_panes = "B2"
            wb.save(path_out)

            self._log(f"  ✔  Fichier créé → {path_out}", "ok")
            self._log_duration("Génération Histo_TJM totale", t_global_tjm, tag="section")
            self._log("━" * 54, "section")
            self._log(" HISTO_TJM GÉNÉRÉ AVEC SUCCÈS ✔", "ok")
            self._log("━" * 54, "section")
            self._set_status(
                f"Terminé — {len(intervenants)} intervenants | "
                f"{len(anomalies)} anomalie(s) → Histo_TJM.xlsx")
            self.after(100, lambda: messagebox.showinfo(
                "Histo TJM généré",
                f"Fichier Histo_TJM.xlsx créé avec succès !\n\n"
                f"  • {len(intervenants)} intervenant(s)\n"
                f"  • {len(anomalies)} anomalie(s) TJM détectée(s)\n\n"
                f"Emplacement : {path_out}"))

        except Exception as exc:
            import traceback
            self._log(f"✘ ERREUR : {exc}", "error")
            self._log(traceback.format_exc(), "error")
            self._set_status(f"Erreur : {exc}")
            self.after(100, lambda: messagebox.showerror(
                "Erreur de traitement", str(exc)))
        finally:
            self.after(100, lambda: self.btn_tjm.config(state="normal"))

    # ── Création TCD natifs via win32com ──────────────────────────────────
    def _create_pivot_tables(self, path_xlsx, hdr_projet, hdr_task):
        try:
            import win32com.client as win32
            import pythoncom
        except ImportError:
            return False

        pythoncom.CoInitialize()
        xl = None
        wb = None
        try:
            xl = win32.DispatchEx("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.AutomationSecurity = 1
            try:
                xl.Application.FileDialog(1)
            except Exception:
                pass

            wb = xl.Workbooks.Open(
                path_xlsx,
                UpdateLinks=False,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True,
                Notify=False)

            xlDatabase    = 1
            xlDataField   = 4
            xlRowField    = 1
            xlColumnField = 2
            xlPageField   = 3
            xlSum         = -4157

            def _make_pivot(wb, src_sheet_name, src_headers,
                            dest_sheet_name,
                            row_field, col_field, data_field,
                            filter_field=None, filter_value=None):
                ws_src    = wb.Sheets(src_sheet_name)
                last_row  = ws_src.UsedRange.Rows.Count
                last_col  = ws_src.UsedRange.Columns.Count
                src_range = ws_src.Range(
                    ws_src.Cells(1, 1),
                    ws_src.Cells(last_row, last_col))

                try:
                    ws_dest = wb.Sheets(dest_sheet_name)
                    ws_dest.Cells.Clear()
                except Exception:
                    ws_dest = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
                    ws_dest.Name = dest_sheet_name

                pc = wb.PivotCaches().Create(
                    SourceType=xlDatabase,
                    SourceData=src_range)

                pt = pc.CreatePivotTable(
                    TableDestination=ws_dest.Cells(3, 1),
                    TableName=dest_sheet_name)

                pt.CompactLayoutRowHeader = row_field

                if filter_field:
                    pf = pt.PivotFields(filter_field)
                    pf.Orientation = xlPageField
                    pf.Position = 1
                    pf.CurrentPage = filter_value

                pf_row = pt.PivotFields(row_field)
                pf_row.Orientation = xlRowField
                pf_row.Position = 1

                pf_col = pt.PivotFields(col_field)
                pf_col.Orientation = xlColumnField
                pf_col.Position = 1

                pf_data = pt.PivotFields(data_field)
                pf_data.Orientation = xlDataField
                pf_data.Function = xlSum
                pf_data.NumberFormat = "# ##0"
                pf_data.Name = f"Somme de {data_field}"

                pt.TableStyle2 = "PivotStyleMedium2"
                ws_dest.Columns.AutoFit()
                self._log(f"  ✔  TCD '{dest_sheet_name}' créé.", "ok")

            self._log("  ℹ  Création TCD 1 (TCD_Projets)...", "info")
            _make_pivot(wb,
                        src_sheet_name  = "Diana_Projet",
                        src_headers     = hdr_projet,
                        dest_sheet_name = "TCD_Projets",
                        row_field       = "ProjectName",
                        col_field       = "dianamonth",
                        data_field      = "conso_diana_ttnr")

            self._log("  ℹ  Création TCD 2 (TCD_P02508)...", "info")
            _make_pivot(wb,
                        src_sheet_name  = "Diana_Task",
                        src_headers     = hdr_task,
                        dest_sheet_name = "TCD_P02508",
                        row_field       = "stream",
                        col_field       = "dianamonth",
                        data_field      = "conso_diana_ttnr",
                        filter_field    = "Project Code",
                        filter_value    = "P02508")

            self._log("  ℹ  Sauvegarde du fichier avec TCD...", "info")
            xl.DisplayAlerts = False
            xl.EnableEvents  = False
            for addin in xl.COMAddIns:
                try:
                    if "label" in addin.Description.lower() or \
                       "confidential" in addin.Description.lower() or \
                       "sensitivity" in addin.ProgId.lower() or \
                       "aip" in addin.ProgId.lower():
                        addin.Connect = False
                except Exception:
                    pass
            wb.Save()
            self._log("  ℹ  wb.Save() exécuté.", "info")
            xl.EnableEvents  = True
            xl.DisplayAlerts = False
            wb.Close(SaveChanges=True)
            self._log("  ℹ  wb.Close(SaveChanges=True) exécuté.", "info")
            return True

        except Exception as exc:
            import traceback
            self._log(f"  ✘  Erreur création TCD : {exc}", "error")
            self._log(traceback.format_exc(), "error")
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            return False
        finally:
            try:
                if xl is not None:
                    xl.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()

    # ── Bouton MAJ PDC ────────────────────────────────────────────────────
    def _open_pdc_maj_window(self):
        """Ouvre la fenêtre de mise à jour du plan de charges.
        Déclenche le chargement du cache PDC si pas encore prêt (v4.63 : au clic, plus au démarrage).
        """
        if not HAS_PANDAS:
            messagebox.showerror("Dépendance manquante",
                                  "pandas est requis : pip install pandas openpyxl")
            return
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()
        if not path_jh or not path_eur:
            messagebox.showwarning("Fichiers manquants",
                                    "Sélectionnez d'abord les fichiers PDC (2 et 3).")
            return

        if not self._pdc_cache_ready:
            # Lancer le chargement du cache en arrière-plan,
            # puis ouvrir la fenêtre une fois le cache prêt
            self._set_status("Chargement du cache PDC en cours…")
            self._preload_pdc_cache(on_done=self._open_pdc_maj_window_now)
        else:
            self._open_pdc_maj_window_now()

    def _open_pdc_maj_window_now(self):
        """Ouvre effectivement la fenêtre PdcMajWindow (cache disponible)."""
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()
        win = PdcMajWindow(self, path_jh, path_eur,
                           cache_usernames=self.pdc_usernames,
                           cache_jh=self.pdc_jh_data,
                           cache_eur=self.pdc_eur_data,
                           cache_eur_detail=self.pdc_eur_detail,
                           diana_jh=self.diana_jh_data,
                           diana_eur=self.diana_eur_data,
                           histo_tjm=self.histo_tjm_data,
                           cache_ready=self._pdc_cache_ready,
                           nucleus_exclusions=self._nucleus_exclusions)
        win.focus_set()

    def _update_diana_btn(self):
        has_diana = bool(self.file1_path.get())
        self.btn_diana.config(state="normal" if has_diana else "disabled")

    def _open_diana_window(self):
        path = self.file1_path.get()
        if not path:
            messagebox.showwarning("Fichier manquant",
                                    "Sélectionnez d'abord le fichier Diana (fichier 1).")
            return
        flt = os.path.splitext(path)[0] + "_avec_filtre.csv"
        win = DianaWindow(self, flt if os.path.isfile(flt) else path)
        win.focus_set()

    # ── Pré-chargement cache PDC ──────────────────────────────────────────
    def _preload_pdc_cache(self, on_done=None):
        """Pré-charge les structures PDC.
        v4.63 : déclenché au clic sur MAJ PDC (plus au démarrage).
        Utilise les CSV filtrés si disponibles via _preprocess_sources().
        on_done : callback appelé dans le thread principal une fois le cache prêt.
        """
        path1 = self.file1_path.get()
        path2 = self.file2_path.get()
        path3 = self.file3_path.get()
        if not (path1 and path2 and path3):
            self._log("  ℹ  Pré-chargement PDC ignoré (fichiers incomplets).", "info")
            return

        self._log("  ⏳  Pré-chargement cache PDC (MAJ Plan de charges)…", "info")
        self._set_status("Pré-chargement cache PDC…")

        def _run():
            try:
                annee = datetime.now().year

                # ── Résolution des chemins CSV filtrés ────────────────────────
                # On réutilise _preprocess_sources pour obtenir les bons chemins
                # (génère les CSV s'ils n'existent pas, ou réutilise s'ils existent)
                try:
                    path_diana_eff, path_jh_eff, path_eur_eff = self._preprocess_sources()
                except Exception as e_pre:
                    self.after(0, lambda: self._log(
                        f"  ⚠  Prétraitement PDC cache : {e_pre} — utilisation des chemins bruts.", "warn"))
                    path_diana_eff = path1
                    path_jh_eff    = path2
                    path_eur_eff   = path3

                # ── Intervenants périmètre ─────────────────────────────────────
                rgu_users      = {_norm_name(n): n for n in self._allowed_list}
                usernames_list = sorted(rgu_users.values())

                # ── Lecture PDC CSV filtré ────────────────────────────────────
                def _read_pdc_flt(path):
                    """Lit un CSV filtré (sep=;, utf-8-sig, header ligne 1)."""
                    df = pd.read_csv(path, dtype=str, sep=SEP_OUTPUT,
                                     encoding="utf-8-sig", encoding_errors="replace",
                                     on_bad_lines="skip")
                    df.columns = [c.strip() for c in df.columns]
                    return df.fillna("")

                def _norm_col(s):
                    return s.lower().replace(" ", "").replace("_", "")

                df_jh  = _read_pdc_flt(path_jh_eff)
                df_eur = _read_pdc_flt(path_eur_eff)

                def _find(df, name):
                    if name in df.columns: return name
                    k = _norm_col(name)
                    return next((c for c in df.columns if _norm_col(c) == k), None)

                ic_jh  = _find(df_jh,  COL_INTERVENANT)
                mc_jh  = _find(df_jh,  COL_MOIS)
                pc_jh  = _find(df_jh,  COL_PROJET)
                nc_jh  = _find(df_jh,  COL_NBJOURS)
                yc_jh  = next((c for c in df_jh.columns  if "year" in c.lower()), None)
                ic_eur = _find(df_eur, COL_INTERVENANT)
                mc_eur = _find(df_eur, COL_MOIS)
                pc_eur = _find(df_eur, COL_PROJET)
                nc_eur = _find(df_eur, COL_NBJOURS)
                yc_eur = next((c for c in df_eur.columns if "year" in c.lower()), None)

                def _norm_c(s): return s.lower().replace(" ","").replace("_","")
                lc_jh = next(
                    (c for c in df_jh.columns
                     if _norm_c(c) in ("prjlibelleprojet","libelleprojet","libprojet")
                     or ("libelle" in c.lower() and "projet" in c.lower())
                     or ("libelle" in c.lower() and "prj" in c.lower())), None)
                if not lc_jh:
                    lc_jh = next(
                        (c for c in df_jh.columns if "libelle" in c.lower()), None)

                def _to_mois(v):
                    v2 = str(v).strip().lower()
                    r = MOIS_ORDRE.get(v2)
                    if r: return r
                    try: return int(float(v2))
                    except: return None

                if yc_jh:
                    df_jh = df_jh[df_jh[yc_jh].astype(str).str.strip() == str(annee)]
                if yc_eur:
                    df_eur = df_eur[df_eur[yc_eur].astype(str).str.strip() == str(annee)]

                # ── Dict JH ───────────────────────────────────────────────────
                jh_data  = {}
                pdc_only = {}
                for _, row in df_jh.iterrows():
                    interv_raw  = str(row.get(ic_jh, "")).strip()
                    interv_norm = _norm_name(interv_raw)
                    prj         = str(row.get(pc_jh, "")).strip()
                    in_csv      = interv_norm in rgu_users
                    target_proj = prj in TARGET_PROJECTS
                    if not in_csv:
                        if interv_raw and interv_norm and target_proj:
                            pdc_only[interv_norm] = interv_raw
                        if not target_proj:
                            continue
                    mois = _to_mois(row.get(mc_jh, ""))
                    jh   = 0.0
                    try: jh = float(str(row.get(nc_jh, "0")).replace(",", "."))
                    except: pass
                    if not prj or mois is None: continue
                    if interv_norm not in jh_data:
                        jh_data[interv_norm] = {}
                    if prj not in jh_data[interv_norm]:
                        lib = str(row.get(lc_jh, "")) if lc_jh else ""
                        jh_data[interv_norm][prj] = {"_lib": lib}
                    d = jh_data[interv_norm][prj]
                    d[int(mois)] = d.get(int(mois), 0.0) + jh

                # Compléter libellés depuis EUR
                lc_eur = next(
                    (c for c in df_eur.columns
                     if _norm_c(c) in ("prjlibelleprojet","libelleprojet","libprojet")
                     or ("libelle" in c.lower() and "projet" in c.lower())
                     or ("libelle" in c.lower() and "prj" in c.lower())), None)
                if not lc_eur:
                    lc_eur = next(
                        (c for c in df_eur.columns if "libelle" in c.lower()), None)

                if lc_eur:
                    lib_from_eur = {}
                    for _, row in df_eur.iterrows():
                        prj2 = str(row.get(pc_eur, "")).strip()
                        lib2 = str(row.get(lc_eur, "")).strip()
                        if prj2 and lib2 and prj2 not in lib_from_eur:
                            lib_from_eur[prj2] = lib2
                    for u_norm, prj_dict in jh_data.items():
                        for prj2, prj_info in prj_dict.items():
                            if not prj_info.get("_lib") and prj2 in lib_from_eur:
                                prj_info["_lib"] = lib_from_eur[prj2]

                projets_rgu = set()
                for prj_dict in jh_data.values():
                    projets_rgu.update(prj_dict.keys())

                csv_only = sorted(
                    rgu_users[n] for n in rgu_users if n not in jh_data)
                pdc_only_list = sorted(pdc_only.values())
                if csv_only:
                    self._log(f"  ⚠  Dans CSV mais absents du PDC ({len(csv_only)}) :", "warn")
                    for name in csv_only:
                        self._log(f"    • {name}", "warn")
                else:
                    self._log("  ✔  Tous les intervenants CSV ont des données PDC.", "ok")
                if pdc_only_list:
                    self._log(f"  ⚠  Dans PDC mais absents du CSV ({len(pdc_only_list)}) :", "warn")
                    for name in pdc_only_list:
                        self._log(f"    • {name}", "warn")
                else:
                    self._log("  ✔  Tous les intervenants PDC sont dans le CSV.", "ok")

                # ── Dict EUR ──────────────────────────────────────────────────
                eur_data   = {}
                eur_detail = {}
                for _, row in df_eur.iterrows():
                    interv_norm = _norm_name(str(row.get(ic_eur, "")).strip())
                    prj  = str(row.get(pc_eur, "")).strip()
                    if prj not in TARGET_PROJECTS and interv_norm not in rgu_users: continue
                    mois_e = _to_mois(row.get(mc_eur, ""))
                    eur = 0.0
                    try: eur = float(str(row.get(nc_eur, "0")).replace(",", "."))
                    except: pass
                    if not prj: continue
                    eur_data[prj] = eur_data.get(prj, 0.0) + eur
                    if mois_e:
                        key_d = (interv_norm, prj, int(mois_e))
                        eur_detail[key_d] = eur_detail.get(key_d, 0.0) + eur

                # ── Diana JH réels (mois passés) depuis CSV filtré ────────────
                annee_str  = str(annee)
                mois_cur   = datetime.now().month
                HDR_PC     = "Project Code"
                HDR_UFIRST = "User First Name"
                HDR_ULAST  = "User Last Name"
                HDR_DATE   = "Timesheet Entry Details Calendar Date"
                HDR_VAL    = "Timesheet Entry Details Value"

                diana_jh = {}

                with open(path_diana_eff, "r", encoding="utf-8-sig", errors="replace") as f:
                    d_lines = f.readlines()

                d_hdrs = [h.strip().strip('"') for h in d_lines[0].split(SEP_OUTPUT)]
                def _dci(n):
                    try: return d_hdrs.index(n)
                    except ValueError: return -1

                di_pc  = _dci(HDR_PC)
                di_uf  = _dci(HDR_UFIRST); di_ul = _dci(HDR_ULAST)
                di_dt  = _dci(HDR_DATE);   di_vl = _dci(HDR_VAL)
                di_org = _dci("Organization Unit Name")

                for line in d_lines[1:]:
                    line = line.rstrip("\n\r")
                    if not line.strip(): continue
                    flds = [f.strip().strip('"') for f in line.split(SEP_OUTPUT)]
                    while len(flds) < len(d_hdrs): flds.append("")
                    def _fv(i): return flds[i] if 0 <= i < len(flds) else ""

                    pc_v = _fv(di_pc)

                    # Retenir si Organization Unit Name contient "RGU" OU projet piloté
                    if "RGU" not in _fv(di_org).upper() and pc_v not in TARGET_PROJECTS:
                        continue

                    date_v = _fv(di_dt)
                    if not date_v or len(date_v) < 7: continue
                    try:
                        y_v = int(date_v[:4]); m_v = int(date_v[5:7])
                    except ValueError: continue
                    if y_v != annee or m_v >= mois_cur: continue

                    u_orig = f"{_fv(di_ul)}, {_fv(di_uf)}"
                    u_norm = _norm_name(u_orig)
                    if not u_norm: continue

                    if pc_v not in TARGET_PROJECTS and u_norm not in rgu_users: continue
                    jh_v = 0.0
                    try: jh_v = float(_fv(di_vl).replace(",","."))
                    except: pass

                    if u_norm not in diana_jh: diana_jh[u_norm] = {}
                    if pc_v not in diana_jh[u_norm]:
                        diana_jh[u_norm][pc_v] = {}
                    diana_jh[u_norm][pc_v][m_v] = diana_jh[u_norm][pc_v].get(m_v, 0.0) + jh_v

                # ── Histo_TJM ─────────────────────────────────────────────────
                import os as _os2
                path_histo = _os2.path.join(
                    _os2.path.dirname(_os2.path.abspath(__file__)),
                    "Histo_TJM.xlsx")
                histo_tjm = {}
                if HAS_OPENPYXL and _os2.path.isfile(path_histo):
                    wb_h = openpyxl.load_workbook(path_histo, read_only=True, data_only=True)
                    ws_h = wb_h.active
                    for ri, row in enumerate(ws_h.iter_rows(values_only=True)):
                        if ri == 0 or row[0] is None: continue
                        u_norm_h = _norm_name(str(row[0]).strip())
                        for mois_i in range(1, 13):
                            val = row[mois_i] if len(row) > mois_i else None
                            if val is not None:
                                try: histo_tjm[(u_norm_h, mois_i)] = float(val)
                                except (ValueError, TypeError): pass
                    wb_h.close()

                # ── EUR Diana (JH × TJM) ──────────────────────────────────────
                diana_eur = {}
                for u_norm, prj_dict in diana_jh.items():
                    for prj, mois_dict in prj_dict.items():
                        for mois, jh_v in mois_dict.items():
                            tjm = histo_tjm.get((u_norm, mois))
                            if tjm and jh_v:
                                eur_v = round(jh_v * tjm, 2)
                                if u_norm not in diana_eur: diana_eur[u_norm] = {}
                                if prj not in diana_eur[u_norm]: diana_eur[u_norm][prj] = {}
                                diana_eur[u_norm][prj][mois] = eur_v

                # ── Mise à jour thread principal ──────────────────────────────
                def _done():
                    self.pdc_usernames   = usernames_list
                    self.pdc_jh_data     = jh_data
                    self.pdc_eur_data    = eur_data
                    self.pdc_eur_detail  = eur_detail
                    self.pdc_projets_rgu = projets_rgu
                    self.diana_jh_data   = diana_jh
                    self.diana_eur_data  = diana_eur
                    self.histo_tjm_data  = histo_tjm
                    self._pdc_cache_ready = True
                    self._log(
                        f"  ✔  Cache PDC prêt : {len(usernames_list)} intervenant(s) périmètre, "
                        f"{len(projets_rgu)} projet(s), "
                        f"{len(diana_jh)} username(s) Diana chargés, "
                        f"Histo_TJM : {len(histo_tjm)} entrée(s).", "ok")
                    self._set_status("Cache PDC + Diana prêt.")
                    if on_done:
                        self.after(0, on_done)

                self.after(0, _done)

            except Exception as exc:
                import traceback
                self.after(0, lambda: self._log(
                    f"  ✘  Erreur pré-chargement PDC : {exc}\n{traceback.format_exc()}", "error"))

        import threading
        threading.Thread(target=_run, daemon=True).start()

    # ── Bouton Crapull ────────────────────────────────────────────────────
    def _update_crapull_btn(self):
        has_pdc = bool(self.file2_path.get() and self.file3_path.get())
        self.btn_crapull.config(state="normal" if has_pdc else "disabled")
        self.btn_pdc_maj.config(state="normal" if has_pdc else "disabled")

    def _open_crapull_window(self):
        if not HAS_PANDAS:
            messagebox.showerror("Dépendance manquante",
                                  "pandas est requis : pip install pandas openpyxl")
            return
        path_jh  = self.file2_path.get()
        path_eur = self.file3_path.get()
        if not path_jh or not path_eur:
            messagebox.showwarning("Fichiers manquants",
                                    "Sélectionnez d'abord les fichiers PDC (2 et 3).")
            return

        def _filtered(path):
            """Retourne le chemin du CSV filtré s'il existe, sinon le fichier source."""
            flt = os.path.splitext(path)[0] + "_avec_filtre.csv"
            return flt if os.path.isfile(flt) else path

        win = CrapullWindow(self, _filtered(path_jh), _filtered(path_eur))
        win.focus_set()

    # ── Helpers ───────────────────────────────────────────────────────────
    def _log(self, text, tag="info"):
        ts = datetime.now().strftime("%H:%M:%S")
        prefixed = text if text.strip().startswith("━") else f"[{ts}] {text}"
        def _do():
            self.log_box.config(state="normal")
            self.log_box.insert("end", prefixed + "\n", tag)
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _do)

    def _log_duration(self, label, t0, tag="ok"):
        elapsed = time.perf_counter() - t0
        dur_str = (f"{elapsed:.1f}s" if elapsed < 60
                   else f"{int(elapsed)//60}m {int(elapsed)%60:02d}s")
        self._log(f"  ⏱  {label} terminé en {dur_str}", tag)

    def _clear_log(self):
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")
        self._set_status("Journal effacé.")

    def _set_status(self, msg):
        self.after(0, lambda: self.status_var.set(msg))

    def _update_rgu_box(self, names):
        def _do():
            self.rgu_box.config(state="normal")
            self.rgu_box.delete("1.0", "end")
            self.rgu_box.insert(
                "1.0",
                "\n".join(names) if names
                else "(aucun intervenant périmètre chargé)")
            self.rgu_box.config(state="disabled")
        self.after(0, _do)

    def _update_inco_box(self, inco_names):
        def _do():
            self.inco_tree.delete(*self.inco_tree.get_children())
            rows = sorted(inco_names.items(), key=lambda x: x[1][0])
            for i, (norm, (diana_orig, tjm_orig)) in enumerate(rows):
                tag = "even" if i % 2 == 0 else "odd"
                self.inco_tree.insert("", "end",
                    values=(diana_orig, tjm_orig, norm, norm),
                    tags=(tag,))
        self.after(0, _do)

    def _update_anom_box(self, anomalies, tjm_not_found=None):
        def _do():
            self.anom_box.config(state="normal")
            self.anom_box.delete("1.0", "end")
            content_lines = []

            if anomalies:
                content_lines.append("═" * 60)
                content_lines.append(f"  CONFLITS TJM ({len(anomalies)} cas)")
                content_lines.append("  Plusieurs TJM différents pour un même intervenant/mois")
                content_lines.append("═" * 60)
                for interv, mois, tjm_max, detail in anomalies:
                    content_lines.append(
                        f"\n[Mois {mois:02d}] {interv}\n"
                        f"  TJM par projet   : {detail}\n"
                        f"  TJM retenu (max) : {tjm_max:.2f} €")
            else:
                content_lines.append("✔  Aucun conflit TJM détecté.")

            content_lines.append("")
            if tjm_not_found:
                content_lines.append("═" * 60)
                content_lines.append(f"  TJM MANQUANTS ({len(tjm_not_found)} combinaison(s))")
                content_lines.append("  conso_diana_ttnr non calculable (TJM absent dans Histo_TJM)")
                content_lines.append("═" * 60)
                for (username, mois_num), details in sorted(tjm_not_found.items()):
                    content_lines.append(
                        f"\n[Mois {mois_num:02d}] {username}")
                    for proj_name, wpn, cjh in details:
                        content_lines.append(
                            f"  ↳ Projet  : {proj_name}\n"
                            f"     WPN    : {wpn}\n"
                            f"     Conso  : {cjh:.4f} jh")
            else:
                content_lines.append("✔  Tous les TJM ont été trouvés dans Histo_TJM.")

            self.anom_box.insert("1.0", "\n".join(content_lines))
            self.anom_box.config(state="disabled")
        self.after(0, _do)


# ─────────────────────────────────────────────
#  FENÊTRE VÉRIFICATION CRAPULL
# ─────────────────────────────────────────────
CRAPULL_COLS = ["Intervenant", "Type Intervenant", "Code UT", "Year of charge",
                "Mois", "NbJours", "PRJ_CodeProjet", "PRJ_LibelleProjet"]
FILTER_COLS  = ["Intervenant", "Mois", "PRJ_CodeProjet", "PRJ_LibelleProjet"]


class CrapullWindow(tk.Toplevel):
    """Fenêtre de visualisation et filtrage des données Crapull (PDC JH + EUR)."""

    def __init__(self, parent, path_jh, path_eur):
        super().__init__(parent)
        self.title("Vérification données Crapull")
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)
        self.minsize(1000, 700)

        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w, h = min(1300, sw - 80), min(820, sh - 80)
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        self.path_jh  = path_jh
        self.path_eur = path_eur
        self._df_jh   = None
        self._df_eur  = None
        self._filter_vars  = {}
        self._real_col_names = {}

        self._build_ui()
        self.after(100, self._load_data)

    def _load_data(self):
        self._status.set("Chargement des fichiers PDC…")
        self.update_idletasks()
        try:
            self._df_jh  = self._read_pdc(self.path_jh)
            self._df_eur = self._read_pdc(self.path_eur)
            self._status.set(
                f"JH : {len(self._df_jh)} lignes  |  EUR : {len(self._df_eur)} lignes")
            self._populate_filters()
            self._refresh_headings()
            self._apply_filters()
        except Exception as e:
            self._status.set(f"Erreur : {e}")
            messagebox.showerror("Erreur chargement", str(e), parent=self)

    def _read_pdc(self, path):
        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            enc = "utf-8-sig"
            for candidate in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
                try:
                    with open(path, "r", encoding=candidate) as f:
                        sample = f.read(8192)
                    enc = candidate
                    break
                except UnicodeDecodeError:
                    continue
            sep = ";" if sample.count(";") >= sample.count(",") else ","
            df = pd.read_csv(path, dtype=str, sep=sep, encoding=enc,
                              encoding_errors="replace", on_bad_lines="skip")
        else:
            df = pd.read_excel(path, dtype=str, header=3)

        df.columns = [c.strip() for c in df.columns]

        def _norm(s):
            return s.lower().replace(" ", "").replace("_", "")

        col_map = {_norm(c): c for c in df.columns}

        rename   = {}
        real_map = {}

        for crapull_col in CRAPULL_COLS:
            key = _norm(crapull_col)
            if crapull_col in df.columns:
                real_map[crapull_col] = crapull_col
            elif key in col_map:
                real_name = col_map[key]
                rename[real_name] = crapull_col
                real_map[crapull_col] = real_name
            else:
                df[crapull_col] = ""
                real_map[crapull_col] = crapull_col

        if rename:
            df = df.rename(columns=rename)

        if not hasattr(self, "_real_col_names"):
            self._real_col_names = real_map
        else:
            for k, v in real_map.items():
                self._real_col_names.setdefault(k, v)

        return df[CRAPULL_COLS].fillna("").astype(str)

    def _build_ui(self):
        hdr = tk.Frame(self, bg=BG_MAIN, pady=8)
        hdr.pack(fill="x", padx=16)
        tk.Label(hdr, text="🔍  VÉRIFICATION DONNÉES CRAPULL",
                 bg=BG_MAIN, fg=ACCENT2, font=FONT_HEAD).pack(side="left")
        tk.Label(hdr, text=os.path.basename(self.path_jh),
                 bg=BG_MAIN, fg=TEXT_SEC, font=FONT_SMALL).pack(side="right")

        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        filter_frame = tk.LabelFrame(
            self, text="  Filtres  ", bg=BG_PANEL, fg=ACCENT2,
            font=FONT_SMALL, bd=1, relief="flat",
            highlightbackground=BORDER, highlightthickness=1)
        filter_frame.pack(fill="x", padx=10, pady=(8, 4), ipady=6)

        self._combo_widgets = {}
        for i, col in enumerate(FILTER_COLS):
            col_frame = tk.Frame(filter_frame, bg=BG_PANEL)
            col_frame.grid(row=0, column=i*2, padx=(10, 2), pady=4, sticky="w")
            tk.Label(col_frame, text=col, bg=BG_PANEL,
                     fg=TEXT_SEC, font=FONT_SMALL).pack(anchor="w")
            var = tk.StringVar(value="(Tous)")
            self._filter_vars[col] = var

            style = ttk.Style()
            style.theme_use("default")
            style.configure("Dark.TCombobox",
                            fieldbackground=BG_ENTRY, background=BG_CARD,
                            foreground=TEXT_PRI, selectbackground=ACCENT2,
                            font=FONT_SMALL)

            cb = ttk.Combobox(col_frame, textvariable=var,
                               state="readonly", width=22,
                               style="Dark.TCombobox", font=FONT_SMALL)
            cb.pack()
            cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())
            self._combo_widgets[col] = cb

        reset_frame = tk.Frame(filter_frame, bg=BG_PANEL)
        reset_frame.grid(row=0, column=len(FILTER_COLS)*2, padx=10)
        make_button(reset_frame, "↺  Réinitialiser",
                    self._reset_filters, width=16).pack(pady=8)

        self._build_table_section("📋  Consommés JH", "jh")
        self._build_table_section("💶  Consommés EUR (TTNR)", "eur")

        self._status = tk.StringVar(value="Chargement…")
        tk.Label(self, textvariable=self._status, bg=BG_CARD,
                 fg=TEXT_SEC, font=FONT_SMALL, anchor="w",
                 padx=12).pack(fill="x", side="bottom")

    def _build_table_section(self, title, tag):
        section = tk.Frame(self, bg=BG_MAIN)
        section.pack(fill="both", expand=True, padx=10, pady=(4, 0))

        tk.Label(section, text=title, bg=BG_MAIN,
                 fg=ACCENT2, font=FONT_BODY).pack(anchor="w", padx=4)

        count_var = tk.StringVar(value="")
        tk.Label(section, textvariable=count_var, bg=BG_MAIN,
                 fg=TEXT_SEC, font=FONT_SMALL).pack(anchor="w", padx=4)

        tbl_frame = tk.Frame(section, bg=BG_CARD,
                              highlightbackground=BORDER, highlightthickness=1)
        tbl_frame.pack(fill="both", expand=True, pady=(2, 6))

        vsb = tk.Scrollbar(tbl_frame, orient="vertical")
        hsb = tk.Scrollbar(tbl_frame, orient="horizontal")
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")

        tv = ttk.Treeview(tbl_frame,
                           columns=CRAPULL_COLS,
                           show="headings",
                           yscrollcommand=vsb.set,
                           xscrollcommand=hsb.set)
        tv.pack(fill="both", expand=True)
        vsb.config(command=tv.yview)
        hsb.config(command=tv.xview)

        style = ttk.Style()
        style.configure("Dark.Treeview",
                         background=BG_CARD, foreground=TEXT_PRI,
                         fieldbackground=BG_CARD, rowheight=22,
                         font=FONT_SMALL)
        style.configure("Dark.Treeview.Heading",
                         background=BG_ENTRY, foreground=ACCENT2,
                         font=("Consolas", 9, "bold"))
        style.map("Dark.Treeview",
                  background=[("selected", ACCENT2)],
                  foreground=[("selected", "#000000")])
        tv.config(style="Dark.Treeview")

        col_widths = {
            "Intervenant": 180, "Type Intervenant": 210, "Code UT": 80,
            "Year of charge": 100, "Mois": 55, "NbJours": 90,
            "PRJ_CodeProjet": 110, "PRJ_LibelleProjet": 220,
        }
        center_cols = {"Mois", "NbJours", "Code UT", "Year of charge"}
        for col in CRAPULL_COLS:
            real_name = getattr(self, "_real_col_names", {}).get(col, col)
            tv.heading(col, text=real_name,
                        command=lambda c=col, t=tv: self._sort_col(t, c))
            anchor = "center" if col in center_cols else "w"
            tv.column(col, width=col_widths.get(col, 100),
                       anchor=anchor, stretch=True)

        tv.tag_configure("odd",  background=BG_CARD)
        tv.tag_configure("even", background=BG_ENTRY)

        setattr(self, f"_tv_{tag}", tv)
        setattr(self, f"_count_{tag}", count_var)

    def _refresh_headings(self):
        for tag in ("jh", "eur"):
            tv = getattr(self, f"_tv_{tag}", None)
            if tv is None:
                continue
            for col in CRAPULL_COLS:
                real_name = self._real_col_names.get(col, col)
                tv.heading(col, text=real_name)

    def _populate_filters(self):
        if self._df_jh is None:
            return
        combined = pd.concat([self._df_jh, self._df_eur], ignore_index=True)
        for col, cb in self._combo_widgets.items():
            vals = sorted(combined[col].unique().tolist())
            cb["values"] = ["(Tous)"] + vals

    def _reset_filters(self):
        for var in self._filter_vars.values():
            var.set("(Tous)")
        self._apply_filters()

    def _apply_filters(self):
        if self._df_jh is None or self._df_eur is None:
            return

        def _filter(df):
            mask = pd.Series([True] * len(df), index=df.index)
            for col, var in self._filter_vars.items():
                val = var.get()
                if val != "(Tous)" and col in df.columns:
                    mask &= (df[col] == val)
            return df[mask]

        df_jh_f  = _filter(self._df_jh)
        df_eur_f = _filter(self._df_eur)

        self._fill_table(self._tv_jh,  df_jh_f,  is_eur=False)
        self._fill_table(self._tv_eur, df_eur_f, is_eur=True)
        self._count_jh.set(f"{len(df_jh_f)} ligne(s)")
        self._count_eur.set(f"{len(df_eur_f)} ligne(s)")
        self._status.set(
            f"Filtre appliqué — JH : {len(df_jh_f)} lignes  |  EUR : {len(df_eur_f)} lignes")

    def _fill_table(self, tv, df, is_eur=False):
        tv.delete(*tv.get_children())
        for i, (_, row) in enumerate(df.iterrows()):
            tag = "even" if i % 2 == 0 else "odd"
            values = []
            for c in CRAPULL_COLS:
                val = row[c]
                if is_eur and c == "NbJours":
                    try:
                        num = float(str(val).replace(",", ".").strip())
                        formatted = f"{num:,.2f}".replace(",", " ")
                        val = formatted
                    except (ValueError, TypeError):
                        pass
                values.append(val)
            tv.insert("", "end", values=values, tags=(tag,))

    def _sort_col(self, tv, col):
        data = [(tv.set(k, col), k) for k in tv.get_children("")]
        data.sort(reverse=getattr(tv, f"_sort_rev_{col}", False))
        for i, (_, k) in enumerate(data):
            tv.move(k, "", i)
            tv.item(k, tags=("even" if i % 2 == 0 else "odd",))
        setattr(tv, f"_sort_rev_{col}",
                not getattr(tv, f"_sort_rev_{col}", False))


# ─────────────────────────────────────────────
#  FENÊTRE VÉRIFICATION DIANA
# ─────────────────────────────────────────────
DIANA_COLS = [
    "Project Code", "Project Is Alive", "WorkPackage Name", "Task Name",
    "Organization Unit Name", "User Country Name", "User Contract Type",
    "TimeSheet Mode Name", "Status", "Is Published", "Application Cluster Name",
    "User First Name", "User Last Name",
    "Project Name",
    "Timesheet Entry Details Calendar Date",
    "Timesheet Entry Details Value",
]
DIANA_FILTER_COLS = [
    "User First Name", "User Last Name",
    "Project Code", "Project Name",
    "WorkPackage Name", "Task Name",
]
DIANA_COL_WIDTHS = {
    "Project Code": 90, "Project Is Alive": 80, "WorkPackage Name": 160,
    "Task Name": 160, "Organization Unit Name": 180, "User Country Name": 100,
    "User Contract Type": 120, "TimeSheet Mode Name": 130, "Status": 80,
    "Is Published": 80, "Application Cluster Name": 160,
    "User First Name": 100, "User Last Name": 130, "Project Name": 200,
    "Timesheet Entry Details Calendar Date": 130,
    "Timesheet Entry Details Value": 100,
}


class DianaWindow(tk.Toplevel):

    def __init__(self, parent, path):
        super().__init__(parent)
        self.title("Vérification données Diana")
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)
        self.minsize(1000, 680)
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w, h = min(1400, sw - 80), min(830, sh - 80)
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        self.path         = path
        self._df          = None
        self._real_cols   = {}
        self._filter_vars = {}

        self._build_ui()
        self.after(100, self._load_data)

    def _load_data(self):
        self._status.set("Chargement du fichier Diana…")
        self.update_idletasks()
        try:
            enc = "utf-8-sig"
            for candidate in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
                try:
                    with open(self.path, "r", encoding=candidate) as f:
                        f.read(8192)
                    enc = candidate
                    break
                except UnicodeDecodeError:
                    continue

            # Détecter le séparateur (~ pour .txt, ; pour CSV filtré)
            ext = os.path.splitext(self.path)[1].lower()
            sep = SEP_INPUT if ext == ".txt" else SEP_OUTPUT

            with open(self.path, "r", encoding=enc, errors="replace") as f:
                lines = f.readlines()

            if not lines:
                raise ValueError("Fichier vide.")

            raw_headers = [h.strip().strip('"') for h in lines[0].split(sep)]

            def _norm(s):
                return s.lower().replace(" ", "").replace("_", "")

            col_map = {_norm(h): h for h in raw_headers}

            col_indices = {}
            self._real_cols = {}
            for dc in DIANA_COLS:
                key = _norm(dc)
                if dc in raw_headers:
                    col_indices[dc] = raw_headers.index(dc)
                    self._real_cols[dc] = dc
                elif key in col_map:
                    real = col_map[key]
                    col_indices[dc] = raw_headers.index(real)
                    self._real_cols[dc] = real
                else:
                    col_indices[dc] = -1
                    self._real_cols[dc] = dc

            records = []
            for line in lines[1:]:
                line = line.rstrip("\n\r")
                if not line.strip():
                    continue
                fields = [f.strip().strip('"') for f in line.split(sep)]
                while len(fields) < len(raw_headers):
                    fields.append("")
                row = {}
                for dc in DIANA_COLS:
                    idx = col_indices[dc]
                    row[dc] = fields[idx] if idx >= 0 and idx < len(fields) else ""
                records.append(row)

            self._df = pd.DataFrame(records, columns=DIANA_COLS)
            self._status.set(f"{len(self._df)} lignes chargées.")
            self._populate_filters()
            self._refresh_headings()
            self._apply_filters()

        except Exception as e:
            self._status.set(f"Erreur : {e}")
            messagebox.showerror("Erreur chargement", str(e), parent=self)

    def _build_ui(self):
        hdr = tk.Frame(self, bg=BG_MAIN, pady=8)
        hdr.pack(fill="x", padx=16)
        tk.Label(hdr, text="📋  VÉRIFICATION DONNÉES DIANA",
                 bg=BG_MAIN, fg=ACCENT2, font=FONT_HEAD).pack(side="left")
        tk.Label(hdr, text=os.path.basename(self.path),
                 bg=BG_MAIN, fg=TEXT_SEC, font=FONT_SMALL).pack(side="right")
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        filter_frame = tk.LabelFrame(
            self, text="  Filtres  ", bg=BG_PANEL, fg=ACCENT2,
            font=FONT_SMALL, bd=1, relief="flat",
            highlightbackground=BORDER, highlightthickness=1)
        filter_frame.pack(fill="x", padx=10, pady=(8, 4), ipady=6)

        self._combo_widgets = {}
        for i, col in enumerate(DIANA_FILTER_COLS):
            col_frame = tk.Frame(filter_frame, bg=BG_PANEL)
            col_frame.grid(row=0, column=i, padx=(10, 2), pady=4, sticky="w")
            tk.Label(col_frame, text=col, bg=BG_PANEL,
                     fg=TEXT_SEC, font=FONT_SMALL).pack(anchor="w")
            var = tk.StringVar(value="(Tous)")
            self._filter_vars[col] = var
            cb = ttk.Combobox(col_frame, textvariable=var,
                               state="readonly", width=20,
                               font=FONT_SMALL)
            cb.pack()
            cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())
            self._combo_widgets[col] = cb

        reset_frame = tk.Frame(filter_frame, bg=BG_PANEL)
        reset_frame.grid(row=0, column=len(DIANA_FILTER_COLS), padx=10)
        make_button(reset_frame, "↺  Réinitialiser",
                    self._reset_filters, width=16).pack(pady=8)

        section = tk.Frame(self, bg=BG_MAIN)
        section.pack(fill="both", expand=True, padx=10, pady=(4, 0))

        self._count_var = tk.StringVar(value="")
        tk.Label(section, textvariable=self._count_var,
                 bg=BG_MAIN, fg=TEXT_SEC, font=FONT_SMALL).pack(anchor="w", padx=4)

        tbl_frame = tk.Frame(section, bg=BG_CARD,
                              highlightbackground=BORDER, highlightthickness=1)
        tbl_frame.pack(fill="both", expand=True, pady=(2, 6))
        vsb = tk.Scrollbar(tbl_frame, orient="vertical")
        hsb = tk.Scrollbar(tbl_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self._tv = ttk.Treeview(tbl_frame, columns=DIANA_COLS,
                                 show="headings",
                                 yscrollcommand=vsb.set,
                                 xscrollcommand=hsb.set)
        self._tv.pack(fill="both", expand=True)
        vsb.config(command=self._tv.yview)
        hsb.config(command=self._tv.xview)

        style = ttk.Style()
        style.configure("Diana.Treeview",
                         background=BG_CARD, foreground=TEXT_PRI,
                         fieldbackground=BG_CARD, rowheight=22, font=FONT_SMALL)
        style.configure("Diana.Treeview.Heading",
                         background=BG_ENTRY, foreground=ACCENT2,
                         font=("Consolas", 9, "bold"))
        style.map("Diana.Treeview",
                  background=[("selected", ACCENT2)],
                  foreground=[("selected", "#000000")])
        self._tv.config(style="Diana.Treeview")

        for col in DIANA_COLS:
            self._tv.heading(col, text=col,
                              command=lambda c=col: self._sort_col(c))
            self._tv.column(col, width=DIANA_COL_WIDTHS.get(col, 100),
                             anchor="w", stretch=True)
        self._tv.tag_configure("odd",  background=BG_CARD)
        self._tv.tag_configure("even", background=BG_ENTRY)

        self._status = tk.StringVar(value="Chargement…")
        tk.Label(self, textvariable=self._status, bg=BG_CARD,
                 fg=TEXT_SEC, font=FONT_SMALL, anchor="w",
                 padx=12).pack(fill="x", side="bottom")

    def _refresh_headings(self):
        for col in DIANA_COLS:
            real = self._real_cols.get(col, col)
            self._tv.heading(col, text=real)

    def _populate_filters(self):
        if self._df is None:
            return
        for col, cb in self._combo_widgets.items():
            if col in self._df.columns:
                vals = sorted(self._df[col].dropna().unique().tolist())
                cb["values"] = ["(Tous)"] + vals

    def _reset_filters(self):
        for var in self._filter_vars.values():
            var.set("(Tous)")
        self._apply_filters()

    def _apply_filters(self):
        if self._df is None:
            return
        mask = pd.Series([True] * len(self._df), index=self._df.index)
        for col, var in self._filter_vars.items():
            val = var.get()
            if val != "(Tous)" and col in self._df.columns:
                mask &= (self._df[col] == val)
        df_f = self._df[mask]
        self._fill_table(df_f)
        self._count_var.set(f"{len(df_f)} ligne(s)")
        self._status.set(f"Filtre appliqué — {len(df_f)} ligne(s) affichée(s)")

    def _fill_table(self, df):
        self._tv.delete(*self._tv.get_children())
        for i, (_, row) in enumerate(df.iterrows()):
            tag = "even" if i % 2 == 0 else "odd"
            self._tv.insert("", "end",
                             values=[row[c] for c in DIANA_COLS],
                             tags=(tag,))

    def _sort_col(self, col):
        data = [(self._tv.set(k, col), k) for k in self._tv.get_children("")]
        data.sort(reverse=getattr(self._tv, f"_sort_rev_{col}", False))
        for i, (_, k) in enumerate(data):
            self._tv.move(k, "", i)
            self._tv.item(k, tags=("even" if i % 2 == 0 else "odd",))
        setattr(self._tv, f"_sort_rev_{col}",
                not getattr(self._tv, f"_sort_rev_{col}", False))


# ─────────────────────────────────────────────
#  FENÊTRE MISE À JOUR PLAN DE CHARGES
# ─────────────────────────────────────────────
MOIS_ABBREV = ["Jan.", "Fév.", "Mar.", "Avr.", "Mai", "Juin",
               "Juil.", "Août", "Sep.", "Oct.", "Nov.", "Déc."]
COL_LIBELLE = "PRJ_LibelleProjet"
COL_OU      = "Code UT"


class PdcMajWindow(tk.Toplevel):

    def __init__(self, parent, path_jh, path_eur,
                 cache_usernames=None, cache_jh=None,
                 cache_eur=None, cache_eur_detail=None,
                 diana_jh=None, diana_eur=None, histo_tjm=None,
                 cache_ready=False, nucleus_exclusions=None):
        super().__init__(parent)
        self.title("Mise à jour Plan de charges")
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)
        self.minsize(1100, 700)
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w, h = min(1500, sw - 60), min(860, sh - 60)
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        self.path_jh   = path_jh
        self.path_eur  = path_eur
        self._df_jh    = None
        self._df_eur   = None
        self._interv_var  = tk.StringVar()
        self._annee       = datetime.now().year
        self._mois_cur    = datetime.now().month
        self._modifs      = {}
        self._added_rows  = set()
        self._cells       = {}
        self._cache_ready     = cache_ready
        self._cache_usernames = cache_usernames or []
        self._cache_jh        = cache_jh  or {}
        self._cache_eur       = cache_eur or {}
        self._cache_eur_detail = cache_eur_detail or {}
        self._diana_jh           = diana_jh  or {}
        self._diana_eur          = diana_eur or {}
        self._histo_tjm          = histo_tjm or {}
        self._nucleus_exclusions = nucleus_exclusions or set()
        self._nucleus_only_var   = tk.BooleanVar(value=False)
        import os as _os
        self._log_path = _os.path.join(
            _os.path.dirname(_os.path.abspath(path_jh)),
            "Modifs_plan_de_charge.log")

        self._budgets, self._survcomm_projects, self._vacance_projects, self._analytic_projects = self._load_budgets()
        self._survcomm_intervenants = set()
        if self._survcomm_projects:
            for u_orig in (self._cache_usernames or []):
                u_norm = _norm_name(u_orig)
                projs  = set(self._cache_jh.get(u_norm, {}).keys())
                projs_hors_vacance = projs - self._vacance_projects
                if projs_hors_vacance and projs_hors_vacance.issubset(self._survcomm_projects):
                    self._survcomm_intervenants.add(u_orig)
        self._analytic_intervenants = set()
        if self._analytic_projects:
            for u_orig in (self._cache_usernames or []):
                u_norm = _norm_name(u_orig)
                projs  = set(self._cache_jh.get(u_norm, {}).keys())
                projs_hors_vacance = projs - self._vacance_projects
                if projs_hors_vacance and projs_hors_vacance.issubset(self._analytic_projects):
                    self._analytic_intervenants.add(u_orig)
        self._intervenants_orig, self._intervenants_norm = self._load_all_intervenants()
        self._build_ui()
        self.after(100, self._load_data)

    def _load_budgets(self):
        import os, csv
        log = self.master._log

        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidates = [
            os.path.join(os.getcwd(), "Budgets_2026.csv"),
            os.path.join(script_dir,  "Budgets_2026.csv"),
        ]
        path = next((p for p in candidates if os.path.isfile(p)), None)

        result   = {}
        survcomm = set()
        vacance  = set()
        analytic = set()

        if path is None:
            log("  ⚠  Budgets_2026.csv introuvable. Chemins recherchés :", "warn")
            for p in candidates:
                log(f"      • {p}", "warn")
            return result, survcomm, vacance, analytic

        enc = "utf-8-sig"
        for c in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
            try:
                with open(path, "r", encoding=c) as f:
                    f.read(1024)
                enc = c
                break
            except (UnicodeDecodeError, LookupError):
                continue
        try:
            with open(path, "r", encoding=enc, errors="replace", newline="") as f:
                reader = csv.DictReader(f, delimiter=";")
                raw_fields = reader.fieldnames or []
                norm_fields = {
                    fn.strip().lower().replace(" ", "_"): fn
                    for fn in raw_fields
                }
                col_pc  = norm_fields.get("project_code", "Project Code")
                col_bud = norm_fields.get("budget",       "Budget")
                col_top      = norm_fields.get("topsurvcomm",  "TopSurvComm")
                col_vac      = norm_fields.get("topabsence",
                               norm_fields.get("top_absence",
                               norm_fields.get("topvacance",
                               norm_fields.get("top_vacance",  "TopAbsence"))))
                col_analytic = norm_fields.get("topanalytic", "TopAnalytic")

                for row in reader:
                    pc  = str(row.get(col_pc,  "")).strip()
                    raw = str(row.get(col_bud, "")).strip().replace(",", ".")
                    top   = str(row.get(col_top,      "")).strip()
                    vac   = str(row.get(col_vac,      "")).strip().split(".")[0]
                    top_a = str(row.get(col_analytic, "")).strip()
                    if not pc:
                        continue
                    try:
                        result[pc] = float(raw)
                    except ValueError:
                        pass
                    if top:
                        survcomm.add(pc)
                    if vac == "1":
                        vacance.add(pc)
                    if top_a:
                        analytic.add(pc)

            log("━" * 54, "section")
            log(f"  Budgets_2026.csv chargé depuis {path} (encodage : {enc})", "ok")
            log(f"  Structure : {len(raw_fields)} colonne(s) → "
                f"{', '.join(raw_fields)}", "info")
            log(f"  Colonnes résolues : Project Code='{col_pc}' | "
                f"Budget='{col_bud}' | TopSurvComm='{col_top}' | "
                f"TopAbsence='{col_vac}' | TopAnalytic='{col_analytic}'", "info")
            with open(path, "r", encoding=enc, errors="replace", newline="") as f2:
                raw_lines = [l.rstrip("\n\r") for l in f2.readlines()]
            log(f"  Contenu brut ({len(raw_lines)} ligne(s)) :", "info")
            for line in raw_lines:
                log(f"    {line}", "info")
            log(f"  {len(result)} projet(s) avec budget, "
                f"{len(survcomm)} projet(s) SurvComm, "
                f"{len(vacance)} projet(s) TopAbsence, "
                f"{len(analytic)} projet(s) Analytic.", "ok")

            survcomm_users = []
            for u_orig in (self._cache_usernames or []):
                u_norm = _norm_name(u_orig)
                projs  = set(self._cache_jh.get(u_norm, {}).keys())
                projs_hors_vacance = projs - vacance
                if projs_hors_vacance and projs_hors_vacance.issubset(survcomm):
                    survcomm_users.append(u_orig)
            if survcomm_users:
                log(f"  Intervenant(s) exclusivement SurvComm "
                    f"({len(survcomm_users)}) — exclus des écarts JH/JO :", "info")
                for u in survcomm_users:
                    log(f"      • {u}", "info")
            else:
                log("  Aucun intervenant exclusivement SurvComm.", "info")

        except Exception as e:
            log(f"  ✘  Erreur lecture Budgets_2026.csv : {e}", "error")

        return result, survcomm, vacance, analytic

    def _load_all_intervenants(self):
        """Charge intervenants.csv et retourne (set noms originaux, set noms normalisés)
        pour TOUS les intervenants présents dans le fichier (sans filtre de colonne)."""
        import os, csv as _csv
        log = self.master._log
        path = getattr(self.master, "file4_path", None)
        path = path.get() if path else ""
        if not path or not os.path.isfile(path):
            log("  ⚠  intervenants.csv introuvable — restriction P02508 non-intervenants désactivée.", "warn")
            return set(), set()
        orig = set()
        norm = set()
        try:
            with open(path, encoding="utf-8-sig", errors="replace") as f:
                reader = _csv.DictReader(f, delimiter=";")
                for row in reader:
                    name = str(row.get("Intervenant", "")).strip()
                    if name:
                        orig.add(name)
                        norm.add(_norm_name(name))
        except Exception as e:
            log(f"  ✘  Erreur lecture intervenants.csv : {e}", "error")
        return orig, norm

    def _load_data(self):
        if self._cache_ready and self._cache_usernames:
            self._cb_interv["values"] = self._cache_usernames
            if self._cache_usernames:
                self._interv_var.set(self._cache_usernames[0])
                self._refresh_tables()
            self._status.set(
                f"Cache utilisé — {len(self._cache_usernames)} username(s) RGU")
            return

        self._status.set("Chargement des fichiers PDC…")
        self.update_idletasks()
        try:
            self._df_jh  = self._read_pdc_df(self.path_jh)
            self._df_eur = self._read_pdc_df(self.path_eur)

            cols_jh = list(self._df_jh.columns)
            self._status.set(f"Colonnes JH : {cols_jh}")
            self.update_idletasks()

            def _norm_col(s):
                return s.lower().replace(" ", "").replace("_", "")
            col_map_jh = {_norm_col(c): c for c in cols_jh}

            interv_col = col_map_jh.get(_norm_col(COL_INTERVENANT), COL_INTERVENANT)
            mois_col   = col_map_jh.get(_norm_col(COL_MOIS), COL_MOIS)
            yr_col     = next((c for c in cols_jh if "year" in c.lower()), None)
            ou_col     = next((c for c in cols_jh
                               if "organization" in c.lower()
                               or _norm_col(c) in ("codeut", "code ut")
                               or "type" in c.lower() and "intervenant" in c.lower()), None)

            self._interv_col = interv_col
            self._mois_col   = mois_col
            self._yr_col     = yr_col
            self._ou_col     = ou_col

            def _to_mois(v):
                v2 = str(v).strip().lower()
                r  = MOIS_ORDRE.get(v2)
                if r: return r
                try: return int(float(v2))
                except: return None

            self._df_jh["_mois_num"]  = self._df_jh[mois_col].apply(_to_mois)
            self._df_eur["_mois_num"] = self._df_eur[col_map_jh.get(_norm_col(mois_col), mois_col)
                                                      if mois_col in self._df_eur.columns
                                                      else COL_MOIS].apply(_to_mois) \
                if mois_col in self._df_eur.columns else pd.Series([None]*len(self._df_eur))

            RESP_COL = "PRJ_RespNiv1"
            RESP_VAL = "GUITTARD, Raphael"
            resp_col = next(
                (c for c in cols_jh if _norm_col(c) == _norm_col(RESP_COL)), None)
            if resp_col:
                resp_mask = (self._df_jh[resp_col].astype(str).str.strip() == RESP_VAL)
                intervenants = sorted(
                    self._df_jh[resp_mask][interv_col].dropna()
                    .replace("", pd.NA).dropna().unique().tolist())
                msg = (f"Chargé — {len(intervenants)} intervenant(s) "
                       f"({resp_col} = '{RESP_VAL}')")
            else:
                intervenants = sorted(
                    self._df_jh[interv_col].dropna()
                    .replace("", pd.NA).dropna().unique().tolist())
                msg = (f"⚠ Colonne '{RESP_COL}' introuvable — "
                       f"{len(intervenants)} intervenant(s) sans filtre. "
                       f"Colonnes : {cols_jh[:8]}")
            self._cb_interv["values"] = intervenants
            if intervenants:
                self._interv_var.set(intervenants[0])
                self._refresh_tables()
            self._status.set(msg)
        except Exception as e:
            import traceback as _tb
            self._status.set(f"Erreur : {e}")
            messagebox.showerror("Erreur chargement", _tb.format_exc(), parent=self)

    def _read_pdc_df(self, path):
        ext = path.lower().rsplit(".", 1)[-1]
        if ext == "csv":
            enc = "utf-8-sig"
            for c in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
                try:
                    with open(path, "r", encoding=c) as f: f.read(8192)
                    enc = c; break
                except UnicodeDecodeError: continue
            with open(path, "r", encoding=enc, errors="replace") as f:
                sample = f.read(8192)
            sep = ";" if sample.count(";") >= sample.count(",") else ","
            df = pd.read_csv(path, dtype=str, sep=sep, encoding=enc,
                              encoding_errors="replace", on_bad_lines="skip")
        else:
            df = pd.read_excel(path, dtype=str, header=3)
        df.columns = [c.strip() for c in df.columns]
        return df.fillna("")

    def _refresh_tables(self, *_):
        interv = self._interv_var.get()
        if not interv:
            return
        if not self._cache_ready and self._df_jh is None:
            return
        self._modifs.clear()
        self._added_rows.clear()
        self._build_table_jh(interv)
        self._build_table_eur(interv)
        self._build_ecart_list()

    def _get_jh_data(self, interv):
        if self._cache_ready:
            interv_norm = _norm_name(interv)
            mois_cur    = self._mois_cur
            pdc_data    = self._cache_jh.get(interv_norm, {})
            diana_data  = self._diana_jh.get(interv_norm, {})

            result = {}
            all_projs = set(pdc_data.keys()) | set(diana_data.keys())
            for prj in all_projs:
                lib = pdc_data.get(prj, {}).get("_lib", "")
                result[prj] = {"_lib": lib}
                for m in range(1, mois_cur):
                    result[prj][m] = diana_data.get(prj, {}).get(m, 0.0)
                for m in range(mois_cur, 13):
                    result[prj][m] = pdc_data.get(prj, {}).get(m, 0.0)
            return result

        ic = getattr(self, "_interv_col", COL_INTERVENANT)
        yc = getattr(self, "_yr_col",     None)

        def _find_col(df, name):
            if name in df.columns: return name
            k = name.lower().replace(" ","").replace("_","")
            return next((c for c in df.columns
                         if c.lower().replace(" ","").replace("_","") == k), name)

        pc_col  = _find_col(self._df_jh, COL_PROJET)
        jh_col  = _find_col(self._df_jh, COL_NBJOURS)
        lib_col = next((c for c in self._df_jh.columns
                        if "libelle" in c.lower() and "projet" in c.lower()), None)

        mask = self._df_jh[ic].astype(str).str.strip() == interv
        df   = self._df_jh[mask].copy()
        if yc and yc in df.columns:
            df = df[df[yc].astype(str).str.strip() == str(self._annee)]

        result = {}
        for _, row in df.iterrows():
            pc   = str(row.get(pc_col, "")).strip()
            mois = row.get("_mois_num")
            jh   = 0.0
            try: jh = float(str(row.get(jh_col, "0")).replace(",", "."))
            except: pass
            if not pc or mois is None: continue
            if pc not in result:
                lib = str(row.get(lib_col, "")) if lib_col else ""
                result[pc] = {"_lib": lib}
            result[pc][int(mois)] = result[pc].get(int(mois), 0.0) + jh
        return result

    def _get_eur_data(self, interv):
        if self._cache_ready:
            interv_norm = _norm_name(interv)
            projets_interv = set(self._cache_jh.get(interv_norm, {}).keys())
            return {p: v for p, v in self._cache_eur.items()
                    if p in projets_interv}

        def _find_col(df, name):
            if name in df.columns: return name
            k = name.lower().replace(" ","").replace("_","")
            return next((c for c in df.columns
                         if c.lower().replace(" ","").replace("_","") == k), name)

        pc_col  = _find_col(self._df_eur, COL_PROJET)
        eur_col = _find_col(self._df_eur, COL_NBJOURS)
        yc      = getattr(self, "_yr_col", None)

        df = self._df_eur.copy()
        if yc and yc in df.columns:
            df = df[df[yc].astype(str).str.strip() == str(self._annee)]
        result = {}
        for _, row in df.iterrows():
            pc  = str(row.get(pc_col, "")).strip()
            eur = 0.0
            try: eur = float(str(row.get(eur_col, "0")).replace(",", "."))
            except: pass
            if not pc: continue
            result[pc] = result.get(pc, 0.0) + eur
        return result

    def _build_ecart_list(self):
        for w in self._frame_ecart.winfo_children():
            w.destroy()

        JOURS_OUVRES  = _calc_jours_ouvres(self._annee)
        mois_futurs   = list(range(self._mois_cur, 13))

        nucleus_only = self._nucleus_only_var.get()
        ecarts = []
        for u_orig in (self._cache_usernames or []):
            if nucleus_only and u_orig in (self._survcomm_intervenants | self._analytic_intervenants):
                continue
            u_norm   = _norm_name(u_orig)
            pdc_prjs = self._cache_jh.get(u_norm, {})
            ecart_total   = 0.0
            has_ecart_fut = False
            for m in mois_futurs:
                jh_m = sum(
                    prj_data.get(m, 0.0)
                    for prj, prj_data in pdc_prjs.items()
                    if isinstance(prj_data, dict)
                    # Règle 1a : GUITTARD/ABDELLI — exclure leur JH sur projets SurvComm
                    and not (nucleus_only
                             and u_norm in _NUCLEUS_HARDCODED_SURVCOMM_NORM
                             and prj in self._survcomm_projects)
                    # Règle 1b : GUITTARD/TCHIEGAING — exclure leur JH sur projets Analytic
                    and not (nucleus_only
                             and u_norm in _NUCLEUS_HARDCODED_ANALYTIC_NORM
                             and prj in self._analytic_projects)
                )
                e = jh_m - JOURS_OUVRES[m - 1]
                ecart_total += e
                if abs(e) > 0.01:
                    has_ecart_fut = True
            if has_ecart_fut:
                ecarts.append((u_orig, ecart_total))

        C_HDR    = "#1e3a5f"
        W_INTERV = 31
        W_ECART  = 10

        tk.Label(self._frame_ecart, text="Intervenant", bg=C_HDR, fg=ACCENT2,
                 font=("Consolas", 9, "bold"), width=W_INTERV,
                 anchor="w", relief="flat", bd=1).grid(
            row=0, column=0, sticky="ew", padx=1, pady=1)
        tk.Label(self._frame_ecart, text="Écart JH", bg=C_HDR, fg=ACCENT2,
                 font=("Consolas", 9, "bold"), width=W_ECART,
                 anchor="e", relief="flat", bd=1).grid(
            row=0, column=1, sticky="ew", padx=1, pady=1)

        self._ecart_usernames = []
        for i, (u_orig, ecart) in enumerate(ecarts, start=1):
            bg_r  = "#1a2530" if i % 2 == 0 else BG_CARD
            sign  = "+" if ecart > 0 else ""
            e_str = f"{sign}{round(ecart)}"
            fg_e  = WARN if ecart > 0.01 else (ACCENT if ecart < -0.01 else TEXT_PRI)

            lbl_n = tk.Label(self._frame_ecart, text=u_orig[:W_INTERV],
                             bg=bg_r, fg=TEXT_PRI, font=FONT_SMALL,
                             width=W_INTERV, anchor="w")
            lbl_n.grid(row=i, column=0, sticky="ew", padx=1, pady=1)
            lbl_e = tk.Label(self._frame_ecart, text=e_str,
                             bg=bg_r, fg=fg_e, font=FONT_SMALL,
                             width=W_ECART, anchor="e")
            lbl_e.grid(row=i, column=1, sticky="ew", padx=1, pady=1)

            self._ecart_usernames.append(u_orig)

            def _on_click(event, u=u_orig):
                if u == self._interv_var.get():
                    return
                self._interv_var.set(u)
                self._refresh_tables()

            lbl_n.bind("<Button-1>", _on_click)
            lbl_e.bind("<Button-1>", _on_click)

    def _sync_ecart_height(self):
        pass

    def _build_table_jh(self, interv):
        for w in self._frame_jh.winfo_children():
            w.destroy()
        self._cells.clear()

        data  = self._get_jh_data(interv)
        now_m = self._mois_cur

        C_HDR    = "#1e3a5f"
        C_PAST   = "#2a3a2a"
        C_PAST_H = "#d4e8c2"
        C_OVER   = "#f0883e"

        JOURS_OUVRES = _calc_jours_ouvres(self._annee)

        tk.Label(self._frame_jh, text="Project Code", bg=C_HDR, fg=ACCENT2,
                 font=("Consolas", 9, "bold"), width=12, anchor="w",
                 relief="flat", bd=1).grid(row=0, column=0, sticky="ew", padx=1, pady=1)
        tk.Label(self._frame_jh, text="Project Name", bg=C_HDR, fg=ACCENT2,
                 font=("Consolas", 9, "bold"), width=34, anchor="w",
                 relief="flat", bd=1).grid(row=0, column=1, sticky="ew", padx=1, pady=1)
        for m in range(1, 13):
            bg_h = C_PAST if m < now_m else C_HDR
            fg_h = C_PAST_H if m < now_m else ACCENT2
            tk.Label(self._frame_jh, text=MOIS_ABBREV[m-1], bg=bg_h, fg=fg_h,
                     font=("Consolas", 9, "bold"), width=6, anchor="center",
                     relief="flat", bd=1).grid(row=0, column=m+1, sticky="ew", padx=1, pady=1)
        tk.Label(self._frame_jh, text="Total", bg=C_HDR, fg=ACCENT2,
                 font=("Consolas", 9, "bold"), width=7, anchor="center",
                 relief="flat", bd=1).grid(row=0, column=14, sticky="ew", padx=1, pady=1)

        projs = sorted(data.keys())
        self._proj_rows = projs[:]

        lib_map = {}
        if self._cache_ready:
            interv_norm = _norm_name(interv)
            for pc2, prj_dict in self._cache_jh.get(interv_norm, {}).items():
                lib_map[pc2] = prj_dict.get("_lib", prj_dict.get("_libelle", ""))

        for ri, pc in enumerate(projs, start=1):
            info = data[pc]
            lib  = lib_map.get(pc, info.get("_lib", info.get("_libelle", "")))
            if not lib:
                if self._cache_ready:
                    for u_norm, pd2 in self._cache_jh.items():
                        if pc in pd2 and pd2[pc].get("_lib"):
                            lib = pd2[pc]["_lib"]; break
            bg_r = "#1a2530" if ri % 2 == 0 else BG_CARD
            tk.Label(self._frame_jh, text=pc, bg=bg_r, fg=TEXT_PRI,
                     font=FONT_SMALL, width=12, anchor="w").grid(
                row=ri, column=0, sticky="ew", padx=1, pady=1)
            tk.Label(self._frame_jh, text=lib[:40], bg=bg_r, fg=TEXT_SEC,
                     font=FONT_SMALL, width=34, anchor="w").grid(
                row=ri, column=1, sticky="ew", padx=1, pady=1)
            row_total = 0.0
            for m in range(1, 13):
                val = info.get(m, 0.0)
                row_total += val
                editable = (m < now_m)
                bg_c = C_PAST if m < now_m else bg_r
                fg_c = C_PAST_H if m < now_m else TEXT_PRI
                disp = str(int(val)) if val == int(val) else f"{val:.1f}"
                if editable:
                    var = tk.StringVar(value=disp)
                    e = tk.Entry(self._frame_jh, textvariable=var,
                                 bg=bg_c, fg=fg_c, font=FONT_SMALL,
                                 width=6, justify="center", bd=0,
                                 insertbackground=ACCENT,
                                 relief="flat",
                                 highlightbackground=BORDER,
                                 highlightthickness=1)
                    e.grid(row=ri, column=m+1, sticky="ew", padx=1, pady=1)
                    e.bind("<FocusOut>",
                           lambda ev, p=pc, mo=m, v=var, orig=val:
                           self._on_cell_edit(p, mo, v, orig))
                    self._cells[(pc, m)] = (var, val)
                else:
                    tk.Label(self._frame_jh, text=disp if val else "",
                             bg=bg_c, fg=fg_c, font=FONT_SMALL,
                             width=6, anchor="center").grid(
                        row=ri, column=m+1, sticky="ew", padx=1, pady=1)
            tot_disp = str(int(row_total)) if row_total == int(row_total) else f"{row_total:.1f}"
            tk.Label(self._frame_jh, text=tot_disp, bg=bg_r, fg=ACCENT2,
                     font=("Consolas", 9, "bold"), width=7, anchor="center").grid(
                row=ri, column=14, sticky="ew", padx=1, pady=1)

        n_rows  = len(projs)
        tot_row = n_rows + 1
        tk.Label(self._frame_jh, text="Total", bg=C_HDR, fg=ACCENT2,
                 font=("Consolas", 9, "bold"), width=12, anchor="e").grid(
            row=tot_row, column=0, columnspan=2, sticky="ew", padx=1, pady=1)
        grand_total = 0.0
        for m in range(1, 13):
            col_tot = sum(data[pc].get(m, 0.0) for pc in projs)
            grand_total += col_tot
            jo   = JOURS_OUVRES[m-1]
            over = abs(col_tot - jo) > 0.01 and m >= now_m
            bg_t = C_PAST if m < now_m else (C_OVER if over else C_HDR)
            fg_t = "#1a1a1a" if over else (C_PAST_H if m < now_m else ACCENT2)
            disp = str(int(col_tot)) if col_tot == int(col_tot) else f"{col_tot:.1f}"
            tk.Label(self._frame_jh, text=disp if col_tot else "",
                     bg=bg_t, fg=fg_t,
                     font=("Consolas", 9, "bold"), width=6, anchor="center").grid(
                row=tot_row, column=m+1, sticky="ew", padx=1, pady=1)
        gt = str(int(grand_total)) if grand_total == int(grand_total) else f"{grand_total:.1f}"
        tk.Label(self._frame_jh, text=gt, bg=C_HDR, fg=ACCENT2,
                 font=("Consolas", 9, "bold"), width=7, anchor="center").grid(
            row=tot_row, column=14, sticky="ew", padx=1, pady=1)

        jo_row = tot_row + 1
        tk.Label(self._frame_jh, text="Jours ouvrés", bg=BG_PANEL, fg=TEXT_SEC,
                 font=FONT_SMALL, width=12, anchor="e").grid(
            row=jo_row, column=0, columnspan=2, sticky="ew", padx=1, pady=1)
        jo_total = 0
        for m in range(1, 13):
            jo = JOURS_OUVRES[m-1]
            jo_total += jo
            bg_j = C_PAST if m < now_m else BG_PANEL
            tk.Label(self._frame_jh, text=str(jo), bg=bg_j, fg=TEXT_SEC,
                     font=FONT_SMALL, width=6, anchor="center").grid(
                row=jo_row, column=m+1, sticky="ew", padx=1, pady=1)
        tk.Label(self._frame_jh, text=str(jo_total), bg=BG_PANEL, fg=TEXT_SEC,
                 font=FONT_SMALL, width=7, anchor="center").grid(
            row=jo_row, column=14, sticky="ew", padx=1, pady=1)

    def _build_table_eur(self, interv):
        for w in self._frame_eur.winfo_children():
            w.destroy()
        for w in self._frame_detail.winfo_children():
            w.destroy()
        self._selected_proj = None

        mois_passes = list(range(1, self._mois_cur))

        all_projs = set()
        if self._cache_ready:
            for prj_dict in self._diana_eur.values():
                all_projs.update(prj_dict.keys())
            for (u_norm, prj, m) in self._cache_eur_detail:
                all_projs.add(prj)
        projs = sorted(
            (p for p in all_projs if p not in EXCLUDED_PROJECTS),
            key=lambda p: self._budgets.get(p) or 0,
            reverse=True
        )

        lib_lookup = {}
        if self._cache_ready:
            for prj_dict in self._cache_jh.values():
                for prj, prj_data in prj_dict.items():
                    if prj not in lib_lookup and isinstance(prj_data, dict):
                        lib_lookup[prj] = prj_data.get("_lib", "")

        C_HDR = "#1e3a5f"
        self._eur_row_widgets = {}

        cols_hdr = [
            ("Project Code",    12, "w"),
            ("Project Name",    34, "w"),
            ("Consommé à date", 16, "e"),
            ("RAF",             16, "e"),
            ("Atterrissage",    16, "e"),
            ("Budget",          14, "e"),
        ]
        for ci, (txt, w, anc) in enumerate(cols_hdr):
            tk.Label(self._frame_eur, text=txt, bg=C_HDR, fg=ACCENT2,
                     font=("Consolas", 9, "bold"), width=w, anchor=anc,
                     relief="flat", bd=1).grid(
                row=0, column=ci, sticky="ew", padx=1, pady=1)

        def _fmt(v):
            return f"{round(v):,} €".replace(",", "\u202f") if v else ""

        mois_futurs = list(range(self._mois_cur, 13))

        for ri, pc in enumerate(projs, start=1):
            lib = lib_lookup.get(pc, "")

            nucleus_only = self._nucleus_only_var.get()

            conso_date = 0.0
            if self._cache_ready and self._diana_eur:
                for u_norm, prj_dict in self._diana_eur.items():
                    if nucleus_only and u_norm in self._nucleus_exclusions:
                        continue
                    # Règle 1a : GUITTARD/ABDELLI — exclure leurs consommés SurvComm
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_SURVCOMM_NORM and pc in self._survcomm_projects:
                        continue
                    # Règle 1b : GUITTARD/TCHIEGAING — exclure leurs consommés Analytic
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_ANALYTIC_NORM and pc in self._analytic_projects:
                        continue
                    for m, eur_v in prj_dict.get(pc, {}).items():
                        if m in mois_passes:
                            conso_date += eur_v
            elif self._cache_ready and self._cache_eur_detail:
                for (u_norm, prj, m), eur_val in self._cache_eur_detail.items():
                    if nucleus_only and u_norm in self._nucleus_exclusions:
                        continue
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_SURVCOMM_NORM and prj in self._survcomm_projects:
                        continue
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_ANALYTIC_NORM and prj in self._analytic_projects:
                        continue
                    if prj == pc and m in mois_passes:
                        conso_date += eur_val

            raf = 0.0
            if self._cache_ready and self._cache_eur_detail:
                for (u_norm, prj, m), eur_val in self._cache_eur_detail.items():
                    if nucleus_only and u_norm in self._nucleus_exclusions:
                        continue
                    # Règle 1a : GUITTARD/ABDELLI — exclure leur RAF sur projets SurvComm
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_SURVCOMM_NORM and prj in self._survcomm_projects:
                        continue
                    # Règle 1b : GUITTARD/TCHIEGAING — exclure leur RAF sur projets Analytic
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_ANALYTIC_NORM and prj in self._analytic_projects:
                        continue
                    # Règle 2 : non-intervenants — exclure leur RAF hors P02508
                    if nucleus_only and prj != _NUCLEUS_P02508_ONLY and u_norm not in self._intervenants_norm:
                        continue
                    if prj == pc and m in mois_futurs:
                        raf += eur_val

            atterr = round(conso_date + raf, 2)

            bg_r = "#1a2530" if ri % 2 == 0 else BG_CARD
            row_data = [
                (pc,              12, "w", TEXT_PRI),
                (lib[:40],        34, "w", TEXT_SEC),
                (_fmt(conso_date),16, "e", TEXT_PRI),
                (_fmt(raf),       16, "e", TEXT_PRI),
                (_fmt(atterr),    16, "e", ACCENT2),
                (_fmt(self._budgets.get(pc)), 14, "e", TEXT_SEC),
            ]
            row_lbls = []
            for ci, (txt, w, anc, fg) in enumerate(row_data):
                lbl = tk.Label(self._frame_eur, text=txt, bg=bg_r,
                               fg=fg, font=FONT_SMALL, width=w, anchor=anc,
                               cursor="hand2")
                lbl.grid(row=ri, column=ci, sticky="ew", padx=1, pady=1)
                lbl.bind("<Button-1>",
                         lambda e, p=pc, bg=bg_r: self._on_proj_click(p, bg))
                row_lbls.append(lbl)
            self._eur_row_widgets[pc] = (row_lbls, bg_r)

    def _on_proj_click(self, pc, bg_orig):
        if self._selected_proj and self._selected_proj in self._eur_row_widgets:
            lbls, bg_r = self._eur_row_widgets[self._selected_proj]
            for l in lbls: l.config(bg=bg_r)
        if pc in self._eur_row_widgets:
            lbls, _ = self._eur_row_widgets[pc]
            for l in lbls: l.config(bg="#2a3a5f")
        self._selected_proj = pc
        self._show_proj_detail(pc)

    def _compute_nucleus_exclusions(self):
        """Retourne les noms normalisés des intervenants à exclure du filtre Nucleus.
        Intervenants exclusifs SurvComm ET intervenants exclusifs Analytic.
        """
        return {_norm_name(u) for u in self._survcomm_intervenants | self._analytic_intervenants}

    def _on_nucleus_filter(self):
        """Rappelé au clic sur la checkbox — calcule les exclusions, log, rafraîchit le détail."""
        log = getattr(self.master, "_log", None)
        checked = self._nucleus_only_var.get()

        if log:
            log("─" * 50, "section")
            log(f"  Périmètre Nucleus uniquement : {'✔ activé' if checked else '✘ désactivé'}", "info")

        if checked:
            # Calcul dynamique depuis le cache JH
            self._nucleus_exclusions = self._compute_nucleus_exclusions()
            if self._nucleus_exclusions:
                # Récupérer le nom original pour l'affichage
                norm_to_orig = {_norm_name(u): u for u in (self._cache_usernames or [])}
                if log:
                    log(f"  Intervenants exclus (périmètre SurvComm/Analytic exclusif — {len(self._nucleus_exclusions)}) :", "info")
                    for norm in sorted(self._nucleus_exclusions):
                        display = norm_to_orig.get(norm, norm)
                        log(f"    • {display}", "info")
            else:
                if log:
                    log("  ⚠  Aucun intervenant à exclure — cache JH vide ou tous multi-périmètres.", "warn")

            # Exclusions en dur : GUITTARD/ABDELLI — projets SurvComm
            if log:
                log(f"  Exclusions partielles (multiprojets — projets SurvComm) :", "info")
                for name in _NUCLEUS_HARDCODED_SURVCOMM_ORIG:
                    log(f"    • {name} : consommés JH et EUR sur projets SurvComm exclus", "info")
            # Exclusions en dur : GUITTARD/TCHIEGAING — projets Analytic
            if log:
                log(f"  Exclusions partielles (multiprojets — projets Analytic) :", "info")
                for name in _NUCLEUS_HARDCODED_ANALYTIC_ORIG:
                    log(f"    • {name} : consommés JH et EUR sur projets Analytic exclus", "info")

            # Exclusions en dur : non-intervenants — RAF hors P02508
            if log:
                non_interv = [
                    u for u in (self._cache_usernames or [])
                    if _norm_name(u) not in self._intervenants_norm
                    and _norm_name(u) not in self._nucleus_exclusions
                    and _norm_name(u) not in _NUCLEUS_HARDCODED_NORM
                ]
                if non_interv:
                    log(f"  Intervenants hors intervenants.csv — RAF limité à {_NUCLEUS_P02508_ONLY} ({len(non_interv)}) :", "info")
                    for u in non_interv:
                        log(f"    • {u}", "info")
                else:
                    log(f"  Aucun intervenant hors intervenants.csv dans le cache.", "info")
        else:
            self._nucleus_exclusions = set()

        # Rafraîchir les tableaux
        interv = self._interv_var.get()
        if interv:
            if log:
                log(f"  Rafraîchissement des tableaux pour : {interv}", "info")
            prev_proj = self._selected_proj
            self._build_ecart_list()
            self._build_table_eur(interv)
            # Ré-afficher le détail si un projet était sélectionné
            if prev_proj:
                self._show_proj_detail(prev_proj)
        else:
            if log:
                log("  ℹ  Aucun intervenant sélectionné.", "info")

    def _show_proj_detail(self, pc):
        for w in self._frame_detail.winfo_children():
            w.destroy()

        C_HDR = "#1e3a5f"
        mois_passes_d = list(range(1, self._mois_cur))
        mois_futurs_d = list(range(self._mois_cur, 13))

        def _fmt_d(v):
            return f"{round(v):,} €".replace(",", "\u202f") if v else ""

        for ci, (txt, w) in enumerate([
            ("Intervenant",     29), ("Consommé à date", 16),
            ("RAF",             14), ("Atterrissage",    16)
        ]):
            tk.Label(self._frame_detail, text=txt, bg=C_HDR, fg=ACCENT2,
                     font=("Consolas", 9, "bold"), width=w,
                     anchor="w" if ci == 0 else "e",
                     relief="flat", bd=1).grid(
                row=0, column=ci, sticky="ew", padx=1, pady=1)

        nucleus_only = self._nucleus_only_var.get()
        detail = {}
        if self._cache_ready:
            for u_norm, prj_dict in self._cache_jh.items():
                if pc in prj_dict:
                    # Filtre Nucleus : exclure les intervenants uniquement Surv.Comm. ou Analytic
                    if nucleus_only and u_norm in self._nucleus_exclusions:
                        continue
                    # Règle 1a : GUITTARD/ABDELLI — exclure entièrement sur projets SurvComm
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_SURVCOMM_NORM and pc in self._survcomm_projects:
                        continue
                    # Règle 1b : GUITTARD/TCHIEGAING — exclure entièrement sur projets Analytic
                    if nucleus_only and u_norm in _NUCLEUS_HARDCODED_ANALYTIC_NORM and pc in self._analytic_projects:
                        continue
                    u_orig = next(
                        (v for v in (self._cache_usernames or [])
                         if _norm_name(v) == u_norm), u_norm)
                    c_date = sum(
                        self._diana_eur.get(u_norm, {}).get(pc, {}).get(m, 0.0)
                        for m in mois_passes_d)
                    raf_v = sum(
                        v for (un, pr, m), v in self._cache_eur_detail.items()
                        if un == u_norm and pr == pc and m in mois_futurs_d)
                    # Règle 2 : non-intervenants — exclure leur RAF hors P02508
                    if nucleus_only and pc != _NUCLEUS_P02508_ONLY and u_norm not in self._intervenants_norm:
                        raf_v = 0.0
                    if c_date or raf_v:
                        detail[u_orig] = (c_date, raf_v)

        tot_cd = tot_raf = 0.0
        for ri, (u_orig, (cd, raf_v)) in enumerate(
                sorted(detail.items(), key=lambda x: -(x[1][0]+x[1][1])), start=1):
            bg_r = "#1a2530" if ri % 2 == 0 else BG_CARD
            atterr_v = round(cd + raf_v, 2)
            for ci, (val, w, fg) in enumerate([
                (u_orig[:29],    29, TEXT_PRI),
                (_fmt_d(cd),     16, TEXT_PRI),
                (_fmt_d(raf_v),  14, TEXT_PRI),
                (_fmt_d(atterr_v),16, ACCENT2),
            ]):
                tk.Label(self._frame_detail, text=val, bg=bg_r,
                         fg=fg, font=FONT_SMALL, width=w,
                         anchor="w" if ci == 0 else "e").grid(
                    row=ri, column=ci, sticky="ew", padx=1, pady=1)
            tot_cd  += cd
            tot_raf += raf_v

        tot_row    = 1 + len(detail)
        tot_atterr = round(tot_cd + tot_raf, 2)
        for ci, (val, w) in enumerate([
            ("Total",             29), (_fmt_d(tot_cd),     16),
            (_fmt_d(tot_raf),     14), (_fmt_d(tot_atterr), 16)
        ]):
            tk.Label(self._frame_detail, text=val, bg=C_HDR, fg=ACCENT2,
                     font=("Consolas", 9, "bold"), width=w,
                     anchor="e").grid(
                row=tot_row, column=ci, sticky="ew", padx=1, pady=1)

    def _on_cell_edit(self, proj_code, mois, var, orig_val):
        raw = var.get().strip().replace(",", ".")
        try:
            new_val = float(raw) if raw else 0.0
        except ValueError:
            var.set(str(int(orig_val)) if orig_val == int(orig_val) else f"{orig_val:.1f}")
            return
        if new_val == orig_val:
            return
        self._modifs[(proj_code, mois)] = (orig_val, new_val)
        self._write_log(proj_code, mois, orig_val, new_val)
        self._status.set(
            f"Modif. enregistrée : {proj_code} / mois {mois:02d} : "
            f"{orig_val} → {new_val} JH")

    def _write_log(self, proj_code, mois, old_val, new_val):
        interv = self._interv_var.get()
        ts     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line   = (f"{ts} | {interv} | {proj_code} | mois {mois:02d} | "
                  f"{old_val:.2f} JH → {new_val:.2f} JH\n")
        try:
            with open(self._log_path, "a", encoding="utf-8") as f:
                f.write(line)
        except Exception as e:
            self._status.set(f"⚠  Impossible d'écrire le log : {e}")

    def _build_ui(self):
        hdr = tk.Frame(self, bg=BG_MAIN, pady=8)
        hdr.pack(fill="x", padx=16)
        tk.Label(hdr, text="📝  MISE À JOUR PLAN DE CHARGES",
                 bg=BG_MAIN, fg=ACCENT2, font=FONT_HEAD).pack(side="left")
        tk.Label(hdr, text=f"Année {self._annee}",
                 bg=BG_MAIN, fg=TEXT_SEC, font=FONT_SMALL).pack(side="right")
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        sel_frame = tk.Frame(self, bg=BG_PANEL, pady=6)
        sel_frame.pack(fill="x", padx=10, pady=(6, 2))
        tk.Label(sel_frame, text="UserName :", bg=BG_PANEL,
                 fg=TEXT_SEC, font=FONT_SMALL).pack(side="left", padx=(10, 4))
        self._cb_interv = ttk.Combobox(
            sel_frame, textvariable=self._interv_var,
            state="readonly", width=35, font=FONT_SMALL)
        self._cb_interv.pack(side="left")
        self._cb_interv.bind("<<ComboboxSelected>>", self._refresh_tables)
        make_button(sel_frame, "↺  Rafraîchir", self._refresh_tables,
                    width=14).pack(side="left", padx=12)
        tk.Checkbutton(
            sel_frame, text="Périmètre Nucleus uniquement",
            variable=self._nucleus_only_var,
            bg=BG_PANEL, fg=TEXT_SEC, font=FONT_SMALL,
            selectcolor=BG_ENTRY, activebackground=BG_PANEL,
            activeforeground=TEXT_PRI,
            command=self._on_nucleus_filter,
        ).pack(side="right", padx=(0, 12))

        self._status = tk.StringVar(value="Chargement…")
        tk.Label(self, textvariable=self._status, bg=BG_CARD,
                 fg=TEXT_SEC, font=FONT_SMALL, anchor="w",
                 padx=12).pack(fill="x", side="bottom")

        self._selected_proj = None

        body = tk.Frame(self, bg=BG_MAIN)
        body.pack(fill="both", expand=True, padx=4, pady=4)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=2)
        body.rowconfigure(1, weight=3)

        def _make_quad(parent, title, col, autowidth=False, vscroll=False):
            outer = tk.Frame(parent, bg=BG_PANEL,
                             highlightbackground=BORDER, highlightthickness=1)
            outer.grid(row=0, column=col, sticky="nsew", padx=4, pady=4)
            tk.Label(outer, text=title, bg=BG_PANEL, fg=ACCENT2,
                     font=FONT_BODY, anchor="w").pack(fill="x", padx=6, pady=(4, 2))
            tk.Frame(outer, bg=BORDER, height=1).pack(fill="x")
            cf = tk.Frame(outer, bg=BG_PANEL)
            cf.pack(fill="both", expand=True)
            if vscroll or not autowidth:
                vsb = tk.Scrollbar(cf, orient="vertical")
                vsb.pack(side="right", fill="y")
            c = tk.Canvas(cf, bg=BG_MAIN, bd=0, highlightthickness=0,
                          yscrollcommand=vsb.set if (vscroll or not autowidth) else None)
            if vscroll or not autowidth:
                vsb.config(command=c.yview)
            c.pack(side="left", fill="both", expand=True)
            inner = tk.Frame(c, bg=BG_MAIN)
            cw = c.create_window((0, 0), window=inner, anchor="nw")
            if autowidth:
                inner.bind("<Configure>",
                           lambda e, cv=c, w=cw:
                               cv.config(width=e.width, scrollregion=cv.bbox("all")))
            else:
                inner.bind("<Configure>",
                           lambda e, cv=c: cv.configure(scrollregion=cv.bbox("all")))
                c.bind("<Configure>",
                       lambda e, cv=c, w=cw: cv.itemconfig(w, width=e.width))
            if vscroll or not autowidth:
                c.bind("<Enter>",
                       lambda e, cv=c: cv.bind_all(
                           "<MouseWheel>",
                           lambda ev, cv=cv: cv.yview_scroll(
                               -1 if ev.delta > 0 else 1, "units")))
                c.bind("<Leave>",
                       lambda e, cv=c: cv.unbind_all("<MouseWheel>"))
            return inner

        row1 = tk.Frame(body, bg=BG_MAIN)
        row1.grid(row=0, column=0, sticky="nsew")
        row1.rowconfigure(0, weight=1)
        row1.columnconfigure(0, weight=0)
        row1.columnconfigure(1, weight=1)

        self._frame_jh    = _make_quad(row1, "Plan de charges en JH  (mois passés modifiables)", 0, autowidth=True)
        self._frame_ecart = _make_quad(row1, "Écarts JH / JO", 1)

        row2 = tk.Frame(body, bg=BG_MAIN)
        row2.grid(row=1, column=0, sticky="nsew")
        row2.rowconfigure(0, weight=1)
        row2.columnconfigure(0, weight=0)
        row2.columnconfigure(1, weight=1)

        # Cadran bas-gauche : Plan de charges TTNR
        self._frame_eur = _make_quad(row2, "Plan de charges TTNR", 0, autowidth=True, vscroll=True)

        # Cadran bas-droite : Détail intervenants (simple)
        inner_det = _make_quad(row2, "Détail intervenants", 1)
        self._frame_detail_container = inner_det
        self._frame_detail = tk.Frame(inner_det, bg=BG_MAIN)
        self._frame_detail.pack(anchor="nw", padx=4, pady=4)


# ─────────────────────────────────────────────
#  POINT D'ENTRÉE
# ─────────────────────────────────────────────
if __name__ == "__main__":
    app = SuiviConsoApp()
    app.mainloop()
